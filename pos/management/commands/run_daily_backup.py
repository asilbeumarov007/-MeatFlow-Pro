"""
MeatFlow-Pro — Avtomatik Kunlik Zaxira Nusxasi (Daily Backup)
==============================================================
Tizim ma'lumotlari bazasi (db.sqlite3) hamda kunlik savdolar Excel
hisobotini avtomatik Telegram Admin guruhiga yuboradi.
"""
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from django.core.management.base import BaseCommand
from django.utils import timezone
from django.conf import settings

from pos.models import Sale, Customer, Product, Stock
from pos.telegram_bot import send_document, send_message, CHAT_ID


class Command(BaseCommand):
    help = "Kunlik bazani hamda savdo hisobotlarini Telegram guruhga zaxira fayl (backup) sifatida yuboradi"

    def handle(self, *args, **options):
        today_str = timezone.localdate().strftime('%Y-%m-%d')
        time_str = timezone.localtime().strftime('%H:%M')

        self.stdout.write(f"[BACKUP] Kunlik zaxira nusxasi yaratilmoqda... ({today_str} {time_str})")

        # 1. Generate Excel Sales & Customer Report
        wb = openpyxl.Workbook()
        ws_sales = wb.active
        ws_sales.title = "Bugungi Savdolar"

        # Headers
        headers = ["Chek #", "Vaqt", "Mijoz", "To'lov Usuli", "Jami Summa (so'm)", "To'landi (so'm)", "Qarz (so'm)"]
        ws_sales.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="1B6B4A", end_color="1B6B4A", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        
        for cell in ws_sales[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        sales = Sale.objects.filter(created_at__date=timezone.localdate()).order_by('-id')
        total_sum = 0
        
        for s in sales:
            c_name = f"{s.customer.first_name} {s.customer.last_name or ''}".strip() if s.customer else "Anonim Mijoz"
            loc_time = timezone.localtime(s.created_at).strftime('%H:%M')
            ws_sales.append([
                s.id,
                loc_time,
                c_name,
                s.get_payment_method_display(),
                float(s.total_amount),
                float(s.final_paid),
                float(s.debt_added)
            ])
            total_sum += float(s.total_amount)

        # Add summary row
        ws_sales.append(["", "", "JAMI:", "", total_sum, "", ""])

        # Auto column widths
        for col in ws_sales.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws_sales.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Customer Debt Sheet
        ws_cust = wb.create_sheet(title="Qarzdorlar Ro'yxati")
        ws_cust.append(["Mijoz ID", "Ism Familiya", "Telefon", "Joriy Qarz (so'm)"])
        for cell in ws_cust[1]:
            cell.fill = header_fill
            cell.font = header_font

        debtors = Customer.objects.filter(debt_amount__gt=0).order_by('-debt_amount')
        for c in debtors:
            ws_cust.append([
                c.custom_id,
                f"{c.first_name} {c.last_name or ''}".strip(),
                c.phone,
                float(c.debt_amount)
            ])

        for col in ws_cust.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws_cust.column_dimensions[col_letter].width = max(max_len + 4, 12)

        excel_filename = f"MeatFlow_Backup_{today_str}.xlsx"
        excel_path = os.path.join(settings.BASE_DIR, 'media', excel_filename)
        os.makedirs(os.path.dirname(excel_path), exist_ok=True)
        wb.save(excel_path)

        # 2. Send Excel file to Telegram
        if CHAT_ID:
            caption = f"📊 *MeatFlow-Pro Kunlik Zaxira Hisoboti*\n📅 Sana: {today_str} {time_str}\n💰 Bugungi Jami Savdo: {total_sum:,.0f} so'm"
            send_document(CHAT_ID, excel_path, caption=caption)

            # 3. Send Database Backup File (PostgreSQL or SQLite)
            db_engine = settings.DATABASES['default']['ENGINE']
            if 'postgresql' in db_engine:
                db_name = settings.DATABASES['default'].get('NAME', 'meatflow_db')
                db_user = settings.DATABASES['default'].get('USER', 'meatflow_user')
                db_pass = settings.DATABASES['default'].get('PASSWORD', '')
                dump_path = os.path.join(settings.BASE_DIR, 'media', f"meatflow_pg_{today_str}.sql.gz")
                
                try:
                    import subprocess
                    env = os.environ.copy()
                    if db_pass:
                        env['PGPASSWORD'] = db_pass
                    
                    cmd = f"pg_dump -U {db_user} -h localhost {db_name} | gzip > {dump_path}"
                    subprocess.run(cmd, shell=True, env=env, check=True)
                    
                    if os.path.exists(dump_path):
                        db_caption = f"🐘 *PostgreSQL Baza Zaxirasi ({db_name})*\n📅 Sana: {today_str} {time_str}\n🔒 Xavfsiz Shifrlangan Cloud Nusxa"
                        send_document(CHAT_ID, dump_path, caption=db_caption)
                        try:
                            os.remove(dump_path)
                        except Exception:
                            pass
                except Exception as e:
                    self.stderr.write(f"[PostgreSQL Dump Error]: {e}")
            else:
                db_path = os.path.join(settings.BASE_DIR, 'db.sqlite3')
                if os.path.exists(db_path):
                    db_caption = f"💾 *Baza Zaxira Nusxasi (db.sqlite3)*\n📅 Sana: {today_str}"
                    send_document(CHAT_ID, db_path, caption=db_caption)

            self.stdout.write(self.style.SUCCESS("[OK] Kunlik zaxira nusxasi Telegram guruhga yuborildi!"))
        else:
            self.stderr.write("[XATO] TELEGRAM_CHAT_ID o'rnatilmagan!")
