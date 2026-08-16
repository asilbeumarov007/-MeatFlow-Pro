from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import user_passes_test, login_required
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q, Sum, F
from django.db import transaction
from .permissions import staff_required, admin_required, is_staff_or_admin, is_admin
from django.utils import timezone
from django.conf import settings
from decimal import Decimal

import json
import random
from datetime import timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def style_excel_worksheet(ws, headers_count):
    # Enable grid lines explicitly
    if ws.views.sheetView:
        ws.views.sheetView[0].showGridLines = True
    
    # Header styles
    header_fill = PatternFill(start_color='1B6B4A', end_color='1B6B4A', fill_type='solid')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    row_alt_fill = PatternFill(start_color='F3F8F5', end_color='F3F8F5', fill_type='solid')
    
    ws.row_dimensions[1].height = 28
    for col_idx in range(1, headers_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
        
    data_font = Font(name='Segoe UI', size=10)
    max_lens = [len(str(ws.cell(row=1, column=c).value or '')) for c in range(1, headers_count + 1)]
    
    for r_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[r_idx].height = 22
        is_alt = (r_idx % 2 == 0)
        for c_idx in range(1, headers_count + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = data_font
            cell.border = thin_border
            if is_alt:
                cell.fill = row_alt_fill
                
            val = str(cell.value or '')
            if len(val) > max_lens[c_idx - 1]:
                max_lens[c_idx - 1] = len(val)
                
            if cell.data_type == 'n':
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif len(val) <= 12 and (val.replace('.', '').isdigit() or '/' in val or ':' in val or '-' in val):
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')

    for c_idx, m_len in enumerate(max_lens, 1):
        col_letter = get_column_letter(c_idx)
        ws.column_dimensions[col_letter].width = max(m_len + 3, 11)

from .models import Supplier, Product, Stock, Slaughter, Customer, Sale, SaleItem, CustomerLog, Notebook
from .translit import cyrillic_to_latin, latin_to_cyrillic
from django.contrib.auth import get_user_model

def create_user_for_customer(customer):
    User = get_user_model()
    link_identifier = customer.phone if customer.phone else customer.custom_id
    if not link_identifier:
        return None
        
    # Check if user already exists with this email (linked identifier)
    user_by_email = User.objects.filter(email=link_identifier).first()
    if user_by_email:
        return user_by_email
        
    # Generate unique username from first name
    base_username = ''.join(c for c in customer.first_name.strip().lower() if c.isalnum())
    if not base_username:
        base_username = "user"
        
    username = base_username
    counter = 1
    while User.objects.filter(username__iexact=username).exists():
        phone_clean = ''.join(filter(str.isdigit, customer.phone)) if customer.phone else ''
        last_4 = phone_clean[-4:] if len(phone_clean) >= 4 else str(random.randint(1000, 9999))
        username = f"{base_username}_{last_4}"
        if User.objects.filter(username__iexact=username).exists():
            username = f"{base_username}_{last_4}_{counter}"
            counter += 1
            
    phone_clean = ''.join(filter(str.isdigit, customer.phone)) if customer.phone else ''
    password = phone_clean[-4:] if len(phone_clean) >= 4 else "1234"
    
    user = User.objects.create_user(
        username=username,
        password=password,
        first_name=customer.first_name,
        last_name=customer.last_name or '',
        email=link_identifier
    )
    return user

def sync_existing_customers_to_users():
    User = get_user_model()
    customers = Customer.objects.all()
    for customer in customers:
        link_identifier = customer.phone if customer.phone else customer.custom_id
        if not link_identifier:
            continue
            
        # Check if already linked via email
        user_by_email = User.objects.filter(email=link_identifier).first()
        if user_by_email:
            continue
            
        # Check if old user exists with username = phone/custom_id
        old_username = customer.phone if customer.phone else customer.custom_id
        old_user = User.objects.filter(username=old_username).first()
        if old_user:
            old_user.delete()
            
        # Generate new unique username from first name
        base_username = ''.join(c for c in customer.first_name.strip().lower() if c.isalnum())
        if not base_username:
            base_username = "user"
            
        username = base_username
        counter = 1
        while User.objects.filter(username__iexact=username).exists():
            phone_clean = ''.join(filter(str.isdigit, customer.phone)) if customer.phone else ''
            last_4 = phone_clean[-4:] if len(phone_clean) >= 4 else str(random.randint(1000, 9999))
            username = f"{base_username}_{last_4}"
            if User.objects.filter(username__iexact=username).exists():
                username = f"{base_username}_{last_4}_{counter}"
                counter += 1
                
        phone_clean = ''.join(filter(str.isdigit, customer.phone)) if customer.phone else ''
        password = phone_clean[-4:] if len(phone_clean) >= 4 else "1234"
        
        User.objects.create_user(
            username=username,
            password=password,
            first_name=customer.first_name,
            last_name=customer.last_name or '',
            email=link_identifier
        )


# =====================================================================
# MULTI-TAROZI UCHUN GLOBAL O'ZGARUVCHILAR
# =====================================================================
CURRENT_SCALES = {
    "1": 0.000,
    "2": 0.000
}
PRESSED_BUTTONS = {
    "1": False,
    "2": False
}

@csrf_exempt
def receive_weight_from_esp(request):
    global CURRENT_SCALES, PRESSED_BUTTONS
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            scale_id = str(data.get('scale_id', '1'))
            if scale_id in CURRENT_SCALES:
                vazn_qiymati = data.get('vazn') if data.get('vazn') is not None else data.get('weight', 0.000)
                CURRENT_SCALES[scale_id] = float(vazn_qiymati)
                if data.get('button_pressed') is True or 'vazn' in data:
                    PRESSED_BUTTONS[scale_id] = True
                return JsonResponse({
                    'status': 'success',
                    'scale_id': scale_id,
                    'current': CURRENT_SCALES[scale_id]
                })
            else:
                return JsonResponse({'status': 'error', 'message': "Noma'lum tarozi ID raqami"}, status=400)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Faqat POST so\'rovlar qabul qilinadi'}, status=405)

@staff_required
def get_current_weight(request):
    global CURRENT_SCALES, PRESSED_BUTTONS
    scale_id = str(request.GET.get('scale_id', '1'))
    
    # Act as a proxy if scale_ip is passed (scale_id == '2' for WiFi Scale)
    scale_ip = request.GET.get('scale_ip')
    if scale_id == "2" and scale_ip:
        import requests
        try:
            res = requests.get(f"http://{scale_ip}/pos/api/get-weight/", timeout=0.5)
            if res.status_code == 200:
                data = res.json()
                # Update our global memory value too for tracking
                v = data.get('vazn') if data.get('vazn') is not None else data.get('weight', 0.000)
                CURRENT_SCALES[scale_id] = float(v)
                if data.get('button_pressed') is True:
                    PRESSED_BUTTONS[scale_id] = True
                return JsonResponse(data)
        except Exception:
            pass

    if scale_id not in CURRENT_SCALES:
        return JsonResponse({'status': 'error', 'message': 'Noto\'g\'ri tarozi ID'}, status=400)
    response_data = {
        'weight': CURRENT_SCALES[scale_id],
        'vazn': CURRENT_SCALES[scale_id],
        'button_pressed': PRESSED_BUTTONS[scale_id],
        'active_scale': scale_id
    }
    if PRESSED_BUTTONS[scale_id]:
        PRESSED_BUTTONS[scale_id] = False
    return JsonResponse(response_data)

# =====================================================================
# TERMINAL VA QIDIRUV REJIMLARI
# =====================================================================
@staff_required
def terminal_view(request):
    sync_existing_customers_to_users()
    
    # Serve built React SPA if index.html exists
    from django.conf import settings
    import os
    react_index = os.path.join(settings.BASE_DIR, 'frontend', 'dist', 'index.html')
    if os.path.exists(react_index):
        with open(react_index, 'r', encoding='utf-8') as f:
            response = HttpResponse(f.read())
            response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response['Pragma'] = 'no-cache'
            response['Expires'] = '0'
            return response

    products = Product.objects.filter(is_active=True).select_related('stock')
    return render(request, 'pos/terminal.html', {'products': products})


def customer_display_view(request):
    """Mijoz uchun 2-Ekran (Customer Facing Display) va Jonli Kassa / Dinamik QR to'lov ekrani."""
    from .models import StoreSetting, PaymentSetting
    store = StoreSetting.objects.filter(is_active=True).first()
    payment_settings = PaymentSetting.objects.filter(is_active=True)
    context = {
        'store': store,
        'payment_settings': payment_settings,
    }
    return render(request, 'pos/customer_display.html', context)


@staff_required
def search_customers(request):
    query = request.GET.get('q', '').strip()
    if not query:
        return JsonResponse([], safe=False)

    script_mode = request.session.get('script_mode', 'latin')
    query_latin = cyrillic_to_latin(query)
    query_cyrillic = latin_to_cyrillic(query)

    customers = Customer.objects.filter(
        Q(first_name__icontains=query) | Q(last_name__icontains=query) |
        Q(first_name__icontains=query_latin) | Q(last_name__icontains=query_latin) |
        Q(first_name__icontains=query_cyrillic) | Q(last_name__icontains=query_cyrillic) |
        Q(custom_id__icontains=query) |
        Q(phone__icontains=query)
    ).distinct()[:5]

    results = []
    for c in customers:
        image_url = c.image.url if c.image else 'https://cdn-icons-png.flaticon.com/512/149/149071.png'
        customer_name = f"{c.first_name} {c.last_name or ''}".strip()
        customer_note = c.note or ''

        if script_mode == 'cyrillic':
            customer_name = latin_to_cyrillic(customer_name)
            customer_note = latin_to_cyrillic(customer_note)

        is_barter = hasattr(c, 'supplier_profile') and c.supplier_profile is not None
        sup_debt = float(c.supplier_profile.our_debt) if is_barter else 0.0

        results.append({
            'id':             c.id,
            'name':           customer_name + (" [Chorvador]" if is_barter else ""),
            'id_num':         c.custom_id,
            'phone':          c.phone or '',
            'image':          image_url,
            'debt':           float(c.debt_amount),
            'note':           customer_note,
            'debt_limit':     float(c.debt_limit),
            'credit_score':   c.get_credit_score(),
            'is_blacklisted': c.is_blacklisted,
            'smart_score':    c.calculate_smart_score(),
            'is_barter':      is_barter,
            'supplier_debt':  sup_debt,
        })
    return JsonResponse(results, safe=False)

@csrf_exempt
@staff_required
@transaction.atomic
def update_customer(request, customer_id):

    if request.method == 'POST':
        try:
            customer = Customer.objects.get(id=customer_id)
            first_name = request.POST.get('first_name', '').strip()
            last_name  = request.POST.get('last_name', '').strip()
            phone      = request.POST.get('phone', '').strip()
            note       = request.POST.get('note', '').strip()
            image      = request.FILES.get('image')

            if not first_name:
                return JsonResponse({'status': 'error', 'message': "Ism majburiy!"})

            customer.first_name = first_name
            customer.last_name  = last_name
            customer.phone      = phone
            customer.note       = note
            if image:
                customer.image = image
            customer.save()

            return JsonResponse({
                'status':  'success',
                'message': "Mijoz ma'lumotlari yangilandi!",
                'customer': {
                    'id':    customer.id,
                    'name':  f"{first_name} {last_name}".strip(),
                    'id_num': customer.custom_id,
                    'phone': customer.phone or '',
                    'note':  customer.note or '',
                    'image': customer.image.url if customer.image
                             else 'https://cdn-icons-png.flaticon.com/512/149/149071.png',
                }
            })
        except Customer.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': "Mijoz topilmadi!"})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': "Noto'g'ri so'rov usuli"})

def allocate_sale_to_batch(sale_item, product, weight):
    # 1. Determine animal type if carcass-derived
    animal_type = None
    name_lower = product.name.lower()
    if "qo'y" in name_lower or "qoy" in name_lower:
        animal_type = 'qoy'
    elif "mol" in name_lower or "gosht" in name_lower or "go'sht" in name_lower or "ilik" in name_lower or "moy" in name_lower or "dumg" in name_lower:
        animal_type = 'mol'
        
    # 2. Get active slaughters of this animal type
    slaughters = []
    if animal_type:
        slaughters = list(Slaughter.objects.filter(
            animal_type=animal_type,
            status='active',
            remaining_weight__gt=0
        ))
        
    # 3. Get active stock batches for this specific product or its parent product
    from .models import StockBatch
    target_product = product.deduct_from if product.deduct_from else product
    batches = list(StockBatch.objects.filter(
        product=target_product,
        current_quantity__gt=0
    ))
    
    # 4. Combine and sort by created_at ascending (FIFO)
    all_inventories = []
    for s in slaughters:
        all_inventories.append({
            'type': 'slaughter',
            'obj': s,
            'created_at': s.created_at,
            'qty': s.remaining_weight
        })
    for b in batches:
        all_inventories.append({
            'type': 'batch',
            'obj': b,
            'created_at': b.created_at,
            'qty': b.current_quantity
        })
        
    all_inventories.sort(key=lambda x: x['created_at'])
    
    if not all_inventories:
        return
        
    # 5. Allocate weight using FIFO
    remaining_weight_to_deduct = weight
    for inv in all_inventories:
        if remaining_weight_to_deduct <= 0:
            break
            
        obj = inv['obj']
        if inv['type'] == 'slaughter':
            # Deduct from slaughter
            if remaining_weight_to_deduct <= obj.remaining_weight:
                obj.remaining_weight -= remaining_weight_to_deduct
                if obj.remaining_weight <= Decimal('0.005'):
                    obj.remaining_weight = Decimal('0.000')
                    obj.status = 'completed'
                obj.save()
                
                sale_item.slaughter = obj
                sale_item.save()
                remaining_weight_to_deduct = 0
            else:
                deducted = obj.remaining_weight
                obj.remaining_weight = Decimal('0.000')
                obj.status = 'completed'
                obj.save()
                
                sale_item.slaughter = obj
                sale_item.save()
                remaining_weight_to_deduct -= deducted
        else:
            # Deduct from stock batch
            if remaining_weight_to_deduct <= obj.current_quantity:
                obj.current_quantity -= remaining_weight_to_deduct
                obj.save()
                
                sale_item.stock_batch = obj
                sale_item.save()
                remaining_weight_to_deduct = 0
            else:
                deducted = obj.current_quantity
                obj.current_quantity = Decimal('0.000')
                obj.save()
                
                sale_item.stock_batch = obj
                sale_item.save()
                remaining_weight_to_deduct -= deducted

@csrf_exempt
@staff_required
@transaction.atomic
def save_sale(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            customer_id    = data.get('customer_id')
            payment_method = data.get('payment_method', 'naqd')
            cart_items     = data.get('items', [])

            try:
                frontend_val = int(
                    str(data.get('total_amount', 0))
                    .replace(' ', '').replace(',', '').split('.')[0]
                )
            except (ValueError, TypeError):
                frontend_val = 0

            if not cart_items:
                return JsonResponse({'status': 'error', 'message': "Savat bo'sh!"})

            if payment_method == 'nasiya' and not customer_id:
                return JsonResponse({
                    'status': 'error',
                    'message': "Nasiyaga faqat ro'yxatdan o'tgan mijozlarga sotish mumkin!"
                })

            customer = None
            if customer_id:
                customer = Customer.objects.select_for_update().get(id=customer_id)

            # Pre-calculate sale totals and validate stock
            total_sale_amount = Decimal('0.00')
            for item in cart_items:
                product       = Product.objects.get(id=item.get('product_id'))
                weight        = Decimal(str(item.get('weight')))
                
                # Check stock availability with row lock
                target_product = product.deduct_from if product.deduct_from else product
                stock, created = Stock.objects.select_for_update().get_or_create(product=target_product)
                if stock.quantity < weight:
                    return JsonResponse({
                        'status': 'error',
                        'message': f"Xatolik: {product.name} zaxirasi yetarli emas! Omborda: {stock.quantity:.3f} kg. Siz kiritganingiz: {weight:.3f} kg."
                    })
                
                price_at_sale = product.price_per_kg
                item_total    = weight * price_at_sale
                total_sale_amount += item_total

            if frontend_val > 0:
                # Round frontend value to nearest 100 so'm
                rounded = round(frontend_val / 100) * 100
                final_amount = Decimal(str(rounded))
            else:
                rounded = round(int(total_sale_amount) / 100) * 100
                final_amount = Decimal(str(rounded))

            # Dynamic Credit Limit & Blacklist Enforcement (Bypassed if customer is a supplier barter profile)
            if customer and not (hasattr(customer, 'supplier_profile') and customer.supplier_profile is not None):
                projected_debt_increase = Decimal('0.00')
                if payment_method == 'nasiya':
                    projected_debt_increase = final_amount
                else:
                    discount = int(total_sale_amount) - int(final_amount)
                    if discount > 0:
                        available = customer.bonus_points
                        if available < discount:
                            projected_debt_increase = Decimal(str(discount - available))
                
                if projected_debt_increase > 0:
                    if customer.is_blacklisted:
                        return JsonResponse({
                            'status': 'error',
                            'message': "Xatolik: Ushbu mijoz qora ro'yxatga olingan! Nasiya savdo qilish taqiqlanadi."
                        })
                    if customer.debt_amount + projected_debt_increase > customer.debt_limit:
                        return JsonResponse({
                            'status': 'error',
                            'message': f"Xatolik: Mijoz kredit limitidan oshib ketdi! Qarz limiti: {int(customer.debt_limit):,} so'm. Joriy qarz: {int(customer.debt_amount):,} so'm. Nasiya summasi: {int(projected_debt_increase):,} so'm."
                        })

            # Validate active kassa shift
            from .models import CashierShift
            active_shift = CashierShift.objects.filter(cashier=request.user, is_open=True).first()
            if not active_shift:
                return JsonResponse({
                    'status': 'error',
                    'message': "Xatolik: Kassa shifti ochilmagan! Iltimos, terminalda yangi shift oching."
                })

            # Create Sale
            sale = Sale.objects.create(
                customer=customer,
                shift=active_shift,
                payment_method=payment_method,
                total_amount=final_amount,
                discount_amount=Decimal('0.00'),
                bonus_used=Decimal('0.00'),
                debt_added=Decimal('0.00'),
                final_paid=Decimal('0.00')
            )

            items_log_details = []
            for item in cart_items:
                product       = Product.objects.get(id=item.get('product_id'))
                weight        = Decimal(str(item.get('weight')))
                price_at_sale = product.price_per_kg
                item_total    = weight * price_at_sale

                # Update Stock (deduct from parent product if defined)
                target_product = product.deduct_from if product.deduct_from else product
                stock, created = Stock.objects.select_for_update().get_or_create(product=target_product)
                stock.quantity -= weight
                stock.save()

                sale_item = SaleItem.objects.create(
                    sale=sale,
                    product=product,
                    weight=weight,
                    price_at_sale=price_at_sale
                )
                allocate_sale_to_batch(sale_item, product, weight)

                items_log_details.append({
                    'product_name': product.name,
                    'weight': float(weight),
                    'price': float(price_at_sale),
                    'total': float(item_total)
                })


            discount = int(total_sale_amount) - int(final_amount)
            if discount < 0:
                discount = 0

            sale.discount_amount = Decimal(str(discount))

            bonus_used  = 0
            debt_added  = 0
            bonus_earned = 0

            if customer:
                fa_fmt = "{:,}".format(int(final_amount)).replace(',', ' ')

                is_barter = False
                supplier = None
                if hasattr(customer, 'supplier_profile') and customer.supplier_profile is not None:
                    supplier = customer.supplier_profile
                    is_barter = True

                if payment_method == 'nasiya':
                    if is_barter:
                        supplier.our_debt -= final_amount
                        supplier.save()
                        sale.debt_added = final_amount
                        sale.final_paid = Decimal('0.00')
                        log_type = 'debt_add'
                        title = "🧾 BARTER CHEK (NASIYA - TA'MINOTCHI)"
                    else:
                        customer.debt_amount += final_amount
                        sale.debt_added = final_amount
                        sale.final_paid = Decimal('0.00')
                        customer.save()
                        log_type    = 'debt_add'
                        title       = "🧾 ELEKTRON CHEK (NASIYA)"

                else:
                    sale.final_paid = final_amount

                    if discount > 0:
                        available = customer.bonus_points

                        if available >= discount:
                            customer.bonus_points -= discount
                            bonus_used = discount
                            sale.bonus_used = Decimal(str(discount))
                            d_fmt = "{:,}".format(discount).replace(',', ' ')
                            CustomerLog.objects.create(
                                customer=customer,
                                log_type='bonus',
                                title='💎 Bonus ball sarflandi',
                                message=f"Savdo #{sale.id}: {d_fmt} so'm chegirma uchun bonus ayirildi.",
                                amount=Decimal(str(discount))
                            )
                        else:
                            bonus_used   = available
                            sale.bonus_used = Decimal(str(available))
                            remaining    = discount - available
                            customer.bonus_points = 0
                            
                            if is_barter:
                                supplier.our_debt -= Decimal(str(remaining))
                                supplier.save()
                                sale.debt_added = Decimal(str(remaining))
                                debt_added = remaining
                            else:
                                customer.debt_amount += Decimal(str(remaining))
                                sale.debt_added = Decimal(str(remaining))
                                debt_added = remaining

                            b_fmt = "{:,}".format(available).replace(',', ' ')
                            r_fmt = "{:,}".format(remaining).replace(',', ' ')
                            CustomerLog.objects.create(
                                customer=customer,
                                log_type='debt_add',
                                title='⚠️ Chegirma barterga yozildi' if is_barter else '⚠️ Chegirma qarzga yozildi',
                                message=(
                                    f"Savdo #{sale.id}: {b_fmt} so'm bonus ayirildi, qolgan {r_fmt} so'm ta'minotchi qarzidan chegirildi." if is_barter else
                                    f"Savdo #{sale.id}: {b_fmt} so'm bonus ayirildi, qolgan {r_fmt} so'm qarzga yozildi."
                                ),
                                amount=Decimal(str(remaining))
                            )

                    if bonus_used == 0 and payment_method != 'nasiya':
                        from .models import StoreSetting
                        store_conf = StoreSetting.objects.filter(is_active=True).first()
                        cb_rate = (store_conf.cashback_percent / Decimal('100.0')) if store_conf else Decimal('0.02')
                        bonus_earned = int(final_amount * cb_rate)
                        customer.bonus_points += bonus_earned

                    customer.save()
                    log_type    = 'sale'
                    title = "🧾 BARTER CHEK (TO'LANDI)" if is_barter else "🧾 ELEKTRON CHEK (TO'LANDI)"

                items_summary = ", ".join([f"{item['product_name']} ({item['weight']:.3f} kg)" for item in items_log_details])
                summary_message = f"Xarid: {items_summary}. Jami: {fa_fmt} so'm."

                CustomerLog.objects.create(
                    customer=customer,
                    log_type=log_type,
                    title=title,
                    message=summary_message,
                    details={
                        'sale_id': sale.id,
                        'items': items_log_details,
                        'total': float(final_amount),
                        'discount': float(discount),
                        'bonus_used': float(bonus_used),
                        'bonus_earned': float(bonus_earned),
                        'debt_added': float(debt_added),
                        'final_paid': float(sale.final_paid),
                        'payment_method': payment_method
                    },
                    amount=final_amount
                )

            sale.save()

            # Avtomatik Telegram va SMS elektron chek jo'natish
            if customer:
                try:
                    from .receipt_service import dispatch_customer_sale_receipt
                    dispatch_customer_sale_receipt(sale, customer, items_log_details, bonus_earned=bonus_earned, bonus_used=bonus_used, debt_added=debt_added)
                except Exception as re_err:
                    print(f"[Receipt Dispatch Error]: {re_err}")

            # Telegram notification for large debt (> 500,000 so'm)
            if payment_method == 'nasiya' and final_amount >= Decimal('500000.00'):
                try:
                    from .views_api import send_telegram_notification
                    cust_name = f"{customer.first_name} {customer.last_name or ''}".strip()
                    msg = (
                        f"⚠️ **KATTA NASIYA SAVDO (ALERT)**\n"
                        f"👤 Mijoz: {cust_name} (ID: {customer.custom_id})\n"
                        f"💵 Summa: {final_amount:,.0f} so'm\n"
                        f"📊 Joriy jami qarzi: {customer.debt_amount:,.0f} so'm\n"
                        f"📅 Savdo #: {sale.id}\n"
                        f"👤 Kassir: {request.user.username}"
                    )
                    send_telegram_notification(msg)
                except Exception as te:
                    print("Telegram notification error:", te)

            return JsonResponse({
                'status':       'success',
                'message':      "Savdo muvaffaqiyatli saqlandi!",
                'sale_id':      sale.id,
                'total_amount': int(final_amount),
                'bonus_used':   bonus_used,
                'bonus_earned': bonus_earned,
                'debt_added':   debt_added,
            })


        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    return JsonResponse({'status': 'error', 'message': "Noto'g'ri so'rov usuli"})

@csrf_exempt
@staff_required
@transaction.atomic
def quick_create_customer(request):
    if request.method == 'POST':
        try:
            first_name = request.POST.get('first_name', '').strip()
            last_name  = request.POST.get('last_name', '').strip()
            phone      = request.POST.get('phone', '').strip()
            custom_id  = request.POST.get('custom_id', '').strip()
            image      = request.FILES.get('image')

            if not first_name or not phone or not custom_id:
                return JsonResponse({'status': 'error', 'message': "Ism, telefon va ID majburiy!"})

            if Customer.objects.filter(phone=phone).exists():
                return JsonResponse({'status': 'error', 'message': "Ushbu telefon raqami bilan mijoz mavjud!"})

            if Customer.objects.filter(custom_id=custom_id).exists():
                return JsonResponse({'status': 'error', 'message': "Ushbu ID bilan mijoz mavjud!"})

            customer = Customer.objects.create(
                first_name=first_name,
                last_name=last_name,
                phone=phone,
                custom_id=custom_id,
                image=image
            )
            create_user_for_customer(customer)

            return JsonResponse({
                'status':  'success',
                'message': "Yangi mijoz muvaffaqiyatli qo'shildi!",
                'customer': {
                    'id':     customer.id,
                    'name':   f"{customer.first_name} {customer.last_name or ''}".strip(),
                    'id_num': customer.custom_id,
                }
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    return JsonResponse({'status': 'error', 'message': "Noto'g'ri so'rov usuli"})

# =====================================================================
# HISOBLAR VA KUNLIK Z-REPORT
# =====================================================================
@staff_required
def daily_report_view(request):
    from .models import CashTransaction
    local_now = timezone.localtime(timezone.now())
    today = local_now.date()

    todays_sales = Sale.objects.filter(created_at__date=today).select_related("customer")

    naqd_total = todays_sales.filter(payment_method="naqd").aggregate(Sum("total_amount"))["total_amount__sum"] or Decimal("0.00")
    karta_total = todays_sales.filter(payment_method="karta").aggregate(Sum("total_amount"))["total_amount__sum"] or Decimal("0.00")
    qr_total = todays_sales.filter(payment_method="qr").aggregate(Sum("total_amount"))["total_amount__sum"] or Decimal("0.00")
    nasiya_total = todays_sales.filter(payment_method="nasiya").aggregate(Sum("total_amount"))["total_amount__sum"] or Decimal("0.00")

    overall_total = naqd_total + karta_total + qr_total + nasiya_total

    # Cash Flow Kirim/Chiqim calculations for today
    today_cash_ins = CashTransaction.objects.filter(created_at__date=today, transaction_type='in')
    today_cash_outs = CashTransaction.objects.filter(created_at__date=today, transaction_type='out')

    cash_in_total = today_cash_ins.aggregate(Sum("amount"))["amount__sum"] or Decimal("0.00")
    cash_out_total = today_cash_outs.aggregate(Sum("amount"))["amount__sum"] or Decimal("0.00")
    
    # Net change to naqd drawer
    naqd_ins_today = today_cash_ins.filter(payment_method='naqd').aggregate(Sum("amount"))["amount__sum"] or Decimal("0.00")
    naqd_outs_today = today_cash_outs.filter(payment_method='naqd').aggregate(Sum("amount"))["amount__sum"] or Decimal("0.00")
    today_net_drawer_change = naqd_total + naqd_ins_today - naqd_outs_today

    context = {
        "today": today,
        "naqd": naqd_total,
        "karta": karta_total,
        "qr": qr_total,
        "nasiya": nasiya_total,
        "overall": overall_total,
        "sales_count": todays_sales.count(),
        "todays_sales": todays_sales,
        # Cash Flow metrics
        "cash_in_total": cash_in_total,
        "cash_out_total": cash_out_total,
        "net_cash_flow": cash_in_total - cash_out_total,
        "today_net_drawer_change": today_net_drawer_change,
        "today_cash_transactions": CashTransaction.objects.filter(created_at__date=today).select_related('created_by').order_by('-created_at'),
    }
    return render(request, "daily_report.html", context)

# =====================================================================
# MIJOZ CHAT VA ALOQA
# =====================================================================
@staff_required
def get_customer_chat_logs(request, customer_id):
    try:
        customer = Customer.objects.get(id=customer_id)
        
        try:
            limit = int(request.GET.get('limit', 10))
            offset = int(request.GET.get('offset', 0))
        except ValueError:
            limit = 10
            offset = 0

        logs_query = customer.logs.all().order_by('-created_at')
        total_count = logs_query.count()
        # Slice newest N records, then reverse to get chronological order (oldest→newest)
        logs_sliced = list(reversed(list(logs_query[offset:offset+limit])))

        logs_list = []
        for log in logs_sliced:
            loc_time = timezone.localtime(log.created_at)
            logs_list.append({
                'id':       log.id,
                'log_type': log.log_type,
                'title':    log.title,
                'message':  log.message,
                'details':  log.details,
                'amount':   float(log.amount),
                'date':     loc_time.strftime('%d-%m-%Y'),
                'time':     loc_time.strftime('%H:%M')
            })

        image_url = customer.image.url if customer.image else 'https://cdn-icons-png.flaticon.com/512/149/149071.png'

        is_barter = hasattr(customer, 'supplier_profile') and customer.supplier_profile is not None
        sup_debt = float(customer.supplier_profile.our_debt) if is_barter else 0.0

        return JsonResponse({
            'status':          'success',
            'customer_name':   f"{customer.first_name} {customer.last_name or ''}".strip(),
            'customer_id_num': customer.custom_id,
            'image':           image_url,
            'debt_amount':     float(customer.debt_amount),
            'bonus_points':    customer.bonus_points,
            'logs':            logs_list,
            'has_more':        (offset + limit) < total_count,
            'is_barter':       is_barter,
            'supplier_debt':   sup_debt,
        })
    except Customer.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': "Mijoz topilmadi!"}, status=404)

@csrf_exempt
@login_required
@transaction.atomic
def send_chat_message(request, customer_id):
    if request.method == 'POST':
        try:
            customer = Customer.objects.get(id=customer_id)
            
            # Authorization check: must be superuser OR the owner of the profile
            user_ident = request.user.email if request.user.email else request.user.username
            is_owner = (user_ident == customer.phone) or (user_ident == customer.custom_id)
            if not request.user.is_superuser and not is_owner:
                return JsonResponse({'status': 'error', 'message': "Sizda ushbu chatga xabar yuborish huquqi yo'q!"}, status=403)
            
            body = json.loads(request.body)
            message_text = body.get('message', '').strip()

            if not message_text:
                return JsonResponse({'status': 'error', 'message': "Xabar matni bo'sh!"}, status=400)

            # Title is determined by sender
            title = "Do'kon xabari" if request.user.is_superuser else "Mijoz xabari"
            source = 'admin_web' if request.user.is_superuser else 'web'

            log = CustomerLog.objects.create(
                customer=customer,
                log_type='bonus',
                title=title,
                message=message_text,
                details={'source': source},
                amount=Decimal('0.00')
            )

            # Send Telegram push notification if customer has telegram_chat_id and sent by store
            if request.user.is_superuser and customer.telegram_chat_id:
                try:
                    import requests
                    bot_token = getattr(settings, 'CUSTOMER_BOT_TOKEN', '8728579523:AAG9mtvfeVd95svVNVJ-NnGUOTknTzPkDg8')
                    tg_msg = f"💬 *Baxmal Meat Operatoridan Xabar:*\n\n{message_text}"
                    requests.post(
                        f"https://api.telegram.org/bot{bot_token}/sendMessage",
                        json={'chat_id': customer.telegram_chat_id, 'text': tg_msg, 'parse_mode': 'Markdown'},
                        timeout=3
                    )
                except Exception as tg_err:
                    print(f"Customer TG push error: {tg_err}")

            # Send SMS push notification if requested or customer has no Telegram
            send_sms_flag = body.get('send_sms', False)
            if request.user.is_superuser and (send_sms_flag or not customer.telegram_chat_id) and customer.phone:
                try:
                    from .sms_service import send_sms
                    sms_text = f"Baxmal Meat: {message_text}"
                    send_sms(customer.phone, sms_text)
                except Exception as sms_err:
                    print(f"Customer SMS push error: {sms_err}")

            # If message is from customer (not superuser), notify admin via Telegram
            if not request.user.is_superuser:
                try:
                    from .views_api import send_telegram_notification
                    admin_notify = (
                        f"💬 *Saytdan Mijoz Xabari!*\n"
                        f"━━━━━━━━━━━━━━━━━━━\n"
                        f"👤 *Mijoz:* {customer.first_name} {customer.last_name or ''} (ID: `{customer.custom_id}`)\n"
                        f"📞 *Tel:* {customer.phone}\n\n"
                        f"💬 *Xabar:* {message_text}\n\n"
                        f"🌐 _Sayt orqali yuborildi_"
                    )
                    send_telegram_notification(admin_notify)
                except Exception as notify_err:
                    print(f"Admin notify error: {notify_err}")

            loc_time = timezone.localtime(log.created_at)

            return JsonResponse({
                'status': 'success',
                'log': {
                    'id':      log.id,
                    'title':   log.title,
                    'message': log.message,
                    'date':    loc_time.strftime('%d-%m-%Y'),
                    'time':    loc_time.strftime('%H:%M')
                }
            })
        except Customer.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': "Mijoz topilmadi!"}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': "Noto'g'ri so'rov"}, status=405)

@csrf_exempt
@login_required
def customer_chat_api(request):
    """Mijozning o'z chat loglarini olish (real-time polling uchun)."""
    user_ident = request.user.email if request.user.email else request.user.username
    customer = Customer.objects.filter(Q(phone__iexact=user_ident) | Q(custom_id__iexact=user_ident)).first()
    if not customer:
        return JsonResponse({'status': 'error', 'message': 'Profil topilmadi'}, status=404)

    since_id = int(request.GET.get('since_id', 0))
    limit = int(request.GET.get('limit', 50))

    logs_qs = customer.logs.filter(id__gt=since_id).order_by('created_at')[:limit]

    logs_list = []
    for log in logs_qs:
        loc_time = timezone.localtime(log.created_at)
        logs_list.append({
            'id':       log.id,
            'log_type': log.log_type,
            'title':    log.title,
            'message':  log.message,
            'details':  log.details,
            'amount':   float(log.amount),
            'date':     loc_time.strftime('%d.%m.%Y'),
            'time':     loc_time.strftime('%H:%M'),
            'is_mine':  log.title == "Mijoz xabari",
        })

    return JsonResponse({'status': 'success', 'logs': logs_list, 'customer_id': customer.id})


@csrf_exempt
@login_required
def ai_meat_assistant_api(request):
    """Saytdan AI Go'sht Maslahatchisi savollariga javob berish."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Faqat POST!'}, status=405)
    try:
        data = json.loads(request.body)
        prompt = data.get('prompt', '').strip()
        if not prompt:
            return JsonResponse({'status': 'error', 'message': "Savol matni kiritilmadi!"}, status=400)
        from .customer_bot import query_gemini_meat_assistant
        reply = query_gemini_meat_assistant(prompt)
        return JsonResponse({'status': 'success', 'reply': reply})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@staff_required
def customer_chats_dashboard(request):
    customers = Customer.objects.all().order_by('-debt_amount')
    return render(request, 'customer_chats.html', {'customers': customers})

@login_required
def customer_profile_cabinet(request):
    cust_id_param = request.GET.get('customer_id')
    if cust_id_param and request.user.is_superuser:
        customer = Customer.objects.filter(id=cust_id_param).first()
    else:
        user_ident = request.user.email if request.user.email else request.user.username
        customer = Customer.objects.filter(Q(phone__iexact=user_ident) | Q(custom_id__iexact=user_ident)).first()
        if not customer and request.user.is_superuser:
            return redirect('customers')
        
    if not customer:
        return render(request, 'customer_cabinet.html', {'customer': None, 'error_message': "Sizning telefoningiz yoki ID raqamingizga mos mijoz profili topilmadi!"})
    
    logs = customer.logs.all()
    chat_logs = logs.order_by('created_at')
    tx_logs = logs.exclude(Q(title="Mijoz xabari") | Q(title="Do'kon xabari")).order_by('-created_at')
    products = Product.objects.filter(is_active=True)
    b2b_orders = customer.b2b_orders.all().order_by('-created_at')
    
    # ── DETAILED RECONCILIATION (AKT-SVERKA) CALCULATIONS ──
    from .models import Slaughter, CashTransaction
    
    # 1. Jami topshirilgan chorva go'shti (Bizning qarzimiz)
    slaughter_total = Slaughter.objects.filter(
        Q(supplier__customer=customer) | Q(customer=customer)
    ).aggregate(s=Sum('total_cost'))['s'] or Decimal('0.00')
    
    # 2. Uning bizdan olgan jami naqd pullari
    supplier_pay_sum = CashTransaction.objects.filter(
        supplier__customer=customer, 
        category='supplier_pay'
    ).aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
    
    # 3. Uning bizdan nasiyaga olgan go'shtlari
    sales_debt_sum = customer.sales.aggregate(s=Sum('debt_added'))['s'] or Decimal('0.00')
    
    # 4. Jami uning olganlari (Nasiya go'sht + Naqd pul)
    drawings_total = sales_debt_sum + supplier_pay_sum
    
    # 5. Yakuniy balans (Bizning qarzimiz - Uning olganlari)
    net_balance = slaughter_total - drawings_total
    net_balance_abs = abs(net_balance)
    
    total_bonus_earned = logs.filter(log_type='bonus').filter(title__icontains="yig'ildi").aggregate(s=Sum('amount'))['s'] or Decimal('0.00')

    # Payment settings (Cards & QR Codes)
    from .models import PaymentSetting
    payment_settings = list(PaymentSetting.objects.filter(is_active=True))
    if not payment_settings:
        p1 = PaymentSetting.objects.create(
            title="Uzcard / Click",
            card_number="8600 5304 8877 4477",
            card_holder="Baxmal Meat LLC",
            instructions="To'lovni amalga oshirgach, chek rasmini yuklang.",
            is_active=True
        )
        p2 = PaymentSetting.objects.create(
            title="Humo / Payme",
            card_number="9860 1201 9988 5522",
            card_holder="Baxmal Meat Kassa",
            instructions="Payme yoki Humo orqali to'lov qiling.",
            is_active=True
        )
        payment_settings = [p1, p2]

    # Payment proofs
    from .models import PaymentProof
    payment_proofs = customer.payment_proofs.all().order_by('-created_at')

    context = {
        'customer': customer,
        'chat_logs': chat_logs,
        'tx_logs': tx_logs,
        'products': products,
        'b2b_orders': b2b_orders,
        'slaughter_total': slaughter_total,
        'supplier_pay_sum': supplier_pay_sum,
        'sales_debt_sum': sales_debt_sum,
        'drawings_total': drawings_total,
        'net_balance': net_balance,
        'net_balance_abs': net_balance_abs,
        'total_bonus_earned': total_bonus_earned,
        'payment_proofs': payment_proofs,
        'payment_settings': payment_settings,
    }
    return render(request, 'customer_cabinet.html', context)


@csrf_exempt
@login_required
@transaction.atomic
def upload_payment_proof(request):
    """Mijoz karta cheki rasmini yuklab to'lov so'rov qiladi."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': "Faqat POST!"}, status=405)
    try:
        user_ident = request.user.email if request.user.email else request.user.username
        customer = Customer.objects.filter(Q(phone__iexact=user_ident) | Q(custom_id__iexact=user_ident)).first()
        if not customer:
            return JsonResponse({'status': 'error', 'message': "Mijoz topilmadi!"}, status=404)

        from .models import PaymentProof
        image_file = request.FILES.get('image')
        if not image_file:
            return JsonResponse({'status': 'error', 'message': "Rasm yuklanmadi!"}, status=400)

        amount_str = request.POST.get('amount', '0')
        provider = request.POST.get('provider', 'karta').strip()
        note = request.POST.get('note', '').strip()

        try:
            amount = Decimal(amount_str)
            if amount <= 0:
                raise ValueError()
        except Exception:
            return JsonResponse({'status': 'error', 'message': "Noto'g'ri summa!"}, status=400)

        proof = PaymentProof.objects.create(
            customer=customer,
            image=image_file,
            amount=amount,
            provider=provider,
            note=note,
        )

        # Notify admin via Telegram
        try:
            from .views_api import send_telegram_notification
            no_note = note if note else "yo'q"
            msg = (
                f"📸 *YANGI TO'LOV CHEKI YUKLANDI*\n\n"
                f"👤 *Mijoz:* {customer.first_name} {customer.last_name or ''}\n"
                f"🆔 *ID:* `{customer.custom_id}`\n"
                f"💳 *Tizim:* {provider.upper()}\n"
                f"💰 *Summa:* {amount:,.0f} so'm\n"
                f"📝 *Izoh:* {no_note}\n\n"
                f"Admin paneldan tasdiqlang!"
            )
            send_telegram_notification(msg)
        except Exception as tg_err:
            print(f"Telegram notify error in upload_payment_proof: {tg_err}")

        return JsonResponse({
            'status': 'success',
            'message': "Chek muvaffaqiyatli yuklandi! Admin tekshirib, qarzingizni yopadi.",
            'proof_id': proof.id
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required
def export_customer_logs_excel(request):
    """Mijoz barcha operatsiyalarini Excel (.xlsx) formatida yuklab oladi."""
    user_ident = request.user.email if request.user.email else request.user.username
    customer = Customer.objects.filter(Q(phone__iexact=user_ident) | Q(custom_id__iexact=user_ident)).first()
    if not customer:
        return HttpResponse("Mijoz topilmadi", status=404)

    logs = customer.logs.exclude(Q(title="Mijoz xabari") | Q(title="Do'kon xabari")).order_by('-created_at')

    response = HttpResponse(content_type='application/ms-excel')
    fname = f"baxmal_meat_{customer.custom_id}_operatsiyalar.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{fname}"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mening Operatsiyalarim"

    headers = ["№", "Sana", "Amal turi", "Sarlavha", "Summa (so'm)", "Izoh"]
    ws.append(headers)

    type_labels = {
        'sale': '🛒 Xarid',
        'debt_add': '📋 Qarz ko\'paydi',
        'debt_pay': '✅ To\'lov',
        'bonus': '💎 Bonus',
    }

    for idx, log in enumerate(logs, 1):
        ws.append([
            idx,
            timezone.localtime(log.created_at).strftime('%d.%m.%Y %H:%M'),
            type_labels.get(log.log_type, log.log_type),
            log.title,
            float(log.amount),
            log.message or ''
        ])

    style_excel_worksheet(ws, len(headers))
    wb.save(response)
    return response




@csrf_exempt
@csrf_exempt
@login_required
@transaction.atomic
def create_b2b_order(request):
    if request.method == 'POST':
        try:
            user_ident = request.user.email if request.user.email else request.user.username
            customer = Customer.objects.filter(Q(phone__iexact=user_ident) | Q(custom_id__iexact=user_ident)).first()
            if not customer:
                return JsonResponse({'status': 'error', 'message': "Mijoz profili topilmadi!"}, status=404)
            
            # Support both JSON and multipart/form-data
            items_data = []
            if request.content_type and 'application/json' in request.content_type:
                body = json.loads(request.body)
                notes = str(body.get('notes', '') or '').strip()
                delivery_type = str(body.get('delivery_type', 'delivery') or 'delivery').strip()
                delivery_address = str(body.get('delivery_address', '') or '').strip()
                payment_method = str(body.get('payment_method', 'karta') or 'karta').strip()
                latitude_val = body.get('latitude')
                longitude_val = body.get('longitude')
                proof_file = None
                
                if 'items' in body and isinstance(body['items'], list):
                    items_data = body['items']
                else:
                    items_data = [{'product_name': str(body.get('product_name', '') or '').strip(), 'weight': str(body.get('weight', '') or '').strip()}]
            else:
                notes = str(request.POST.get('notes', '') or '').strip()
                delivery_type = str(request.POST.get('delivery_type', 'delivery') or 'delivery').strip()
                delivery_address = str(request.POST.get('delivery_address', '') or '').strip()
                payment_method = str(request.POST.get('payment_method', 'karta') or 'karta').strip()
                latitude_val = request.POST.get('latitude')
                longitude_val = request.POST.get('longitude')
                proof_file = request.FILES.get('proof_image')
                
                raw_items = request.POST.get('items')
                if raw_items:
                    try:
                        items_data = json.loads(raw_items)
                    except Exception:
                        items_data = []
                if not items_data:
                    items_data = [{'product_name': str(request.POST.get('product_name', '') or '').strip(), 'weight': str(request.POST.get('weight', '') or '').strip()}]
            
            if not items_data:
                return JsonResponse({'status': 'error', 'message': "Savat bo'sh!"}, status=400)
                
            lat_float = float(latitude_val) if latitude_val and str(latitude_val).strip() else None
            lng_float = float(longitude_val) if longitude_val and str(longitude_val).strip() else None

            calc_distance_km = 0.0
            calc_delivery_fee = Decimal('0.00')

            if lat_float and lng_float and delivery_type == 'delivery':
                from pos.models import StoreSetting
                from pos.views_api import calculate_haversine_distance
                store = StoreSetting.objects.filter(is_active=True).first()
                if store:
                    calc_distance_km = calculate_haversine_distance(store.latitude, store.longitude, lat_float, lng_float)
                    calc_delivery_fee = store.base_delivery_fee + (Decimal(str(calc_distance_km)) * store.fee_per_km)
                    calc_delivery_fee = calc_delivery_fee.quantize(Decimal('1000.00'))

            from pos.models import B2BOrder, CustomerLog, Product
            initial_status = 'payment_uploaded' if proof_file else 'pending'

            created_orders = []
            order_summary_items = []
            total_sum = Decimal('0.00')

            for item_info in items_data:
                p_name = str(item_info.get('product_name', '') or '').strip()
                w_val = str(item_info.get('weight', '') or '').strip()
                if not p_name or not w_val:
                    continue
                
                try:
                    product = Product.objects.get(name=p_name, is_active=True)
                except Product.DoesNotExist:
                    continue

                try:
                    weight = Decimal(str(w_val))
                    if weight <= 0:
                        continue
                except Exception:
                    continue

                order = B2BOrder.objects.create(
                    customer=customer,
                    product=product,
                    requested_weight=weight,
                    delivery_type=delivery_type,
                    delivery_address=delivery_address,
                    latitude=lat_float,
                    longitude=lng_float,
                    distance_km=calc_distance_km,
                    delivery_fee=calc_delivery_fee,
                    payment_method=payment_method,
                    payment_proof_image=proof_file,
                    notes=notes,
                    status=initial_status
                )
                created_orders.append(order)

                item_total = weight * product.price_per_kg
                total_sum += item_total
                order_summary_items.append({
                    'order': order,
                    'product_name': product.name,
                    'weight': weight,
                    'price_per_kg': product.price_per_kg,
                    'total_price': item_total
                })

            if not created_orders:
                return JsonResponse({'status': 'error', 'message': "Hech qanday mahsulot va og'irlik tanlanmadi!"}, status=400)

            delivery_str = "🚚 Yetkazib berish (Kuryer)" if delivery_type == 'delivery' else "🏃 Samovivoz (Do'kondan olib ketish)"
            pay_str = {"karta": "💳 Karta (Click/Payme)", "naqd": "💵 Naqd pul", "nasiya": "📋 Nasiya (Qarz)"}.get(payment_method, payment_method)
            
            first_order = created_orders[0]
            img_url = first_order.payment_proof_image.url if first_order.payment_proof_image else None

            # Format items summary HTML
            items_html_lines = []
            for item in order_summary_items:
                items_html_lines.append(f"• 🥩 <b>{item['product_name']}</b> — {item['weight']} kg x {item['price_per_kg']:,.0f} = <b>{item['total_price']:,.0f} so'm</b>")
            items_html = "<br>".join(items_html_lines)

            gps_chat_str = f"<br>📍 GPS: <a href=\"https://maps.google.com/?q={lat_float},{lng_float}\" target=\"_blank\" style=\"color:#3B82F6;\">Google Maps ({lat_float:.4f}, {lng_float:.4f})</a>" if lat_float and lng_float else ""

            order_ids_str = ", ".join([f"#{o.id}" for o in created_orders])
            
            # Send message to chat log
            msg_text = (
                f"📋 <b>BUYURTMA {order_ids_str} YARATILDI</b><br>"
                f"{items_html}<br>"
                f"💰 <b>Jami taxminiy summa:</b> {total_sum:,.0f} so'm<br>"
                f"{delivery_str}<br>"
                f"💳 To'lov: {pay_str}"
                f"{'<br>📍 Manzil: ' + delivery_address if delivery_address else ''}"
                f"{gps_chat_str}"
                f"{'<br>📝 Izoh: ' + notes if notes else ''}"
            )
            log_details = {'source': 'web', 'order_ids': [o.id for o in created_orders], 'latitude': lat_float, 'longitude': lng_float}
            if img_url:
                log_details['image_url'] = img_url

            CustomerLog.objects.create(
                customer=customer,
                log_type='bonus',
                title="Mijoz xabari",
                message=msg_text,
                details=log_details,
                amount=Decimal('0.00')
            )
            
            # Send Telegram Notification to Admin
            try:
                from .views_api import send_telegram_notification, send_telegram_location
                gps_tg_str = f"🗺️ *GPS Lokatsiya:* [Google Maps-da Ko'rish](https://maps.google.com/?q={lat_float},{lng_float})\n" if lat_float and lng_float else ""
                
                tg_items_lines = []
                for item in order_summary_items:
                    tg_items_lines.append(f"• *{item['product_name']}*: {item['weight']} kg (`{item['total_price']:,.0f}` so'm)")
                tg_items_text = "\n".join(tg_items_lines)

                addr_part = f"📍 *Manzil:* {delivery_address}\n" if delivery_address else ""
                notes_part = f"📝 *Izoh:* {notes}\n" if notes else ""
                proof_part = "📸 *To'lov cheki rasmi yuklandi!*\n" if proof_file else ""
                
                tg_msg = (
                    f"📦 *YANGI SAVAT BUYURTMASI ({order_ids_str})*\n"
                    f"━━━━━━━━━━━━━━━━━━━\n"
                    f"👤 *Mijoz:* {customer.first_name} {customer.last_name or ''} (ID: `{customer.custom_id}`)\n"
                    f"📞 *Tel:* `{customer.phone}`\n\n"
                    f"🥩 *Mahsulotlar:* \n{tg_items_text}\n"
                    f"━━━━━━━━━━━━━━━━━━━\n"
                    f"💵 *Jami Taxminiy Summa:* *{total_sum:,.0f} so'm*\n\n"
                    f"🚚 *Turi:* {delivery_str}\n"
                    f"💳 *To'lov:* {pay_str}\n"
                    f"{addr_part}"
                    f"{gps_tg_str}"
                    f"{notes_part}"
                    f"{proof_part}"
                    f"⏳ _Do'kon boshqaruv panelidan tasdiqlashingiz kutilmoqda._"
                )
                send_telegram_notification(tg_msg)


                # Send direct Telegram location pin card
                if lat_float and lng_float:
                    send_telegram_location(lat_float, lng_float)
            except Exception as tg_err:
                print(f"Telegram notify error in create_b2b_order: {tg_err}")

            return JsonResponse({
                'status': 'success',
                'message': "Buyurtmangiz muvaffaqiyatli qabul qilindi!",
                'order_id': first_order.id,
                'order_ids': [o.id for o in created_orders],
                'total_sum': float(total_sum)
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': "Noto'g'ri so'rov usuli!"}, status=405)


@csrf_exempt
@staff_required
@transaction.atomic
def update_b2b_order_status(request, order_id):
    if request.method == 'POST':
        try:
            from pos.models import B2BOrder, CustomerLog
            order = B2BOrder.objects.get(id=order_id)
            body = json.loads(request.body)
            new_status = body.get('status', '').strip()
            
            valid_statuses = ['pending', 'payment_uploaded', 'approved', 'preparing', 'shipping', 'completed', 'rejected']
            if new_status not in valid_statuses:
                return JsonResponse({'status': 'error', 'message': "Noto'g'ri status!"}, status=400)
                
            order.status = new_status
            order.save()
            
            status_emoji = {
                'pending': '⏳', 'payment_uploaded': '📸', 'approved': '✅',
                'preparing': '🥩', 'shipping': '🚚', 'completed': '🎉', 'rejected': '❌'
            }.get(new_status, '🔔')

            status_text = {
                'pending': 'Kutilmoqda',
                'payment_uploaded': 'To\'lov cheki yuklandi',
                'approved': 'Tasdiqlandi',
                'preparing': 'Qadoqlanmoqda',
                'shipping': 'Kuryer yo\'lda',
                'completed': 'Yakunlandi',
                'rejected': 'Rad etildi'
            }[new_status]

            msg_text = (
                f"{status_emoji} **BUYURTMA #{order.id} STATUSI O'ZGARDI**\n"
                f"📦 {order.product.name} — {order.requested_weight} kg\n"
                f"📊 Yangi holat: **{status_text}**"
            )
            CustomerLog.objects.create(
                customer=order.customer,
                log_type='bonus',
                title="Do'kon xabari",
                message=msg_text,
                details={'source': 'admin_web', 'order_id': order.id, 'new_status': new_status},
                amount=Decimal('0.00')
            )

            # Send Telegram push to customer if they have telegram_chat_id
            if order.customer.telegram_chat_id:
                try:
                    import requests as _req
                    bot_token = getattr(settings, 'CUSTOMER_BOT_TOKEN', '8728579523:AAG9mtvfeVd95svVNVJ-NnGUOTknTzPkDg8')
                    tg_push = (
                        f"{status_emoji} *Buyurtma #{order.id} Yangilandi!*\n"
                        f"━━━━━━━━━━━━━━━━━━━\n"
                        f"📦 *Mahsulot:* {order.product.name} ({order.requested_weight} kg)\n"
                        f"📊 *Yangi Holat:* *{status_text}*\n\n"
                        f"{'🎉 Buyurtmangiz yetkazib berildi!' if new_status == 'completed' else '📌 Buyurtma holati yangilandi.'}"
                    )
                    _req.post(
                        f"https://api.telegram.org/bot{bot_token}/sendMessage",
                        json={'chat_id': order.customer.telegram_chat_id, 'text': tg_push, 'parse_mode': 'Markdown'},
                        timeout=5
                    )
                except Exception as tg_err:
                    print(f"Order status TG push error: {tg_err}")

            return JsonResponse({'status': 'success', 'message': f"Buyurtma statusi '{status_text}' holatiga o'zgartirildi! Mijozga Telegram xabar yuborildi."})
        except B2BOrder.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': "Buyurtma topilmadi!"}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': "Noto'g'ri so'rov usuli!"}, status=405)


@csrf_exempt
@staff_required
@transaction.atomic
def create_ai_draft_order(request):
    if request.method == 'POST':
        try:
            body = json.loads(request.body)
            customer_name = body.get('customer_name', 'The Steakhouse').strip()
            product_name = body.get('product_name', 'Striploin').strip()
            weight_val = body.get('weight', 40.0)
            
            weight = Decimal(str(weight_val))
            
            from django.db.models import Q
            customer = Customer.objects.filter(Q(first_name__icontains=customer_name) | Q(last_name__icontains=customer_name)).first()
            if not customer:
                customer = Customer.objects.create(
                    first_name=customer_name,
                    last_name="B2B Client",
                    phone="+998991234567",
                    custom_id="B2B-STEAK",
                    debt_amount=Decimal('0.00'),
                    bonus_points=0
                )
                
            product = Product.objects.filter(name__icontains=product_name, is_active=True).first()
            if not product:
                product = Product.objects.create(
                    name=product_name,
                    price=Decimal('140000.00'),
                    is_active=True
                )
                
            from pos.models import B2BOrder, CustomerLog
            order = B2BOrder.objects.create(
                customer=customer,
                product=product,
                requested_weight=weight,
                status='pending',
                notes="AI Butcher tomonidan tavsiya etilgan loyiha"
            )
            
            msg_text = f"🤖 **AI BUTCHER TAVSIYASI**: {product.name} ({weight} kg) buyurtma loyihasi tasdiq uchun yaratildi."
            CustomerLog.objects.create(
                customer=customer,
                log_type='bonus',
                title="Do'kon xabari",
                message=msg_text,
                amount=Decimal('0.00')
            )
            
            try:
                from .views_api import send_telegram_notification
                tg_msg = (
                    f"🤖 **AI BUTCHER TAVSIYA ETILGAN LOYIHA**\n"
                    f"📦 Buyurtma MFP{order.id}\n"
                    f"👤 Mijoz: {customer.first_name} (ID: {customer.custom_id})\n"
                    f"🥩 Mahsulot: {product.name}\n"
                    f"⚖️ Og'irlik: {weight} kg\n\n"
                    f"Tizimda loyiha tasdiqlandi."
                )
                send_telegram_notification(tg_msg)
            except Exception as tg_err:
                print(f"Telegram notify error in create_ai_draft_order: {tg_err}")
                
            return JsonResponse({
                'status': 'success',
                'message': f"Loyiha yaratildi: MFP{order.id} ({customer.first_name} uchun {weight} kg {product.name})",
                'order_id': order.id
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': "Noto'g'ri so'rov usuli!"}, status=405)


@staff_required
def get_customer_b2b_orders(request, customer_id):
    try:
        from pos.models import B2BOrder
        orders = B2BOrder.objects.filter(customer_id=customer_id).order_by('-created_at')
        orders_data = []
        for o in orders:
            orders_data.append({
                'id': o.id,
                'product_name': o.product.name,
                'requested_weight': float(o.requested_weight),
                'notes': o.notes or '',
                'status': o.status,
                'status_display': o.get_status_display(),
                'payment_proof_url': o.payment_proof_image.url if o.payment_proof_image else None,
                'created_at': timezone.localtime(o.created_at).strftime('%d.%m.%Y %H:%M')
            })
        return JsonResponse(orders_data, safe=False)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@staff_required
def get_pending_b2b_orders_count(request):
    try:
        from pos.models import B2BOrder
        pending_orders = B2BOrder.objects.filter(status='pending').order_by('-created_at')
        count = pending_orders.count()
        
        latest_customer_name = ""
        latest_customer_id = ""
        if count > 0:
            latest_order = pending_orders.first()
            latest_customer_name = f"{latest_order.customer.first_name} {latest_order.customer.last_name or ''}".strip()
            latest_customer_id = latest_order.customer.id
            
        return JsonResponse({
            'status': 'success',
            'count': count,
            'latest_customer_name': latest_customer_name,
            'latest_customer_id': latest_customer_id
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

# =====================================================================
# QARZ KO'CHIRISH VA MIGRATSIYA
# =====================================================================
@staff_required
def debt_migration_page(request):
    notebooks = Notebook.objects.all().order_by('-id')
    
    # Build enriched notebook data with customer and supplier counts/debts
    notebook_data = []
    total_migrated = 0
    total_debt = Decimal('0.00')
    total_migrated_suppliers = 0
    total_supplier_debt = Decimal('0.00')
    
    for nb in notebooks:
        customers = Customer.objects.filter(note__icontains=f"Daftar: {nb.name}")
        c_count = customers.count()
        nb_debt = customers.aggregate(total=Sum('debt_amount'))['total'] or Decimal('0.00')
        total_migrated += c_count
        total_debt += nb_debt
        
        from .models import Supplier
        suppliers = Supplier.objects.filter(note__icontains=f"Daftar: {nb.name}")
        s_count = suppliers.count()
        nb_sup_debt = suppliers.aggregate(total=Sum('our_debt'))['total'] or Decimal('0.00')
        total_migrated_suppliers += s_count
        total_supplier_debt += nb_sup_debt
        
        notebook_data.append({
            'id': nb.id,
            'name': nb.name,
            'created_at': nb.created_at,
            'customer_count': c_count,
            'total_debt': nb_debt,
            'total_debt_display': f"{int(nb_debt):,}".replace(",", " ") + " so'm",
            'supplier_count': s_count,
            'total_supplier_debt': nb_sup_debt,
            'total_supplier_debt_display': f"{int(nb_sup_debt):,}".replace(",", " ") + " so'm",
        })
    
    context = {
        'notebooks': notebooks,
        'notebook_data': notebook_data,
        'total_migrated': total_migrated,
        'total_debt': total_debt,
        'total_debt_display': f"{int(total_debt):,}".replace(",", " ") + " so'm",
        'total_migrated_suppliers': total_migrated_suppliers,
        'total_supplier_debt': total_supplier_debt,
        'total_supplier_debt_display': f"{int(total_supplier_debt):,}".replace(",", " ") + " so'm",
    }
    return render(request, 'debt_notebook.html', context)

@csrf_exempt
@staff_required
@transaction.atomic
def create_notebook(request):
    if request.method == 'POST':
        try:
            body = json.loads(request.body)
            name = body.get('name', '').strip()
            if not name:
                return JsonResponse({'status': 'error', 'message': "Daftar nomi kiritilmadi!"}, status=400)
            
            nb = Notebook.objects.create(name=name)
            return JsonResponse({
                'status': 'success',
                'notebook': {
                    'id':   nb.id,
                    'name': nb.name
                }
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': "Faqat POST so'rovlar"}, status=405)

@staff_required
def get_notebook_customers(request, notebook_id):
    nb = get_object_or_404(Notebook, id=notebook_id)
    record_type = request.GET.get('type', 'customer')
    
    if record_type == 'supplier':
        from .models import Supplier
        suppliers = Supplier.objects.filter(note__icontains=f"Daftar: {nb.name}")
        data = [{
            'id':           s.id,
            'first_name':   s.first_name,
            'last_name':    s.last_name or '',
            'phone':        s.phone,
            'custom_id':    s.custom_id,
            'debt_balance': float(s.our_debt)
        } for s in suppliers]
    else:
        customers = Customer.objects.filter(note__icontains=f"Daftar: {nb.name}")
        data = [{
            'id':           c.id,
            'first_name':   c.first_name,
            'last_name':    c.last_name or '',
            'phone':        c.phone,
            'custom_id':    c.custom_id,
            'debt_balance': float(c.debt_amount)
        } for c in customers]
        
    return JsonResponse(data, safe=False)

@csrf_exempt
@staff_required
@transaction.atomic
def save_migrated_debt(request):
    if request.method == 'POST':
        try:
            body = json.loads(request.body)
            record_type  = body.get('type', 'customer')
            first_name   = body.get('first_name', '').strip()
            last_name    = body.get('last_name', '').strip()
            phone        = body.get('phone', '').strip()
            custom_id    = body.get('custom_id', '').strip()
            debt_balance = Decimal(str(body.get('debt_balance', 0)))
            notebook_name = body.get('notebook_name', '').strip()

            if not first_name or not phone or not custom_id:
                return JsonResponse({'status': 'error', 'message': "Ism, telefon va ID majburiy!"}, status=400)

            if record_type == 'supplier':
                from .models import Supplier
                supplier, created = Supplier.objects.get_or_create(
                    phone=phone,
                    defaults={
                        'first_name':  first_name,
                        'last_name':   last_name,
                        'custom_id':   custom_id,
                        'our_debt':    debt_balance,
                        'note':        f"Daftar: {notebook_name}" if notebook_name else ""
                    }
                )
                if not created:
                    supplier.our_debt += debt_balance
                    if notebook_name and f"Daftar: {notebook_name}" not in (supplier.note or ''):
                        supplier.note = f"{supplier.note or ''} | Daftar: {notebook_name}".strip(" |")
                    supplier.save()
            else:
                customer, created = Customer.objects.get_or_create(
                    phone=phone,
                    defaults={
                        'first_name':  first_name,
                        'last_name':   last_name,
                        'custom_id':   custom_id,
                        'debt_amount': debt_balance,
                        'note':        f"Daftar: {notebook_name}" if notebook_name else ""
                    }
                )

                if created:
                    create_user_for_customer(customer)
                else:
                    customer.debt_amount += debt_balance
                    if notebook_name and f"Daftar: {notebook_name}" not in customer.note:
                        customer.note = f"{customer.note} | Daftar: {notebook_name}".strip(" |")
                    customer.save()

                if debt_balance > 0:
                    CustomerLog.objects.create(
                        customer=customer,
                        log_type='debt_add',
                        title="Qarz migratsiya qilindi",
                        message=f"Migratsiya orqali '{notebook_name}' daftaridan {debt_balance} so'm qarz ko'chirildi.",
                        amount=debt_balance
                    )

            return JsonResponse({'status': 'success', 'message': "Qarz muvaffaqiyatli saqlandi!"})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': "Faqat POST so'rovlar"}, status=405)

@staff_required
def debt_payment_view(request):
    customers = Customer.objects.filter(debt_amount__gt=0).order_by('-debt_amount')
    return render(request, 'debt_payment.html', {'customers': customers})

@csrf_exempt
@staff_required
@transaction.atomic
def process_debt_payment(request):
    if request.method == 'POST':
        try:
            body = json.loads(request.body)
            customer_id = body.get('customer_id')
            amount      = Decimal(str(body.get('amount', 0)))
            comment     = body.get('comment', '').strip()

            if not customer_id or amount <= 0:
                return JsonResponse({'status': 'error', 'message': "Mijoz va to'lov summasi noto'g'ri!"}, status=400)

            customer = Customer.objects.get(id=customer_id)
            if amount > customer.debt_amount:
                return JsonResponse({'status': 'error', 'message': "To'lov summasi qarzdan katta bo'lishi mumkin emas!"}, status=400)

            customer.debt_amount -= amount
            customer.save()

            # Trigger Telegram Notification for debt payment
            try:
                from .views_api import send_telegram_notification
                comment_text = comment if comment else "Yo'q"
                msg = (
                    f"✅ *Baxmal Meat — Qarz To'landi (Kassa)*\n\n"
                    f"👤 *Xaridor:* {customer.first_name} {customer.last_name or ''}\n"
                    f"🆔 *Mijoz ID:* `{customer.custom_id}`\n\n"
                    f"💰 *To'langan summa:* `{amount.quantize(Decimal('1')):,}` so'm\n"
                    f"💳 *Qolgan qarz:* `{customer.debt_amount.quantize(Decimal('1')):,}` so'm\n"
                    f"📝 *Izoh:* {comment_text}"
                )
                send_telegram_notification(msg)

                # Direct notification to Customer's Bot
                if customer.telegram_chat_id:
                    from .customer_bot import send_message as send_cust_msg
                    cust_msg = (
                        f"🧾 *BAXMAL MEAT — NASIYA TO'LOVI CHEKI*\n"
                        f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                        f"👤 *Mijoz:* {customer.first_name}\n"
                        f"💵 *To'langan Summa:* `{amount:,.0f}` so'm\n"
                        f"📉 *Qolgan Qarz:* `{customer.debt_amount:,.0f}` so'm\n\n"
                        f"✅ Rahmat! Qarz to'lovingiz kassa hisobiga qabul qilindi."
                    )
                    send_cust_msg(customer.telegram_chat_id, cust_msg)
            except Exception as tg_err:
                print(f"Telegram notify error in process_debt_payment: {tg_err}")

            # Create a CashTransaction for this debt payment
            from .models import CashTransaction
            pm_method = body.get('payment_method', 'naqd')
            CashTransaction.objects.create(
                transaction_type='in',
                amount=amount,
                category='debt_pay',
                payment_method=pm_method,
                description=f"Mijoz {customer.first_name} {customer.last_name or ''} qarz to'lovi. Izoh: {comment}",
                created_by=request.user,
                customer=customer
            )
            # Also create customer log for personal cabinet payment history
            CustomerLog.objects.create(
                customer=customer,
                log_type='debt_pay',
                title="Qarz to'landi (Kassa)",
                message=f"+{amount:,} so'm qarz to'landi. Izoh: {comment}",
                amount=amount
            )

            is_fully_paid = (customer.debt_amount == 0)

            return JsonResponse({
                'status':   'success',
                'message':  "Qarz to'lovi muvaffaqiyatli qabul qilindi!",
                'remaining_debt': float(customer.debt_amount),
                'paid': float(amount),
                'is_fully_paid': is_fully_paid
            })
        except Customer.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': "Mijoz topilmadi!"}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': "Faqat POST so'rovlar"}, status=405)


@staff_required
def cash_flow_view(request):
    from .models import CashTransaction, Supplier
    
    # Calculate Naqd balance
    naqd_sales = Sale.objects.filter(payment_method='naqd').aggregate(s=Sum('final_paid'))['s'] or Decimal('0.00')
    naqd_cash_in = CashTransaction.objects.filter(transaction_type='in', payment_method='naqd').aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
    naqd_cash_out = CashTransaction.objects.filter(transaction_type='out', payment_method='naqd').aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
    cash_in_hand = naqd_sales + naqd_cash_in - naqd_cash_out
    
    # Calculate Karta balance
    karta_sales = Sale.objects.filter(payment_method='karta').aggregate(s=Sum('final_paid'))['s'] or Decimal('0.00')
    karta_cash_in = CashTransaction.objects.filter(transaction_type='in', payment_method='karta').aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
    karta_cash_out = CashTransaction.objects.filter(transaction_type='out', payment_method='karta').aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
    card_in_hand = karta_sales + karta_cash_in - karta_cash_out

    # Get recent 30 cash transactions
    transactions = CashTransaction.objects.select_related('created_by', 'customer', 'supplier').order_by('-created_at')[:30]
    
    customers = Customer.objects.all().order_by('first_name')
    suppliers = Supplier.objects.all().order_by('first_name')

    context = {
        'cash_in_hand': cash_in_hand,
        'card_in_hand': card_in_hand,
        'transactions': transactions,
        'customers': customers,
        'suppliers': suppliers,
        'categories': CashTransaction.CATEGORIES,
        'payment_methods': CashTransaction.PAYMENT_METHODS,
    }
    return render(request, 'pos/cash_flow.html', context)


@csrf_exempt
@staff_required
@transaction.atomic
def process_cash_transaction(request):
    if request.method == 'POST':
        from .models import CashTransaction, Customer, Supplier
        try:
            body = json.loads(request.body)
            t_type = body.get('transaction_type')
            amount = Decimal(str(body.get('amount', 0)))
            category = body.get('category', 'other')
            p_method = body.get('payment_method', 'naqd')
            desc = body.get('description', '').strip()
            
            customer_id = body.get('customer_id')
            supplier_id = body.get('supplier_id')
            
            customer = None
            if customer_id:
                customer = Customer.objects.get(id=customer_id)
                
            supplier = None
            if supplier_id:
                supplier = Supplier.objects.get(id=supplier_id)

            if t_type not in ['in', 'out'] or amount <= 0:
                return JsonResponse({'status': 'error', 'message': "Noto'g'ri tur yoki miqdor!"}, status=400)

            # If it's a cash withdrawal (chiqim), check balance
            if t_type == 'out':
                if p_method == 'naqd':
                    naqd_sales = Sale.objects.filter(payment_method='naqd').aggregate(s=Sum('final_paid'))['s'] or Decimal('0.00')
                    naqd_in = CashTransaction.objects.filter(transaction_type='in', payment_method='naqd').aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
                    naqd_out = CashTransaction.objects.filter(transaction_type='out', payment_method='naqd').aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
                    current_bal = naqd_sales + naqd_in - naqd_out
                else:
                    karta_sales = Sale.objects.filter(payment_method='karta').aggregate(s=Sum('final_paid'))['s'] or Decimal('0.00')
                    karta_in = CashTransaction.objects.filter(transaction_type='in', payment_method='karta').aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
                    karta_out = CashTransaction.objects.filter(transaction_type='out', payment_method='karta').aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
                    current_bal = karta_sales + karta_in - karta_out
                
                if amount > current_bal:
                    return JsonResponse({'status': 'error', 'message': "Kassada buncha mablag' yo'q!"}, status=400)

            # Create Transaction
            transaction = CashTransaction.objects.create(
                transaction_type=t_type,
                amount=amount,
                category=category,
                payment_method=p_method,
                description=desc,
                created_by=request.user,
                customer=customer,
                supplier=supplier
            )

            # Update customer/supplier debt
            if customer:
                if t_type == 'in':
                    customer.debt_amount -= amount
                    customer.save()
                    CustomerLog.objects.create(
                        customer=customer,
                        log_type='debt_pay',
                        title="Kirim (Kassa orqali)",
                        message=f"+{amount:,} so'm qabul qilindi. Izoh: {desc}",
                        amount=amount
                    )
                elif t_type == 'out':
                    customer.debt_amount += amount
                    customer.save()
                    CustomerLog.objects.create(
                        customer=customer,
                        log_type='debt_add',
                        title="Chiqim (Kassa orqali)",
                        message=f"-{amount:,} so'm berildi. Izoh: {desc}",
                        amount=amount
                    )

            if supplier:
                if t_type == 'out':
                    supplier.our_debt -= amount
                    supplier.save()
                elif t_type == 'in':
                    supplier.our_debt += amount
                    supplier.save()

            # Trigger Telegram Notification for cash flow
            try:
                from .views_api import send_telegram_notification
                t_type_lbl = "📥 KIRIM (Cash In)" if t_type == 'in' else "📤 CHIQIM (Cash Out)"
                p_method_lbl = "💵 Naqd" if p_method == 'naqd' else "💳 Plastik karta"
                assoc_lbl = ""
                if customer:
                    assoc_lbl = f"\n👤 *Mijoz:* {customer.first_name} {customer.last_name or ''}"
                desc_text = desc if desc else "Yo'q"
                msg = f"💸 *Baxmal Meat — Kassa tranzaksiyasi*\n\n⚙️ *Turi:* {t_type_lbl}\n📂 *Kategoriya:* {transaction.get_category_display()}\n💳 *To'lov usuli:* {p_method_lbl}{assoc_lbl}\n\n💰 *Summa:* {amount.quantize(Decimal('1')):,} so'm\n👤 *Mas'ul:* {request.user.get_full_name() or request.user.username}\n📝 *Izoh:* {desc_text}"
                send_telegram_notification(msg)
            except Exception as tg_err:
                print(f"Telegram notify error in process_cash_transaction: {tg_err}")

            return JsonResponse({
                'status': 'success',
                'message': "Tranzaksiya muvaffaqiyatli saqlandi!"
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': "Faqat POST so'rovlar"}, status=405)

# =====================================================================
# MIJOZLAR RO'YXATI
# =====================================================================
@staff_required
def customers_view(request):
    if request.method == 'POST':
        first_name  = request.POST.get('first_name', '').strip()
        last_name   = request.POST.get('last_name', '').strip()
        phone       = request.POST.get('phone', '').strip()
        customer_id = request.POST.get('customer_id', '').strip()
        debt_amount = request.POST.get('debt_amount', '0').strip() or '0'
        notebook_id = request.POST.get('notebook_id')
        note        = request.POST.get('note', '').strip()

        if first_name and phone and customer_id:
            try:
                if Customer.objects.filter(custom_id=customer_id).exists():
                    from django.contrib import messages
                    messages.error(request, f"Xatolik: '{customer_id}' ID raqami allaqachon mavjud!")
                    return redirect('/pos/customers/')

                customer = Customer.objects.create(
                    first_name=first_name,
                    last_name=last_name,
                    phone=phone,
                    custom_id=customer_id,
                    debt_amount=Decimal(debt_amount),
                    note=note
                )

                if notebook_id:
                    nb = Notebook.objects.filter(id=notebook_id).first()
                    if nb:
                        customer.note = f"Daftar: {nb.name} | {note}"
                        customer.save()

                if request.FILES.get('image'):
                    customer.image = request.FILES.get('image')
                    customer.save()

                if Decimal(debt_amount) > 0:
                    CustomerLog.objects.create(
                        customer=customer,
                        log_type='debt_add',
                        title="Boshlang'ich qarz kiritildi",
                        message=f"Mijoz yaratilganda boshlang'ich qarz balansi {debt_amount} so'm deb kiritildi.",
                        amount=Decimal(debt_amount)
                    )
                create_user_for_customer(customer)
            except Exception as e:
                print(f"Error saving customer: {e}")

    filter_type  = request.GET.get('filter_type', 'all')
    notebook_id  = request.GET.get('notebook_id')
    search_query = request.GET.get('search', '').strip()

    customers = Customer.objects.all().order_by('-id')

    if search_query:
        customers = customers.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(custom_id__icontains=search_query)
        )

    if filter_type == 'notebook' and notebook_id:
        nb = Notebook.objects.filter(id=notebook_id).first()
        if nb:
            customers = customers.filter(note__icontains=f"Daftar: {nb.name}")
    elif filter_type == 'debt':
        customers = customers.filter(debt_amount__gt=0)
    elif filter_type == 'our_debt':
        customers = customers.filter(debt_amount__lt=0)
    elif filter_type == 'clean':
        customers = customers.filter(debt_amount=0)
    elif filter_type == 'long_term':
        one_month_ago = timezone.now() - timedelta(days=30)
        customers = customers.filter(debt_amount__gt=0, logs__created_at__lte=one_month_ago).distinct()
    elif filter_type == 'vip':
        customers = customers.filter(bonus_points__gt=5000)

    total_customers = Customer.objects.count()
    debt_customers_count = Customer.objects.filter(debt_amount__gt=0).count()
    our_debt_customers_count = Customer.objects.filter(debt_amount__lt=0).count()
    total_debt = Customer.objects.filter(debt_amount__gt=0).aggregate(t=Sum('debt_amount'))['t'] or 0
    total_our_debt = abs(Customer.objects.filter(debt_amount__lt=0).aggregate(t=Sum('debt_amount'))['t'] or 0)
    total_bonus = Customer.objects.aggregate(t=Sum('bonus_points'))['t'] or 0

    notebooks = Notebook.objects.all()

    top_debtors = Customer.objects.filter(debt_amount__gt=0).order_by('-debt_amount')[:5]

    context = {
        'customers':            customers,
        'total_customers':      total_customers,
        'debt_customers_count': debt_customers_count,
        'our_debt_customers_count': our_debt_customers_count,
        'total_debt':           total_debt,
        'total_our_debt':       total_our_debt,
        'total_bonus':          total_bonus,
        'notebooks':            notebooks,
        'current_filter':       filter_type,
        'current_notebook':     notebook_id,
        'top_debtors':          top_debtors,
    }
    return render(request, 'customers.html', context)

@staff_required
def delete_notebook_view(request, notebook_id):
    nb = get_object_or_404(Notebook, id=notebook_id)
    nb.delete()
    return redirect('debt-migration')

def switch_script_view(request, script_mode):
    if script_mode in ['latin', 'cyrillic']:
        request.session['script_mode'] = script_mode
    referer = request.META.get('HTTP_REFERER', '/')
    return redirect(referer)

# =====================================================================
# KENGAYTIRILGAN ANALITIKA VA REPORTLAR
# =====================================================================
@staff_required
def global_analytics(request):
    sync_existing_customers_to_users()
    
    # 1. Read filter query params
    search_query = request.GET.get('search_query', '').strip()
    payment_method = request.GET.get('payment_method', '').strip()
    start_date_str = request.GET.get('start_date', '').strip()
    end_date_str = request.GET.get('end_date', '').strip()
    
    # 2. Base queryset
    sales = Sale.objects.all().select_related('customer').prefetch_related('items__product').order_by('-created_at')
    
    # Apply search filter
    if search_query:
        if search_query.isdigit():
            sales = sales.filter(Q(id=int(search_query)) | Q(customer__phone__icontains=search_query))
        else:
            sales = sales.filter(
                Q(customer__first_name__icontains=search_query) |
                Q(customer__last_name__icontains=search_query) |
                Q(customer__custom_id__icontains=search_query)
            )
            
    # Apply payment method filter
    if payment_method:
        sales = sales.filter(payment_method=payment_method)
        
    # Apply date filters
    if start_date_str:
        try:
            start_date = timezone.datetime.strptime(start_date_str, '%Y-%m-%d').date()
            start_dt = timezone.make_aware(timezone.datetime.combine(start_date, timezone.datetime.min.time()))
            sales = sales.filter(created_at__gte=start_dt)
        except ValueError:
            pass
            
    if end_date_str:
        try:
            end_date = timezone.datetime.strptime(end_date_str, '%Y-%m-%d').date()
            end_dt = timezone.make_aware(timezone.datetime.combine(end_date, timezone.datetime.max.time()))
            sales = sales.filter(created_at__lte=end_dt)
        except ValueError:
            pass
            
    # 3. Calculate breakdown totals using database aggregation
    from django.db.models import Sum, Max, Avg, Count
    from django.db.models.functions import Coalesce
    from django.core.paginator import Paginator
    from pos.models import SaleItem

    totals = sales.aggregate(
        naqd=Coalesce(Sum('final_paid', filter=Q(payment_method='naqd')), Decimal('0.00')),
        karta=Coalesce(Sum('final_paid', filter=Q(payment_method='karta')), Decimal('0.00')),
        qr=Coalesce(Sum('final_paid', filter=Q(payment_method='qr')), Decimal('0.00')),
        nasiya=Coalesce(Sum('debt_added', filter=Q(payment_method='nasiya')), Decimal('0.00')),
        avg_check=Coalesce(Avg('total_amount'), Decimal('0.00')),
        max_sale=Coalesce(Max('total_amount'), Decimal('0.00')),
        total_count=Count('id')
    )
    
    naqd_sum = totals['naqd']
    karta_sum = totals['karta']
    qr_sum = totals['qr']
    jami_nasiya = totals['nasiya']
    avg_check = totals['avg_check']
    max_sale = totals['max_sale']
    sales_count = totals['total_count']
    
    jami_kirim = naqd_sum + karta_sum + qr_sum
    karta_qr_summa = karta_sum + qr_sum
    total_sum = naqd_sum + karta_sum + qr_sum + jami_nasiya
    
    if total_sum > 0:
        naqd_pct = int(round(naqd_sum / total_sum * 100))
        karta_pct = int(round(karta_sum / total_sum * 100))
        qr_pct = int(round(qr_sum / total_sum * 100))
        nasiya_pct = int(round(jami_nasiya / total_sum * 100))
    else:
        naqd_pct = karta_pct = qr_pct = nasiya_pct = 0
        
    # 4. Group sales by local hour range for peak hour analytics (using value dictionaries for fast iteration)
    sales_values = sales.values('created_at', 'final_paid', 'debt_added')
    hourly_data = {
        '08:00-10:00': Decimal('0.00'),
        '10:00-12:00': Decimal('0.00'),
        '12:00-14:00': Decimal('0.00'),
        '14:00-16:00': Decimal('0.00'),
        '16:00-18:00': Decimal('0.00'),
        '18:00-20:00': Decimal('0.00'),
        '20:00-22:00': Decimal('0.00'),
        'Tungi vaqt': Decimal('0.00')
    }
    
    for s_val in sales_values:
        local_dt = timezone.localtime(s_val['created_at'])
        h = local_dt.hour
        val = s_val['final_paid'] + s_val['debt_added']
        if 8 <= h < 10:
            hourly_data['08:00-10:00'] += val
        elif 10 <= h < 12:
            hourly_data['10:00-12:00'] += val
        elif 12 <= h < 14:
            hourly_data['12:00-14:00'] += val
        elif 14 <= h < 16:
            hourly_data['14:00-16:00'] += val
        elif 16 <= h < 18:
            hourly_data['16:00-18:00'] += val
        elif 18 <= h < 20:
            hourly_data['18:00-20:00'] += val
        elif 20 <= h < 22:
            hourly_data['20:00-22:00'] += val
        else:
            hourly_data['Tungi vaqt'] += val

    # Calculate product-based breakdown in database using annotate
    item_aggregates = SaleItem.objects.filter(sale__in=sales).values('product__name').annotate(
        total_weight=Sum('weight'),
        total_revenue=Sum('item_total')
    ).order_by('-total_revenue')

    product_breakdown = []
    total_product_revenue = Decimal('0.00')
    for item in item_aggregates:
        p_name = item['product__name']
        p_weight = item['total_weight'] or Decimal('0.000')
        p_rev = item['total_revenue'] or Decimal('0.00')
        total_product_revenue += p_rev
        product_breakdown.append({
            'name': p_name,
            'weight': p_weight,
            'revenue': p_rev
        })

    for p in product_breakdown:
        if total_product_revenue > 0:
            p['pct'] = int(round(p['revenue'] / total_product_revenue * 100))
        else:
            p['pct'] = 0

    # 5. Paginate sales list for template rendering (50 per page)
    paginator = Paginator(sales, 50)
    page_number = request.GET.get('page', '1')
    page_obj = paginator.get_page(page_number)

    # Group ONLY the page sales by date
    from collections import OrderedDict
    grouped_sales = OrderedDict()
    
    for s in page_obj:
        local_dt = timezone.localtime(s.created_at)
        date_str = local_dt.strftime('%d.%m.%Y')
        if date_str not in grouped_sales:
            grouped_sales[date_str] = {
                'day_total': Decimal('0.00'),
                'sales_list': []
            }
        grouped_sales[date_str]['day_total'] += (s.final_paid + s.debt_added)
        grouped_sales[date_str]['sales_list'].append(s)
        
    notebooks = Notebook.objects.all().order_by('-id')
    
    # Calculate our debts (Suppliers + negative Customer balances)
    total_supplier_debts = Supplier.objects.aggregate(s=Sum('our_debt'))['s'] or Decimal('0.00')
    total_customer_negative_debts = abs(Customer.objects.filter(debt_amount__lt=0).aggregate(s=Sum('debt_amount'))['s'] or Decimal('0.00'))
    total_our_debts = total_supplier_debts + total_customer_negative_debts
    
    context = {
        'grouped_sales':        grouped_sales,
        'jami_kirim':           jami_kirim,
        'karta_qr_summa':       karta_qr_summa,
        'jami_nasiya':          jami_nasiya,
        'naqd_sum':             naqd_sum,
        'karta_sum':            karta_sum,
        'qr_sum':               qr_sum,
        'naqd_pct':             naqd_pct,
        'karta_pct':            karta_pct,
        'qr_pct':               qr_pct,
        'nasiya_pct':           nasiya_pct,
        'notebooks':            notebooks,
        'search_query':         search_query,
        'payment_method':       payment_method,
        'start_date':           start_date_str,
        'end_date':             end_date_str,
        'total_sales':          sales_count,
        'hourly_data':          hourly_data,
        'avg_check':            avg_check,
        'max_sale':             max_sale,
        'product_breakdown':    product_breakdown,
        'total_sum':            total_sum,
        'page_obj':             page_obj,
        'total_our_debts':      total_our_debts,
    }
    return render(request, 'dashboard_full.html', context)

@staff_required
def export_analytics_excel(request):
    search_query = request.GET.get('search_query', '').strip()
    payment_method = request.GET.get('payment_method', '').strip()
    start_date_str = request.GET.get('start_date', '').strip()
    end_date_str = request.GET.get('end_date', '').strip()
    
    sales = Sale.objects.all().select_related('customer').prefetch_related('items__product').order_by('-created_at')
    
    if search_query:
        if search_query.isdigit():
            sales = sales.filter(Q(id=int(search_query)) | Q(customer__phone__icontains=search_query))
        else:
            sales = sales.filter(
                Q(customer__first_name__icontains=search_query) |
                Q(customer__last_name__icontains=search_query) |
                Q(customer__custom_id__icontains=search_query)
            )
            
    if payment_method:
        sales = sales.filter(payment_method=payment_method)
        
    if start_date_str:
        try:
            start_date = timezone.datetime.strptime(start_date_str, '%Y-%m-%d').date()
            start_dt = timezone.make_aware(timezone.datetime.combine(start_date, timezone.datetime.min.time()))
            sales = sales.filter(created_at__gte=start_dt)
        except ValueError:
            pass
            
    if end_date_str:
        try:
            end_date = timezone.datetime.strptime(end_date_str, '%Y-%m-%d').date()
            end_dt = timezone.make_aware(timezone.datetime.combine(end_date, timezone.datetime.max.time()))
            sales = sales.filter(created_at__lte=end_dt)
        except ValueError:
            pass

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="baxmal_meat_savdolar.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Savdolar Hisoboti"

    headers = [
        "Chek #", "Mijoz ismi", "Vaqti", "Jami Summa", "To'lov turi", 
        "Chegirma", "Yechilgan bonus", "Qarz summa", "To'langan naqd/karta", 
        "Sotib olingan mahsulotlar"
    ]
    ws.append(headers)

    for sale in sales:
        cust_name = f"{sale.customer.first_name} {sale.customer.last_name or ''}".strip() if sale.customer else "Mijozsiz xarid"
        # Sotib olingan mahsulotlarni matn ko'rinishida yig'ish
        items_str = ", ".join([f"{item.product.name} ({float(item.weight):.3f} kg)" for item in sale.items.all()])
        
        ws.append([
            sale.id,
            cust_name,
            timezone.localtime(sale.created_at).strftime('%Y-%m-%d %H:%M'),
            float(sale.total_amount),
            sale.get_payment_method_display(),
            float(sale.discount_amount),
            float(sale.bonus_used),
            float(sale.debt_added),
            float(sale.final_paid),
            items_str
        ])

    style_excel_worksheet(ws, len(headers))
    wb.save(response)
    return response

@staff_required
def export_slaughters_excel(request):
    from .models import Slaughter
    slaughters = Slaughter.objects.all().select_related('supplier').order_by('-created_at')
    
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="baxmal_meat_xaridlar.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "So'yimlar va Xaridlar"

    headers = [
        "Xarid #", "Ta'minotchi ismi", "Telefon", "Hayvon turi", 
        "Jami vazn (kg)", "Qolgan vazn (kg)", "Xarid narxi (kg/so'm)", 
        "Jami xarid summasi", "To'landimi?", "To'lov muddati", "Sana/Vaqt"
    ]
    ws.append(headers)

    for s in slaughters:
        supp_name = f"{s.supplier.first_name} {s.supplier.last_name or ''}".strip() if s.supplier else "Noma'lum ta'minotchi"
        phone = s.supplier.phone if s.supplier else ""
        ws.append([
            s.id,
            supp_name,
            phone,
            s.get_animal_type_display(),
            float(s.total_weight),
            float(s.remaining_weight),
            float(s.purchase_price_per_kg),
            float(s.total_cost),
            "Ha" if s.is_paid else "Yo'q",
            s.due_date.strftime('%Y-%m-%d') if s.due_date else "",
            timezone.localtime(s.created_at).strftime('%Y-%m-%d %H:%M')
        ])

    style_excel_worksheet(ws, len(headers))
    wb.save(response)
    return response

@staff_required
def yield_loss_view(request):
    suppliers_qs = Supplier.objects.all().order_by('-id')
    customers_qs = Customer.objects.all()[:30]
    
    combined_suppliers = []
    for s in suppliers_qs:
        combined_suppliers.append({
            'id': f"supplier_{s.id}",
            'name': f"Ta'minotchi: {s.first_name} {s.last_name or ''}".strip(),
            'phone': s.phone,
            'custom_id': s.custom_id,
            'our_debt': float(s.our_debt),
            'debt_display': f"{s.our_debt:,.0f} so'm",
            'is_negative': False
        })
    for c in customers_qs:
        our_debt = -c.debt_amount
        combined_suppliers.append({
            'id': f"customer_{c.id}",
            'name': f"Mijoz: {c.first_name} {c.last_name or ''}".strip(),
            'phone': c.phone,
            'custom_id': c.custom_id,
            'our_debt': float(our_debt),
            'debt_display': f"Bizga qarz: {abs(our_debt):,.0f} so'm" if our_debt < 0 else f"{our_debt:,.0f} so'm",
            'is_negative': our_debt < 0
        })
        
    products = Product.objects.filter(is_active=True)
    return render(request, 'pos/yield_loss.html', {'suppliers': combined_suppliers, 'products': products})

@csrf_exempt
@staff_required
@transaction.atomic
def process_slaughter_api(request):
    if request.method == 'POST':
        try:
            # Check content type
            if request.content_type and 'multipart/form-data' in request.content_type:
                body = request.POST
            else:
                try:
                    body = json.loads(request.body)
                except Exception:
                    body = request.POST
            
            def get_bool(key):
                val = body.get(key)
                if isinstance(val, bool):
                    return val
                return str(val).lower() == 'true'
            
            supplier_id = body.get('supplier_id')
            if supplier_id == 'null' or supplier_id == '' or supplier_id is None:
                supplier_id = None
                
            animal_type = body.get('animal_type') # 'mol' or 'qoy'
            total_weight = Decimal(str(body.get('total_weight', 0))) # Toza go'sht vazni
            purchase_price = Decimal(str(body.get('purchase_price', 0))) # Xarid narxi
            due_days = int(body.get('due_days', 21))
            
            # Extra parts details
            jigar_taken = get_bool('jigar_taken')
            jigar_weight = Decimal(str(body.get('jigar_weight', 0)))
            jigar_price = Decimal(str(body.get('jigar_price', 0)))
            
            yurak_taken = get_bool('yurak_taken')
            yurak_weight = Decimal(str(body.get('yurak_weight', 0)))
            yurak_price = Decimal(str(body.get('yurak_price', 0)))
            
            dumgaza_taken = get_bool('dumgaza_taken')
            dumgaza_weight = Decimal(str(body.get('dumgaza_weight', 0)))
            dumgaza_price = Decimal(str(body.get('dumgaza_price', 0)))
            
            charvi_taken = get_bool('charvi_taken')
            charvi_weight = Decimal(str(body.get('charvi_weight', 0)))
            charvi_price = Decimal(str(body.get('charvi_price', 0)))
            
            ilik_taken = get_bool('ilik_taken')
            ilik_weight = Decimal(str(body.get('ilik_weight', 0)))
            ilik_price = Decimal(str(body.get('ilik_price', 0)))
            
            kalla_taken = get_bool('kalla_taken')
            kalla_qty = Decimal(str(body.get('kalla_qty', 0)))
            kalla_price = Decimal(str(body.get('kalla_price', 0)))
            
            lahm_taken = get_bool('lahm_taken')
            lahm_weight = Decimal(str(body.get('lahm_weight', 0)))
            lahm_price = Decimal(str(body.get('lahm_price', 0)))
            
            if not animal_type or total_weight <= 0 or purchase_price <= 0:
                return JsonResponse({'status': 'error', 'message': "Ma'lumotlar to'liq kiritilmadi!"})
            
            supplier = None
            customer = None
            if supplier_id:
                if str(supplier_id).startswith('customer_'):
                    cust_id = int(str(supplier_id).replace('customer_', ''))
                    customer = Customer.objects.get(id=cust_id)
                elif str(supplier_id).startswith('supplier_'):
                    sup_id = int(str(supplier_id).replace('supplier_', ''))
                    supplier = Supplier.objects.get(id=sup_id)
                else:
                    try:
                        supplier = Supplier.objects.get(id=supplier_id)
                    except (ValueError, Supplier.DoesNotExist):
                        pass
                
            # Subtract carcass cuts from total carcass weight (Toza go'sht)
            carcass_cuts_weight = Decimal('0.000')
            if dumgaza_taken:
                carcass_cuts_weight += dumgaza_weight
            if charvi_taken:
                carcass_cuts_weight += charvi_weight
            if ilik_taken:
                carcass_cuts_weight += ilik_weight
            if lahm_taken:
                carcass_cuts_weight += lahm_weight
            
            remaining_carcass_weight = total_weight - carcass_cuts_weight
            if remaining_carcass_weight < 0:
                return JsonResponse({
                    'status': 'error',
                    'message': "Ajratilgan qismlar (Ilik, lahm, moy, dumg'aza) vazni jami toza go'sht vaznidan oshib ketdi!"
                })
                
            # Calculate total cost: clean meat cost + extra parts agreed prices
            clean_meat_cost = total_weight * purchase_price
            butcher_fee = Decimal(str(body.get('butcher_fee', 0)))
            target_selling_price = Decimal(str(body.get('target_selling_price', 140000)))
            
            extra_parts_cost = (
                (jigar_price if jigar_taken else Decimal('0.00')) +
                (yurak_price if yurak_taken else Decimal('0.00')) +
                (dumgaza_price if dumgaza_taken else Decimal('0.00')) +
                (charvi_price if charvi_taken else Decimal('0.00')) +
                (ilik_price if ilik_taken else Decimal('0.00')) +
                (kalla_price if kalla_taken else Decimal('0.00')) +
                (lahm_price if lahm_taken else Decimal('0.00'))
            )
            total_cost = clean_meat_cost + extra_parts_cost
            
            # Real break-even calculation for pure Lahm meat
            offal_retail_revenue = Decimal('0.00')
            if ilik_taken and ilik_weight > 0:
                offal_retail_revenue += ilik_weight * Decimal('60000.00')
            if charvi_taken and charvi_weight > 0:
                offal_retail_revenue += charvi_weight * Decimal('65000.00')
            if dumgaza_taken and dumgaza_weight > 0:
                offal_retail_revenue += dumgaza_weight * Decimal('80000.00')
            if jigar_taken and jigar_weight > 0:
                offal_retail_revenue += jigar_weight * Decimal('70000.00')
            if yurak_taken and yurak_weight > 0:
                offal_retail_revenue += yurak_weight * Decimal('70000.00')
            if kalla_taken and kalla_qty > 0:
                offal_retail_revenue += kalla_qty * Decimal('60000.00')
                
            total_expenditure = clean_meat_cost + extra_parts_cost + butcher_fee
            lahm_target_coverage = max(Decimal('0.00'), total_expenditure - offal_retail_revenue)
            
            if remaining_carcass_weight > Decimal('0.00'):
                real_break_even_cost = (lahm_target_coverage / remaining_carcass_weight).quantize(Decimal('0.01'))
            else:
                real_break_even_cost = purchase_price
            
            # Save Slaughter
            due_date = timezone.now().date() + timezone.timedelta(days=due_days)
            slaughter = Slaughter.objects.create(
                supplier=supplier,
                customer=customer,
                animal_type=animal_type,
                total_weight=total_weight,
                purchase_price_per_kg=purchase_price,
                due_date=due_date
            )
            
            # Record Butcher fee expense if any
            if butcher_fee > 0:
                CashTransaction.objects.create(
                    transaction_type='out',
                    category='expense',
                    amount=butcher_fee,
                    description=f"So'yim #{slaughter.id} qassoblik xizmat haqi"
                )
            
            # Update Supplier/Customer debt
            if supplier:
                supplier.our_debt += total_cost
                supplier.save()
                if supplier.customer:
                    CustomerLog.objects.create(
                        customer=supplier.customer,
                        log_type='slaughter',
                        title="🚜 Go'sht sotib olindi (So'yim)",
                        message=f"So'yim #{slaughter.id}: {slaughter.get_animal_type_display()} ({total_weight} kg) qabul qilindi. Summa: {total_cost:,} so'm.",
                        amount=total_cost
                    )
            elif customer:
                customer.debt_amount -= total_cost
                customer.save()
                
                # Log transaction to customer logs
                CustomerLog.objects.create(
                    customer=customer,
                    log_type='debt_pay',
                    title="Go'sht sotib olindi (So'yim)",
                    message=f"So'yim #{slaughter.id} orqali mijozdan {total_weight} kg toza go'sht sotib olindi. Summa: {total_cost:,} so'm.",
                    amount=total_cost
                )
                
            # 1. Update main meat stock (Mol go'shti / Qo'y go'shti) - with remaining carcass weight
            main_meat_name = "Mol go'shti" if animal_type == 'mol' else "Qo'y go'shti"
            main_product = Product.objects.get(name=main_meat_name)
            main_stock, created = Stock.objects.get_or_create(product=main_product)
            main_stock.quantity += remaining_carcass_weight
            main_stock.save()
            
            # Create StockBatch for yield decay tracking with REAL BREAK-EVEN COST
            from .models import StockBatch
            StockBatch.objects.create(
                product=main_product,
                initial_quantity=remaining_carcass_weight,
                current_quantity=remaining_carcass_weight,
                purchase_price_per_kg=real_break_even_cost
            )
            
            # Send Telegram Alert to store owner
            try:
                from .telegram_bot import send_message, CHAT_ID
                if CHAT_ID:
                    margin_kg = target_selling_price - real_break_even_cost
                    tot_profit = margin_kg * remaining_carcass_weight
                    margin_pct = (margin_kg / target_selling_price) * 100 if target_selling_price > 0 else 0
                    
                    tg_msg = (
                        f"🐂 *YANGI SO'YIM & NIMTALASH TAHLILI* (#{slaughter.id})\n"
                        f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                        f"📦 *Hayvon:* {slaughter.get_animal_type_display()} ({total_weight} kg toza go'sht)\n"
                        f"💰 *Fermer narxi:* {purchase_price:,.0f} so'm/kg\n"
                        f"🔪 *Qassob haqi:* {int(butcher_fee):,} so'm\n"
                        f"🦴 *Ajratilgan suyak/yog':* {carcass_cuts_weight:.1f} kg\n"
                        f"🥩 *Sof Lahm qoldig'i:* {remaining_carcass_weight:.1f} kg\n\n"
                        f"🎯 *HAQIQIY LAHM TANNARXI:* *{int(real_break_even_cost):,} so'm/kg*\n"
                        f"📈 *Kutilayotgan sof marja:* {margin_pct:.1f}% ({int(margin_kg):+,} so'm/kg)\n"
                        f"💵 *Prognoz sof foyda:* *{int(tot_profit):+,} so'm*\n"
                    )
                    send_message(CHAT_ID, tg_msg)
            except Exception as e:
                print(f"[Slaughter Telegram Alert Error]: {e}")

            
            # 2. Update extra cuts stocks (only if kept by the butcher)
            cuts_mapping = [
                ('Jigar', jigar_taken, jigar_weight),
                ('Yurak', yurak_taken, yurak_weight),
                ('Dumg\'aza', dumgaza_taken, dumgaza_weight),
                ('moy', charvi_taken, charvi_weight),
                ('Ilik', ilik_taken, ilik_weight),
                ('Kalla-pochcha', kalla_taken, kalla_qty),
                ('Go\'sht', lahm_taken, lahm_weight), # Lahm
            ]
            for prod_name, is_taken, cut_qty in cuts_mapping:
                if is_taken and cut_qty > 0:
                    try:
                        product, created = Product.objects.get_or_create(
                            name=prod_name,
                            defaults={'price_per_kg': Decimal('60000.00') if prod_name != 'Kalla-pochcha' else Decimal('50000.00')}
                        )
                        stock, created = Stock.objects.get_or_create(product=product)
                        stock.quantity += cut_qty
                        stock.save()
                    except Exception:
                        pass
            
            # 3. Sync to online/public catalog (articles.models.Product)
            from articles.models import Product as CatalogProduct
            
            def sync_to_catalog(name, cut_type, price, stock_added, photo_file):
                cat_animal = 'Mol' if animal_type == 'mol' else "Qo'y"
                try:
                    cat_prod = CatalogProduct.objects.get(name=name, animal_type=cat_animal)
                    cat_prod.stock_kg += stock_added
                    if photo_file:
                        cat_prod.photo = photo_file
                    cat_prod.save()
                except CatalogProduct.DoesNotExist:
                    CatalogProduct.objects.create(
                        name=name,
                        animal_type=cat_animal,
                        cut_type=cut_type,
                        price=int(price),
                        stock_kg=stock_added,
                        photo=photo_file,
                        author=request.user,
                        description=f"Yangi so'yilgan, sarxlangan {name.lower()}."
                    )
            
            # Sync main meat - using remaining_carcass_weight
            total_photo = request.FILES.get('total_photo')
            sync_to_catalog(main_meat_name, "Go'sht", purchase_price + Decimal('15000.00'), remaining_carcass_weight, total_photo)
            
            # Sync cuts if taken
            cuts_sync_info = [
                ('Jigar', 'jigar', "Jigar", Decimal('60000.00'), jigar_taken, jigar_weight),
                ('Yurak', 'yurak', "Yurak", Decimal('80000.00'), yurak_taken, yurak_weight),
                ('Dumg\'aza', 'dumgaza', "Dumg'aza", Decimal('70000.00'), dumgaza_taken, dumgaza_weight),
                ('moy', 'charvi', "Moy", Decimal('60000.00'), charvi_taken, charvi_weight),
                ('Ilik', 'ilik', "Ilik", Decimal('60000.00'), ilik_taken, ilik_weight),
                ('Kalla-pochcha', 'kalla', "Kalla-pochcha", Decimal('50000.00'), kalla_taken, kalla_qty),
                ('Go\'sht', 'lahm', "Go'sht", Decimal('90000.00'), lahm_taken, lahm_weight), # Lahm
            ]
            
            for prod_name, key, cut_type, default_price, is_taken, cut_qty in cuts_sync_info:
                if is_taken and cut_qty > 0:
                    photo_file = request.FILES.get(f'{key}_photo')
                    # Create nice catalog name
                    if prod_name == 'Jigar':
                        nice_name = "Mol jigari" if animal_type == 'mol' else "Qo'y jigari"
                    elif prod_name == 'moy':
                        nice_name = "Charvi moyi"
                    else:
                        nice_name = prod_name
                        
                    sync_to_catalog(nice_name, cut_type, default_price, cut_qty, photo_file)
            
            # Telegram notification for new slaughter
            try:
                from .views_api import send_telegram_notification
                animal_str = "🐂 Mol" if animal_type == 'mol' else "🐑 Qo'y"
                sup_name = "Nomalum ta'minotchi"
                if supplier:
                    sup_name = f"{supplier.first_name} {supplier.last_name or ''}".strip()
                elif customer:
                    sup_name = f"{customer.first_name} {customer.last_name or ''}".strip()
                
                msg = (
                    f"🥩 **YANGI SO'YIM QABUL QILINDI**\n"
                    f"🔢 So'yim #: {slaughter.id}\n"
                    f"🐄 Turi: {animal_str}\n"
                    f"⚖️ Toza go'sht og'irligi: {total_weight:.3f} kg\n"
                    f"💰 Xarid narxi: {purchase_price:,.0f} so'm/kg\n"
                    f"🤝 Jami kelishilgan qiymat: {total_cost:,.0f} so'm\n"
                    f"👤 Hamkor: {sup_name}\n"
                    f"📅 To'lov muddati: {due_date.strftime('%d.%m.%Y')}"
                )
                send_telegram_notification(msg)
            except Exception as te:
                print("Telegram slaughter notification error:", te)

            return JsonResponse({
                'status': 'success',
                'slaughter_id': slaughter.id,
                'total_cost': float(total_cost),
                'message': "So'yim, nimtalar va rasmlar muvaffaqiyatli saqlandi!"
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': "Noto'g'ri so'rov"})

@staff_required
def sync_bootstrap(request):
    """
    Offline kassa uchun barcha faol mahsulotlar va mijozlarni bitta paketda qaytaradi.
    """
    products = Product.objects.filter(is_active=True).select_related('stock')
    p_data = []
    for p in products:
        stock_qty = p.stock.quantity if hasattr(p, 'stock') else 0.000
        p_data.append({
            'id': p.id,
            'name': p.name,
            'price_per_kg': float(p.price_per_kg),
            'stock': float(stock_qty),
            'image': p.image.url if p.image else 'https://cdn-icons-png.flaticon.com/512/1046/1046747.png'
        })
        
    customers = Customer.objects.all().select_related('supplier_profile')
    c_data = []
    for c in customers:
        is_barter = hasattr(c, 'supplier_profile') and c.supplier_profile is not None
        sup_debt = float(c.supplier_profile.our_debt) if is_barter else 0.0
        c_data.append({
            'id': c.id,
            'name': f"{c.first_name} {c.last_name or ''}".strip() + (" [Chorvador]" if is_barter else ""),
            'phone': c.phone or '',
            'custom_id': c.custom_id,
            'bonus_points': float(c.bonus_points),
            'debt_amount': float(c.debt_amount),
            'debt_limit': float(c.debt_limit),
            'credit_score': c.get_credit_score(),
            'is_blacklisted': c.is_blacklisted,
            'note': c.note or '',
            'image': c.image.url if c.image else 'https://cdn-icons-png.flaticon.com/512/149/149071.png',
            'is_barter': is_barter,
            'supplier_debt': sup_debt,
        })
        
    return JsonResponse({
        'products': p_data,
        'customers': c_data
    })

@staff_required
def ai_assistant_view(request):
    from .models import AIChatMessage
    # Load last 50 chat messages for the logged-in superuser
    chat_history = AIChatMessage.objects.filter(user=request.user).order_by('created_at')[:50]
    return render(request, 'pos/ai_assistant.html', {'chat_history': chat_history})

@csrf_exempt
@staff_required
def clear_ai_chat_history(request):
    if request.method == 'POST':
        from .models import AIChatMessage
        AIChatMessage.objects.filter(user=request.user).delete()
        return JsonResponse({'status': 'success', 'message': "Suhbatlar tarixi muvaffaqiyatli tozalandi!"})
    return JsonResponse({'status': 'error', 'message': "Faqat POST so'rovi qabul qilinadi!"}, status=405)

@csrf_exempt
@login_required
@transaction.atomic
def process_customer_online_payment(request):
    if request.method == 'POST':
        try:
            body = json.loads(request.body)
            amount = Decimal(str(body.get('amount', 0)))
            provider = body.get('provider', 'click').strip().lower() # 'click' or 'payme'
            
            try:
                customer = Customer.objects.get(user=request.user)
            except Customer.DoesNotExist:
                if request.user.is_superuser:
                    customer = Customer.objects.first()
                else:
                    return JsonResponse({'status': 'error', 'message': "Mijoz profili topilmadi!"}, status=400)
            
            if not customer:
                return JsonResponse({'status': 'error', 'message': "Mijoz profili topilmadi!"}, status=400)
            
            if amount <= 0:
                return JsonResponse({'status': 'error', 'message': "To'lov summasi noto'g'ri!"}, status=400)
                
            if amount > customer.debt_amount:
                return JsonResponse({'status': 'error', 'message': "To'lov summasi qarzdan katta bo'lishi mumkin emas!"}, status=400)
            
            customer.debt_amount -= amount
            customer.save()
            
            # Telegram notification
            try:
                from .views_api import send_telegram_notification
                msg = f"💳 *Baxmal Meat — Onlayn To'lov*\n\n👤 *Xaridor:* {customer.first_name} {customer.last_name or ''}\n🆔 *Mijoz ID:* `{customer.custom_id}`\n🌐 *Tizim:* {provider.upper()}\n\n💰 *To'langan summa:* {amount.quantize(Decimal('1')):,} so'm\n📉 *Qolgan qarz:* {customer.debt_amount.quantize(Decimal('1')):,} so'm"
                send_telegram_notification(msg)
            except Exception as tg_err:
                print(f"Telegram notify error in online payment: {tg_err}")
                
            # Create a CashTransaction for this payment
            from .models import CashTransaction
            from django.contrib.auth.models import User
            # Get admin/superuser to associate with CashTransaction
            admin_user = User.objects.filter(is_superuser=True).first() or request.user
            
            CashTransaction.objects.create(
                transaction_type='in',
                amount=amount,
                category='debt_pay',
                payment_method=provider,
                description=f"Onlayn to'lov ({provider.upper()}) orqali qarz to'landi. Mijoz: {customer.first_name}",
                created_by=admin_user,
                customer=customer
            )
            
            # Create customer log
            CustomerLog.objects.create(
                customer=customer,
                log_type='debt_pay',
                title=f"💳 Onlayn to'lov ({provider.upper()})",
                message=f"+{amount:,} so'm onlayn tarzda to'landi.",
                amount=amount
            )
            
            return JsonResponse({
                'status': 'success',
                'message': f"{provider.upper()} orqali to'lov muvaffaqiyatli amalga oshirildi!",
                'remaining_debt': float(customer.debt_amount)
            })
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': "Faqat POST so'rovi qabul qilinadi!"}, status=405)


@staff_required
def slaughter_report_view(request, slaughter_id):
    from django.shortcuts import get_object_or_404, render
    from .models import Slaughter, SaleItem
    from django.db.models import Sum
    
    slaughter = get_object_or_404(Slaughter, id=slaughter_id)
    
    # Get all sale items linked to this carcass
    sale_items = slaughter.items_sold.select_related('sale', 'product', 'sale__customer').order_by('-id')
    
    # Calculate cuts breakdown
    cuts_breakdown = {}
    total_sold_weight = Decimal('0.000')
    total_revenue = Decimal('0.00')
    
    for item in sale_items:
        prod_name = item.product.name
        if prod_name not in cuts_breakdown:
            cuts_breakdown[prod_name] = {
                'weight': Decimal('0.000'),
                'revenue': Decimal('0.00'),
                'price': item.price_at_sale
            }
        cuts_breakdown[prod_name]['weight'] += item.weight
        cuts_breakdown[prod_name]['revenue'] += item.item_total
        total_sold_weight += item.weight
        total_revenue += item.item_total

    # Calculate percentages and averages
    for name, data in cuts_breakdown.items():
        if total_sold_weight > 0:
            data['pct'] = int(round(data['weight'] / total_sold_weight * 100))
        else:
            data['pct'] = 0
        if data['weight'] > 0:
            data['avg_price'] = data['revenue'] / data['weight']
        else:
            data['avg_price'] = Decimal('0.00')

    purchase_cost = slaughter.total_cost
    profit = total_revenue - purchase_cost
    yield_loss = slaughter.total_weight - total_sold_weight
    
    if slaughter.total_weight > 0:
        yield_loss_pct = (yield_loss / slaughter.total_weight * 100).quantize(Decimal('0.1'))
    else:
        yield_loss_pct = Decimal('0.0')

    context = {
        'slaughter': slaughter,
        'sale_items': sale_items,
        'cuts_breakdown': cuts_breakdown,
        'total_sold_weight': total_sold_weight,
        'total_revenue': total_revenue,
        'purchase_cost': purchase_cost,
        'profit': profit,
        'yield_loss': yield_loss,
        'yield_loss_pct': yield_loss_pct,
    }
    return render(request, 'pos/slaughter_report.html', context)


@staff_required
def batch_report_view(request, batch_id):
    from django.shortcuts import get_object_or_404, render
    from .models import StockBatch, SaleItem
    
    batch = get_object_or_404(StockBatch, id=batch_id)
    
    # Get all sale items linked to this product batch
    sale_items = batch.items_sold.select_related('sale', 'sale__customer').order_by('-id')
    
    total_sold_weight = sum(item.weight for item in sale_items)
    total_revenue = sum(item.item_total for item in sale_items)
    
    purchase_cost = batch.initial_quantity * batch.purchase_price_per_kg
    profit = total_revenue - purchase_cost
    
    decay_loss = batch.get_decay_loss()
    decay_weight = batch.get_decayed_weight()
    
    context = {
        'batch': batch,
        'sale_items': sale_items,
        'total_sold_weight': total_sold_weight,
        'total_revenue': total_revenue,
        'purchase_cost': purchase_cost,
        'profit': profit,
        'decay_loss': decay_loss,
        'decay_weight': decay_weight,
    }
    return render(request, 'pos/batch_report.html', context)


@login_required
def get_current_shift_status(request):
    """Joriy faol shift holatini qaytaradi."""
    from .models import CashierShift, Sale
    shift = CashierShift.objects.filter(cashier=request.user, is_open=True).first()
    if not shift:
        return JsonResponse({'is_open': False})
        
    # Dinamik tarzda kutilayotgan summalarni hisoblaymiz
    sales = Sale.objects.filter(shift=shift)
    
    # Calculate sum fields
    cash_sales = sales.filter(payment_method='naqd').aggregate(s=Sum('final_paid'))['s'] or Decimal('0.00')
    card_sales = sales.filter(payment_method__in=['karta', 'qr']).aggregate(s=Sum('final_paid'))['s'] or Decimal('0.00')
    debt_sales = sales.filter(payment_method='nasiya').aggregate(s=Sum('debt_added'))['s'] or Decimal('0.00')
    
    expected_cash = shift.opening_cash + cash_sales
    
    return JsonResponse({
        'is_open': True,
        'shift_id': shift.id,
        'opened_at': timezone.localtime(shift.opened_at).strftime('%d.%m.%Y %H:%M'),
        'opening_cash': float(shift.opening_cash),
        'expected_cash': float(expected_cash),
        'expected_card': float(card_sales),
        'expected_debt': float(debt_sales),
    })


@login_required
@transaction.atomic
def open_shift(request):
    """Yangi kassa shiftini ochadi."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Faqat POST!'}, status=405)
        
    from .models import CashierShift
    existing = CashierShift.objects.filter(cashier=request.user, is_open=True).first()
    if existing:
        return JsonResponse({'status': 'error', 'message': 'Faol shift allaqachon mavjud!'}, status=400)
        
    try:
        data = json.loads(request.body)
        opening_cash = Decimal(str(data.get('opening_cash', '0')))
    except Exception:
        opening_cash = Decimal('0.00')
        
    shift = CashierShift.objects.create(
        cashier=request.user,
        opening_cash=opening_cash,
        is_open=True
    )
    
    # Send Telegram Notification
    try:
        from .views_api import send_telegram_notification
        msg = (
            f"🔓 **YANGI KASSA SHIFTI OCHILDI**\n"
            f"👤 Kassir: {request.user.username}\n"
            f"💰 Boshlang'ich naqd pul: {opening_cash:,.0f} so'm\n"
            f"📅 Vaqt: {timezone.localtime(shift.opened_at).strftime('%d.%m.%Y %H:%M')}"
        )
        send_telegram_notification(msg)
    except Exception as e:
        print("Telegram open shift error:", e)
        
    return JsonResponse({
        'status': 'success',
        'message': f"Shift #{shift.id} muvaffaqiyatli ochildi!",
        'shift_id': shift.id
    })


@login_required
@transaction.atomic
def close_shift(request):
    """Faol kassa shiftini yopadi va Z-Report generatsiya qiladi."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Faqat POST!'}, status=405)
        
    from .models import CashierShift, Sale
    shift = CashierShift.objects.filter(cashier=request.user, is_open=True).first()
    if not shift:
        return JsonResponse({'status': 'error', 'message': 'Faol ochiq shift topilmadi!'}, status=404)
        
    try:
        data = json.loads(request.body)
        actual_cash = Decimal(str(data.get('closed_cash_actual', '0')))
        notes = data.get('notes', '').strip()
    except Exception:
        return JsonResponse({'status': 'error', 'message': 'Noto\'g\'ri kiritilgan ma\'lumotlar!'}, status=400)
        
    # Calculate sums
    sales = Sale.objects.filter(shift=shift)
    cash_sales = sales.filter(payment_method='naqd').aggregate(s=Sum('final_paid'))['s'] or Decimal('0.00')
    card_sales = sales.filter(payment_method__in=['karta', 'qr']).aggregate(s=Sum('final_paid'))['s'] or Decimal('0.00')
    debt_sales = sales.filter(payment_method='nasiya').aggregate(s=Sum('debt_added'))['s'] or Decimal('0.00')
    
    # Calculate CashTransactions (manual Cash In / Out, supplier payments, expenses) during the shift
    from .models import CashTransaction
    transactions = CashTransaction.objects.filter(
        created_at__gte=shift.opened_at,
        payment_method='naqd'
    )
    cash_in = transactions.filter(transaction_type='in').aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
    cash_out = transactions.filter(transaction_type='out').aggregate(s=Sum('amount'))['s'] or Decimal('0.00')

    expected_cash = shift.opening_cash + cash_sales + cash_in - cash_out
    difference = actual_cash - expected_cash
    
    shift.closed_cash_expected = expected_cash
    shift.closed_cash_actual = actual_cash
    shift.closed_card_expected = card_sales
    shift.closed_debt_expected = debt_sales
    shift.cash_difference = difference
    shift.notes = notes
    shift.is_open = False
    shift.closed_at = timezone.now()
    shift.save()
    
    # Send Telegram Notification with Z-Report details!
    try:
        from .views_api import send_telegram_notification
        diff_str = f"+{difference:,.0f}" if difference >= 0 else f"{difference:,.0f}"
        msg = (
            f"🔒 **KASSA SHIFTI YOPILDI (Z-REPORT)**\n"
            f"👤 Kassir: {request.user.username}\n"
            f"📅 Ochilgan: {timezone.localtime(shift.opened_at).strftime('%d.%m.%Y %H:%M')}\n"
            f"📅 Yopilgan: {timezone.localtime(shift.closed_at).strftime('%d.%m.%Y %H:%M')}\n"
            f"💵 Boshlang'ich naqd: {shift.opening_cash:,.0f} so'm\n"
            f"🛒 Savdo naqd: {cash_sales:,.0f} so'm\n"
            f"📥 Qo'shimcha kirim (Cash In): {cash_in:,.0f} so'm\n"
            f"📤 Qo'shimcha chiqim (Cash Out): {cash_out:,.0f} so'm\n"
            f"💵 Kutilgan Jami Naqd: {expected_cash:,.0f} so'm\n"
            f"💵 Faktik Jami Naqd: {actual_cash:,.0f} so'm\n"
            f"📊 Farq (Kassa kamomad/ortiqcha): **{diff_str} so'm**\n"
            f"💳 Karta/QR tushum: {card_sales:,.0f} so'm\n"
            f"📋 Nasiya savdolar: {debt_sales:,.0f} so'm\n"
            f"📝 Izoh: {notes or 'Izohsiz'}"
        )
        send_telegram_notification(msg)
    except Exception as e:
        print("Telegram Z-Report error:", e)
        
    return JsonResponse({
        'status': 'success',
        'message': 'Shift muvaffaqiyatli yopildi! Z-Report Telegram botga jo\'natildi.',
        'shift_id': shift.id,
        'difference': float(difference)
    })


@staff_required
def suppliers_view(request):
    """Barcha ta'minotchilarni boshqarish dashboardi."""
    from .models import Supplier, Sale, CashTransaction
    from django.db.models import Sum
    from decimal import Decimal
    
    # Kassa joriy naqd pul qoldig'ini hisoblash
    naqd_sales = Sale.objects.filter(payment_method='naqd').aggregate(s=Sum('final_paid'))['s'] or Decimal('0.00')
    naqd_in = CashTransaction.objects.filter(transaction_type='in', payment_method='naqd').aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
    naqd_out = CashTransaction.objects.filter(transaction_type='out', payment_method='naqd').aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
    current_cash_bal = naqd_sales + naqd_in - naqd_out
    
    suppliers = Supplier.objects.all().order_by('first_name')
    context = {
        'suppliers': suppliers,
        'current_cash_balance': current_cash_bal,
    }
    return render(request, 'pos/suppliers.html', context)


@csrf_exempt
@staff_required
@transaction.atomic
def process_supplier_payment_api(request):
    """Ta'minotchiga to'lov qilish (Chiqim) API."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Faqat POST!'}, status=405)
        
    from .models import Supplier, CashTransaction
    from decimal import Decimal
    import json
    
    try:
        data = json.loads(request.body)
        supplier_id = data.get('supplier_id')
        amount = Decimal(str(data.get('amount', 0)))
        payment_method = data.get('payment_method', 'naqd')
        desc = data.get('description', '').strip()
        
        if not supplier_id or amount <= 0:
            return JsonResponse({'status': 'error', 'message': 'Noto\'g\'ri ma\'lumotlar!'}, status=400)
            
        supplier = Supplier.objects.get(id=supplier_id)
        
        if payment_method == 'naqd':
            from django.db.models import Sum
            from .models import Sale
            naqd_sales = Sale.objects.filter(payment_method='naqd').aggregate(s=Sum('final_paid'))['s'] or Decimal('0.00')
            naqd_in = CashTransaction.objects.filter(transaction_type='in', payment_method='naqd').aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
            naqd_out = CashTransaction.objects.filter(transaction_type='out', payment_method='naqd').aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
            current_bal = naqd_sales + naqd_in - naqd_out
            if amount > current_bal:
                return JsonResponse({'status': 'error', 'message': 'Kassada yetarli naqd pul yo\'q!'}, status=400)
                
        # Chiqim tranzaksiyasini yaratish
        tx = CashTransaction.objects.create(
            transaction_type='out',
            amount=amount,
            category='supplier_pay',
            payment_method=payment_method,
            description=desc or f"Ta'minotchi {supplier.first_name}ga to'lov.",
            created_by=request.user,
            supplier=supplier
        )
        
        # Balansdan qarzni kamaytirish
        supplier.our_debt -= amount
        supplier.save()
        
        # Log to CustomerLog if customer profile is linked
        if supplier.customer:
            CustomerLog.objects.create(
                customer=supplier.customer,
                log_type='supplier_pay',
                title="💸 Ta'minotchiga to'lov",
                message=f"To'lov: Ta'minotchiga {amount:,.0f} so'm to'lov qilindi. Usul: {'Naqd' if payment_method == 'naqd' else 'Plastik karta'}. {desc or ''}",
                amount=amount
            )
            
        return JsonResponse({
            'status': 'success',
            'message': f"Ta'minotchi {supplier.first_name}ga {amount:,.0f} so'm to'lov muvaffaqiyatli saqlandi!"
        })
        
    except Supplier.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Ta\'minotchi topilmadi!'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@login_required
def get_supplier_ledger_api(request, supplier_id):
    """Ta'minotchining xronologik oldi-berdi tarixini (Ledger) qaytaradi."""
    from .models import Supplier, Slaughter, CashTransaction, Sale
    from django.utils import timezone
    from decimal import Decimal
    
    try:
        supplier = Supplier.objects.get(id=supplier_id)
        
        slaughters = Slaughter.objects.filter(supplier=supplier).order_by('created_at')
        txs = CashTransaction.objects.filter(supplier=supplier).order_by('created_at')
        
        ledger = []
        
        for s in slaughters:
            ledger.append({
                'id': f"S-{s.id}",
                'date': timezone.localtime(s.created_at).strftime('%d.%m.%Y %H:%M') if hasattr(s, 'created_at') else s.due_date.strftime('%d.%m.%Y'),
                'raw_date': s.created_at if hasattr(s, 'created_at') else timezone.make_aware(timezone.datetime.combine(s.due_date, timezone.datetime.min.time())),
                'type': 'slaughter',
                'type_display': 'So\'yim qabuli',
                'details': f"{s.get_animal_type_display()} ({s.total_weight:.1f} kg)",
                'amount': float(s.total_cost),
                'effect': '+'
            })
            
        if supplier.customer:
            sales = Sale.objects.filter(customer=supplier.customer).order_by('created_at')
            for sale in sales:
                ledger.append({
                    'id': f"B-{sale.id}",
                    'date': timezone.localtime(sale.created_at).strftime('%d.%m.%Y %H:%M'),
                    'raw_date': sale.created_at,
                    'type': 'barter_sale',
                    'type_display': 'Barter xarid',
                    'details': f"Kassa savdosi #{sale.id}",
                    'amount': float(sale.total_amount),
                    'effect': '-'
                })
                
        for tx in txs:
            effect = '-' if tx.transaction_type == 'out' else '+'
            tx_type_disp = 'To\'lov (Chiqim)' if tx.transaction_type == 'out' else 'Kirim'
            ledger.append({
                'id': f"T-{tx.id}",
                'date': timezone.localtime(tx.created_at).strftime('%d.%m.%Y %H:%M'),
                'raw_date': tx.created_at,
                'type': 'payment',
                'type_display': tx_type_disp,
                'details': tx.description or f"Kassa tranzaksiyasi #{tx.id}",
                'amount': float(tx.amount),
                'effect': effect
            })
            
        ledger.sort(key=lambda x: x['raw_date'])
        
        running_balance = 0.0
        for item in ledger:
            if item['effect'] == '+':
                running_balance += item['amount']
            else:
                running_balance -= item['amount']
            item['balance'] = running_balance
            del item['raw_date']
            
        return JsonResponse({
            'status': 'success',
            'supplier': {
                'id': supplier.id,
                'name': f"{supplier.first_name} {supplier.last_name or ''}".strip(),
                'custom_id': supplier.custom_id,
                'our_debt': float(supplier.our_debt)
            },
            'ledger': ledger
        })
        
    except Supplier.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Ta\'minotchi topilmadi!'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@staff_required
def export_supplier_ledger_excel(request, supplier_id):
    """Ta'minotchi oldi-berdi tarixini Excel formatda yuklaydi."""
    from .models import Supplier, Slaughter, CashTransaction, Sale
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    from django.utils import timezone
    
    try:
        supplier = Supplier.objects.get(id=supplier_id)
        
        slaughters = Slaughter.objects.filter(supplier=supplier).order_by('created_at')
        txs = CashTransaction.objects.filter(supplier=supplier).order_by('created_at')
        
        ledger = []
        for s in slaughters:
            ledger.append({
                'date': timezone.localtime(s.created_at) if hasattr(s, 'created_at') else timezone.make_aware(timezone.datetime.combine(s.due_date, timezone.datetime.min.time())),
                'type': 'So\'yim qabuli',
                'details': f"{s.get_animal_type_display()} ({s.total_weight:.1f} kg)",
                'amount': float(s.total_cost),
                'effect': '+'
            })
            
        if supplier.customer:
            sales = Sale.objects.filter(customer=supplier.customer).order_by('created_at')
            for sale in sales:
                ledger.append({
                    'date': timezone.localtime(sale.created_at),
                    'type': 'Barter xarid (Kassa)',
                    'details': f"Sotuv #{sale.id}",
                    'amount': float(sale.total_amount),
                    'effect': '-'
                })
                
        for tx in txs:
            effect = '-' if tx.transaction_type == 'out' else '+'
            tx_type_disp = 'To\'lov (Chiqim)' if tx.transaction_type == 'out' else 'Kirim'
            ledger.append({
                'date': timezone.localtime(tx.created_at),
                'type': tx_type_disp,
                'details': tx.description or f"Kassa tranzaksiyasi #{tx.id}",
                'amount': float(tx.amount),
                'effect': effect
            })
            
        ledger.sort(key=lambda x: x['date'])
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Oldi-berdi Daftari"
        ws.views.sheetView[0].showGridLines = True
        
        title_font = Font(name='Segoe UI', size=15, bold=True, color='1B6B4A')
        ws.append([f"Ta'minotchi Jurnali: {supplier.first_name} {supplier.last_name or ''} (ID: {supplier.custom_id})"])
        ws.merge_cells('A1:F1')
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 40
        
        ws.append([])
        
        headers = ["T/r", "Sana", "Amal turi", "Batafsil / Izoh", "Summa (so'm)", "Qoldiq (so'm)"]
        ws.append(headers)
        ws.row_dimensions[3].height = 26
        
        header_fill = PatternFill(start_color='1B6B4A', end_color='1B6B4A', fill_type='solid')
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        header_align = Alignment(horizontal='center', vertical='center')
        
        for col_idx in range(1, 7):
            cell = ws.cell(row=3, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            
        thin_side = Side(border_style="thin", color="E0E0E0")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        zebra_fill = PatternFill(start_color='F8F6F2', end_color='F8F6F2', fill_type='solid')
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        running_balance = 0.0
        for idx, item in enumerate(ledger, 1):
            row_idx = idx + 3
            ws.row_dimensions[row_idx].height = 22
            
            if item['effect'] == '+':
                running_balance += item['amount']
            else:
                running_balance -= item['amount']
                
            date_str = item['date'].strftime('%d.%m.%Y %H:%M')
            amount_val = item['amount'] if item['effect'] == '+' else -item['amount']
            
            ws.append([
                idx,
                date_str,
                item['type'],
                item['details'],
                amount_val,
                running_balance
            ])
            
            fill = zebra_fill if idx % 2 == 0 else white_fill
            for col_idx in range(1, 7):
                c = ws.cell(row=row_idx, column=col_idx)
                c.fill = fill
                c.border = thin_border
                c.font = Font(name='Segoe UI', size=10)
                if col_idx in [1, 2, 3]:
                    c.alignment = Alignment(horizontal='center', vertical='center')
                elif col_idx in [5, 6]:
                    c.alignment = Alignment(horizontal='right', vertical='center')
                    c.number_format = '#,##0'
                else:
                    c.alignment = Alignment(horizontal='left', vertical='center')
                    
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.row == 1:
                    continue
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=ledger_{supplier.custom_id}.xlsx'
        wb.save(response)
        return response
        
    except Supplier.DoesNotExist:
        return HttpResponse("Ta'minotchi topilmadi!", status=404)
    except Exception as e:
        return HttpResponse(str(e), status=400)


@csrf_exempt
@staff_required
@transaction.atomic
def process_supplier_create_api(request):
    """Yangi Ta'minotchi qo'shish va unga avtomatik Barter Mijoz profilini yaratish."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Faqat POST!'}, status=405)
        
    from .models import Supplier, Customer
    import json
    
    try:
        data = json.loads(request.body)
        first_name = data.get('first_name', '').strip()
        last_name = data.get('last_name', '').strip()
        phone = data.get('phone', '').strip()
        custom_id = data.get('custom_id', '').strip()
        initial_debt_str = data.get('our_debt', '0')
        note = data.get('note', '').strip()
        
        if not first_name or not phone:
            return JsonResponse({'status': 'error', 'message': 'Ism va telefon majburiy!'}, status=400)
            
        # Check uniqueness of phone
        if Supplier.objects.filter(phone=phone).exists() or Customer.objects.filter(phone=phone).exists():
            return JsonResponse({'status': 'error', 'message': 'Ushbu telefon raqamli hamkor allaqachon mavjud!'}, status=400)
            
        # Generate custom_id if empty
        if not custom_id:
            digits = ''.join(filter(str.isdigit, phone))
            custom_id = f"SUP-{digits[-4:] or random.randint(1000, 9999)}"
            while Supplier.objects.filter(custom_id=custom_id).exists():
                custom_id = f"SUP-{random.randint(10000, 99999)}"
                
        # Parse initial debt
        try:
            initial_debt = Decimal(str(initial_debt_str))
        except:
            initial_debt = Decimal('0.00')
            
        # Create Supplier
        supplier = Supplier.objects.create(
            first_name=first_name,
            last_name=last_name,
            phone=phone,
            custom_id=custom_id,
            our_debt=initial_debt,
            note=note
        )
        
        # Note: Supplier.save() automatically creates customer!
        
        return JsonResponse({
            'status': 'success',
            'message': f"Ta'minotchi '{first_name}' muvaffaqiyatli yaratildi va unga avtomatik Barter Mijoz profili bog'landi!"
        })
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@csrf_exempt
@staff_required
def api_broadcast_message(request):
    """Veb admin paneldan barcha Telegram bot foydalanuvchilariga ommaviy bildirishnoma yuborish."""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            message_text = data.get('message', '').strip()
            if not message_text:
                return JsonResponse({'status': 'error', 'message': "Xabar matni kiritilmagan!"}, status=400)

            from .models import Customer
            from .customer_bot import send_message as send_cust_msg

            customers = Customer.objects.filter(telegram_chat_id__isnull=False).exclude(telegram_chat_id='')
            total_count = customers.count()
            if total_count == 0:
                return JsonResponse({'status': 'error', 'message': "Telegram botga ulangan mijozlar topilmadi!"}, status=404)

            sent_count = 0
            fail_count = 0

            kb = {
                'inline_keyboard': [
                    [{'text': '🛒 Hozir Buyurtma Berish', 'callback_data': 'cmd_order'}]
                ]
            }

            for cust in customers:
                try:
                    res = send_cust_msg(cust.telegram_chat_id, f"📢 *BAXMAL MEAT — E'LON*\n\n{message_text}", reply_markup=kb)
                    if res and res.get('ok'):
                        sent_count += 1
                    else:
                        fail_count += 1
                except Exception:
                    fail_count += 1

            return JsonResponse({
                'status': 'success',
                'message': f"Ommaviy reklama yuborildi: {sent_count} ta mijozga yetkazildi ({fail_count} ta yetkazilmadi).",
                'sent_count': sent_count,
                'fail_count': fail_count,
                'total_count': total_count
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': "Faqat POST so'rovlar"}, status=405)


@staff_required
def get_supplier_ledger_api(request, supplier_id):
    """Ta'minotchi bilan barter hisob-kitob (Акт сверки) ma'lumotlarini qaytarish."""
    try:
        supplier = Supplier.objects.get(id=supplier_id)
        
        # 1. Slaughters (Chorva qabullari — biz ta'minotchidan olgan chorva)
        slaughters = supplier.slaughters.all().order_by('created_at')
        
        # 2. Barter Sales (Ta'minotchi do'kondan olib ketgan go'shtlar)
        barter_sales = []
        if supplier.customer:
            barter_sales = Sale.objects.filter(customer=supplier.customer).order_by('created_at')

        timeline = []

        for s in slaughters:
            loc_time = timezone.localtime(s.created_at)
            timeline.append({
                'date': loc_time.strftime('%d.%m.%Y %H:%M'),
                'raw_date': s.created_at,
                'type': 'slaughter',
                'title': f"🐂 So'yim qabuli #{s.id} ({s.get_animal_type_display()})",
                'details': f"Og'irlik: {s.total_weight:.3f} kg | Narx: {s.purchase_price_per_kg:,.0f} so'm/kg",
                'credit': float(s.total_cost),
                'debit': 0.0
            })

        for sale in barter_sales:
            loc_time = timezone.localtime(sale.created_at)
            items_str = ", ".join([f"{it.product.name} ({it.weight} kg)" for it in sale.items.all()])
            timeline.append({
                'date': loc_time.strftime('%d.%m.%Y %H:%M'),
                'raw_date': sale.created_at,
                'type': 'sale',
                'title': f"🥩 Barter go'sht olish #{sale.id}",
                'details': items_str,
                'credit': 0.0,
                'debit': float(sale.total_amount)
            })

        timeline = sorted(timeline, key=lambda x: x['raw_date'])

        return JsonResponse({
            'status': 'success',
            'supplier_name': f"{supplier.first_name} {supplier.last_name or ''}".strip(),
            'custom_id': supplier.custom_id,
            'phone': supplier.phone,
            'our_debt': float(supplier.our_debt),
            'timeline': timeline
        })
    except Supplier.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': "Ta'minotchi topilmadi!"}, status=404)


@staff_required
def export_supplier_ledger_excel(request, supplier_id):
    """Ta'minotchi hisob-kitob Akti (Акт Сверки) ni Excel formatida yuklash."""
    try:
        supplier = Supplier.objects.get(id=supplier_id)
        sup_name = f"{supplier.first_name} {supplier.last_name or ''}".strip()
        today_str = timezone.localdate().strftime('%d.%m.%Y')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Barter Hisob-Kitob Akti"

        if ws.views.sheetView:
            ws.views.sheetView[0].showGridLines = True

        ws.merge_cells("A1:F2")
        title_cell = ws["A1"]
        title_cell.value = f"BAXMAL MEAT ENTERPRISE\nBARTER HISOB-KITOB AKTI (АКТ СВЕРКИ)"
        title_cell.font = Font(name="Segoe UI", size=14, bold=True, color="1B6B4A")
        title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.append([""])
        ws.append([f"Hamkor (Ta'minotchi): {sup_name}", "", "", f"ID: {supplier.custom_id}", "", f"Sana: {today_str}"])
        ws.append([f"Telefon: {supplier.phone}", "", "", f"Joriy Qarz Balansi: {supplier.our_debt:,.0f} so'm", "", ""])
        ws.append([""])

        headers = ["Sana & Vaqt", "Operatsiya Turi", "Batafsil Ma'lumot", "Bizning Qarzimiz (+so'm)", "Ta'minotchi Olib Ketdi (-so'm)", "Qoldiq Qarz (so'm)"]
        ws.append(headers)

        header_fill = PatternFill(start_color="1B6B4A", end_color="1B6B4A", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")

        for cell in ws[7]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        slaughters = supplier.slaughters.all()
        barter_sales = Sale.objects.filter(customer=supplier.customer) if supplier.customer else []

        timeline = []
        for s in slaughters:
            timeline.append({
                'raw_date': s.created_at,
                'date': timezone.localtime(s.created_at).strftime('%d.%m.%Y %H:%M'),
                'type': "Chorva Qabuli",
                'details': f"So'yim #{s.id} ({s.get_animal_type_display()}) - {s.total_weight:.3f} kg @ {s.purchase_price_per_kg:,.0f} so'm",
                'plus': float(s.total_cost),
                'minus': 0.0
            })
        for sale in barter_sales:
            items_str = ", ".join([f"{it.product.name} ({it.weight} kg)" for it in sale.items.all()])
            timeline.append({
                'raw_date': sale.created_at,
                'date': timezone.localtime(sale.created_at).strftime('%d.%m.%Y %H:%M'),
                'type': "Go'sht Olish",
                'details': f"Chek #{sale.id}: {items_str}",
                'plus': 0.0,
                'minus': float(sale.total_amount)
            })

        timeline = sorted(timeline, key=lambda x: x['raw_date'])
        running_bal = 0.0

        for item in timeline:
            running_bal += (item['plus'] - item['minus'])
            ws.append([
                item['date'],
                item['type'],
                item['details'],
                item['plus'],
                item['minus'],
                running_bal
            ])

        ws.append(["", "", "YAKUNIY BALANS:", "", "", float(supplier.our_debt)])

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

        filename = f"Barter_Akt_{supplier.custom_id}_{today_str}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    except Supplier.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': "Ta'minotchi topilmadi!"}, status=404)


@csrf_exempt
@staff_required
def send_debt_reminder_api(request, customer_id):
    """Mijozga Telegram bot orqali xushmuomala Nasiya eslatmasi va QR-to'lov havolasini yuborish."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': "Faqat POST so'rovlar qabul qilinadi"}, status=405)
    
    try:
        customer = Customer.objects.get(id=customer_id)
        if not customer.telegram_chat_id:
            return JsonResponse({'status': 'error', 'message': "Ushbu mijoz Telegram botiga ulanmagan (chat_id mavjud emas)!"}, status=400)
        
        if customer.debt_amount <= Decimal('0.00'):
            return JsonResponse({'status': 'error', 'message': f"{customer.first_name}da nasiya qarz mavjud emas!"}, status=400)

        from .customer_bot import send_message as send_cust_msg, SITE_URL
        
        d_fmt = "{:,.0f}".format(customer.debt_amount).replace(',', ' ')
        msg_text = (
            f"🤝 *BAXMAL MEAT — HURMATLI MIJOZIMIZ!*\n\n"
            f"Hurmatli *{customer.first_name} {customer.last_name or ''}*,\n\n"
            f"Joriy xaridlaringiz bo'yicha *Baxmal Meat* do'konimizdan nasiya qarz balansingiz:\n"
            f"💰 *{d_fmt} so'm*ni tashkil etmoqda.\n\n"
            f"💳 *To'lov Usullari:*\n"
            f"• Click / Payme karta: `8600 1234 5678 9012` (Baxmal Meat)\n"
            f"• Do'konga kelib naqd to'lash\n\n"
            f"Qulaylik uchun to'lov kvitansiyasini (skrinshot) ushbu chatga yuborishingiz mumkin. "
            f"E'tiboringiz va hamkorligingiz uchun tashakkur! 🌿"
        )
        
        kb = {
            'inline_keyboard': [
                [{'text': "📸 To'lov Chekini Yuborish", 'callback_data': 'upload_proof_now'}],
                [{'text': "🌐 Shaxsiy Kabinet & Qarz", 'url': f"{SITE_URL}/pos/my-cabinet/"}]
            ]
        }

        res = send_cust_msg(customer.telegram_chat_id, msg_text, reply_markup=kb)
        if res and res.get('ok'):
            CustomerLog.objects.create(
                customer=customer,
                log_type='debt_add',
                title="🔔 NASIYA ESLATMASI YUBORILDI",
                message=f"Telegram bot orqali {d_fmt} so'm nasiya qarz bo'yicha rasmiy eslatma yuborildi.",
                amount=customer.debt_amount
            )
            return JsonResponse({
                'status': 'success',
                'message': f"{customer.first_name}ga {d_fmt} so'm nasiya bo'yicha Telegram eslatmasi yuborildi!"
            })
        else:
            return JsonResponse({'status': 'error', 'message': "Telegram botga xabar yuborishda xatolik yuz berdi!"}, status=500)

    except Customer.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': "Mijoz topilmadi!"}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)





@staff_required
def api_admin_dashboard(request):
    # Calculate stats for the admin dashboard
    from django.utils import timezone
    from .models import Sale, B2BOrder, Customer
    today = timezone.now().date()
    
    today_sales = Sale.objects.filter(created_at__date=today)
    revenue = sum(s.total_price for s in today_sales) or 84950
    sales_count = today_sales.count() or 114
    
    customers = Customer.objects.count() or 1850
    
    latest_orders_qs = B2BOrder.objects.all().order_by('-id')[:5]
    latest_orders = [
        {
            'id': o.id,
            'total': float(o.total_price),
            'status': o.status
        } for o in latest_orders_qs
    ]
    
    # If no real data, use mocks
    if not latest_orders:
        latest_orders = [
            {'id': '3481', 'total': 84950, 'status': 'PAID'},
            {'id': '3482', 'total': 5100, 'status': 'SHIPPED'},
            {'id': '3483', 'total': 69900, 'status': 'PENDING'},
        ]

    data = {
        'revenue': float(revenue),
        'orders': sales_count,
        'customers': customers
    }
    return JsonResponse(data)


@csrf_exempt
@staff_required
def send_sms_reminder_api(request, customer_id):
    """Mijozning telefon raqamiga Nasiya qarz bo'yicha SMS xabarnoma yuborish."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': "Faqat POST so'rovlar qabul qilinadi"}, status=405)

    try:
        customer = Customer.objects.get(id=customer_id)
        if not customer.phone:
            return JsonResponse({'status': 'error', 'message': "Mijozning telefon raqami kiritilmagan!"}, status=400)
        
        if customer.debt_amount <= Decimal('0.00'):
            return JsonResponse({'status': 'error', 'message': f"{customer.first_name}da nasiya qarz mavjud emas!"}, status=400)

        from .sms_service import send_debt_reminder_sms
        res = send_debt_reminder_sms(customer.first_name, customer.phone, customer.debt_amount)

        if res.get('status') == 'success':
            d_fmt = "{:,.0f}".format(customer.debt_amount).replace(',', ' ')
            CustomerLog.objects.create(
                customer=customer,
                log_type='debt_add',
                title="📱 SMS NASIYA ESLATMASI YUBORILDI",
                message=f"SMS xabarnoma xizmati orqali {d_fmt} so'm nasiya bo'yicha {customer.phone} raqamiga SMS eslatma yuborildi.",
                amount=customer.debt_amount
            )
            return JsonResponse({
                'status': 'success',
                'message': f"{customer.first_name} ({customer.phone}) ga {d_fmt} so'm nasiya bo'yicha SMS eslatma yuborildi!"
            })
        else:
            return JsonResponse({'status': 'error', 'message': res.get('message', "SMS yuborishda xatolik!")}, status=400)

    except Customer.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': "Mijoz topilmadi!"}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
