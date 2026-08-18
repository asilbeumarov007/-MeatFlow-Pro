"""
Baxmal Meat Kassa — Telegram Bot Engine
========================================
Barcha buyruqlar va inline tugmalar uchun handler funksiyalari.
"""
import os
import requests
from decimal import Decimal
from datetime import date, timedelta, datetime
from django.utils import timezone
from django.db.models import Sum, Count, Q, F

BOT_TOKEN = os.environ.get('TELEGRAM_BOT_TOKEN', '')
CHAT_ID = os.environ.get('TELEGRAM_CHAT_ID', '')
API_URL = f"https://api.telegram.org/bot{BOT_TOKEN}"

# ══════════════════════════════════════════════════════════
# TELEGRAM API HELPERS
# ══════════════════════════════════════════════════════════

def send_message(chat_id, text, reply_markup=None, parse_mode='Markdown'):
    """Telegramga xabar yuborish (Markdown xato bo'lsa oddiy matnda yuboradi)."""
    payload = {
        'chat_id': chat_id,
        'text': text,
    }
    if parse_mode:
        payload['parse_mode'] = parse_mode
    if reply_markup:
        payload['reply_markup'] = reply_markup
    try:
        r = requests.post(f"{API_URL}/sendMessage", json=payload, timeout=10)
        res = r.json()
        if not res.get('ok') and parse_mode:
            payload.pop('parse_mode', None)
            r2 = requests.post(f"{API_URL}/sendMessage", json=payload, timeout=10)
            return r2.json()
        return res
    except Exception as e:
        print(f"[TG Bot] Xabar yuborishda xato: {e}")
        return None


def send_document(chat_id, file_path, caption=""):
    """Telegramga hujjat (fayl) yuborish."""
    url = f"{API_URL}/sendDocument"
    try:
        with open(file_path, 'rb') as f:
            files = {'document': f}
            payload = {'chat_id': chat_id, 'caption': caption}
            r = requests.post(url, data=payload, files=files, timeout=20)
            return r.json()
    except Exception as e:
        print(f"[TG Bot] Hujjat yuborishda xato: {e}")
        return None


def answer_callback(callback_query_id, text=""):
    """Callback query ga javob berish (tugma bosilganda loading o'chirish)."""
    try:
        requests.post(f"{API_URL}/answerCallbackQuery", json={
            'callback_query_id': callback_query_id,
            'text': text,
        }, timeout=5)
    except:
        pass


def edit_message(chat_id, message_id, text, reply_markup=None, parse_mode='Markdown'):
    """Mavjud xabarni tahrirlash (tugma bosilganda)."""
    payload = {
        'chat_id': chat_id,
        'message_id': message_id,
        'text': text,
    }
    if parse_mode:
        payload['parse_mode'] = parse_mode
    if reply_markup:
        payload['reply_markup'] = reply_markup
    try:
        r = requests.post(f"{API_URL}/editMessageText", json=payload, timeout=10)
        res = r.json()
        if not res.get('ok') and parse_mode:
            payload.pop('parse_mode', None)
            r2 = requests.post(f"{API_URL}/editMessageText", json=payload, timeout=10)
            return r2.json()
        return res
    except Exception as e:
        print(f"[TG Bot] Xabar tahrirlashda xato: {e}")
        return None


def send_photo(chat_id, photo_path_or_url, caption="", reply_markup=None, parse_mode='Markdown'):
    """Telegramga rasm yuborish (fayl yoki URL orqali, xavfsiz retry bilan)."""
    import json
    url = f"{API_URL}/sendPhoto"
    data = {
        'chat_id': chat_id,
        'caption': caption,
    }
    if parse_mode:
        data['parse_mode'] = parse_mode
    if reply_markup:
        data['reply_markup'] = json.dumps(reply_markup) if isinstance(reply_markup, dict) else reply_markup

    try:
        if os.path.exists(str(photo_path_or_url)):
            with open(photo_path_or_url, 'rb') as f:
                files = {'photo': f}
                r = requests.post(url, data=data, files=files, timeout=20)
                res = r.json()
                if not res.get('ok') and parse_mode:
                    data.pop('parse_mode', None)
                    f.seek(0)
                    r2 = requests.post(url, data=data, files={'photo': f}, timeout=20)
                    return r2.json()
                return res
        else:
            data['photo'] = str(photo_path_or_url)
            r = requests.post(url, data=data, timeout=20)
            res = r.json()
            if not res.get('ok') and parse_mode:
                data.pop('parse_mode', None)
                r2 = requests.post(url, data=data, timeout=20)
                return r2.json()
            return res
    except Exception as e:
        print(f"[TG Bot Photo Send Error]: {e}")
        return None


def send_payment_proof_photo(proof):
    """Mijoz yuklagan to'lov chekini Admin Telegram botiga rasm va tasdiqlash tugmalari bilan yuborish."""
    if not CHAT_ID:
        return None
    customer = proof.customer
    caption = (
        f"📸 *YANGI ONLAYN TO'LOV CHEKI YUKLANDI*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"👤 *Mijoz:* {customer.first_name} {customer.last_name or ''}\n"
        f"📞 *Telefon:* `{customer.phone}`\n"
        f"🆔 *Mijoz ID:* `{customer.custom_id}`\n"
        f"💳 *To'lov tizimi:* {proof.provider.upper()}\n"
        f"💰 *To'langan Summa:* `{proof.amount:,.0f} so'm`\n"
        f"📉 *Joriy Qarzdorlik:* `{customer.debt_amount:,.0f} so'm`\n"
        f"📝 *Izoh:* {proof.note or 'Kiritilmagan'}\n"
        f"⏳ *Holat:* Kutilmoqda (Admin tasdiqlashi lozim)\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"
    )
    reply_markup = {
        'inline_keyboard': [
            [
                {'text': "✅ Qarzni Yopish (Tasdiqlash)", 'callback_data': f"adm_proof_appr_{proof.id}"},
                {'text': "❌ Rad etish", 'callback_data': f"adm_proof_rej_{proof.id}"}
            ]
        ]
    }
    try:
        if proof.image:
            if hasattr(proof.image, 'path') and os.path.exists(proof.image.path):
                return send_photo(CHAT_ID, proof.image.path, caption=caption, reply_markup=reply_markup)
            elif hasattr(proof.image, 'url'):
                site_url = os.environ.get('SITE_URL', 'https://baxmalmeat.uz')
                full_img_url = proof.image.url if proof.image.url.startswith('http') else f"{site_url}{proof.image.url}"
                return send_photo(CHAT_ID, full_img_url, caption=caption, reply_markup=reply_markup)
    except Exception as e:
        print(f"[TG Bot Payment Proof Image Error]: {e}")

    return send_message(CHAT_ID, caption, reply_markup=reply_markup)


# ══════════════════════════════════════════════════════════
# INLINE KEYBOARD BUILDERS
# ══════════════════════════════════════════════════════════

def main_menu_keyboard():
    """Asosiy menyu tugmalari."""
    site_url = os.environ.get('SITE_URL', 'https://baxmalmeat.uz')
    return {
        'inline_keyboard': [
            [
                {'text': '📱 Admin Boshqaruv (Mini App)', 'web_app': {'url': f'{site_url}/pos/admin-mini-app/'}}
            ],
            [
                {'text': '🌙 Kechki Z-Hisobot', 'callback_data': 'cmd_z_report'},
                {'text': '📊 Bugungi Hisobot', 'callback_data': 'cmd_hisobot'},
            ],
            [
                {'text': '🥩 Vitrina & Qurish (AI)', 'callback_data': 'cmd_decay'},
                {'text': '💰 Qarzdorlar', 'callback_data': 'cmd_qarz'},
            ],
            [
                {'text': '📦 Zaxira (Ombor)', 'callback_data': 'cmd_zaxira'},
                {'text': '🚜 Ta\'minotchilar', 'callback_data': 'cmd_taminotchi'},
            ],
            [
                {'text': '🏪 Shift Holati', 'callback_data': 'cmd_shift'},
                {'text': '💎 Bonus Liderlar', 'callback_data': 'cmd_bonus'},
            ],
            [
                {'text': '📅 Kechagi Hisobot', 'callback_data': 'cmd_kecha'},
            ],
        ]
    }


def main_menu_reply_keyboard():
    """Klaviatura o'rnidagi doimiy tugmalar (Reply Keyboard)."""
    site_url = os.environ.get('SITE_URL', 'https://baxmalmeat.uz')
    return {
        'keyboard': [
            [{'text': '📱 Admin Boshqaruv (Mini App)', 'web_app': {'url': f'{site_url}/pos/admin-mini-app/'}}],
            [{'text': '🌙 Kechki Z-Hisobot'}, {'text': '📊 Bugungi Hisobot'}],
            [{'text': '🥩 Vitrina & Qurish (AI)'}, {'text': '💰 Qarzdorlar'}],
            [{'text': '📦 Zaxira (Ombor)'}, {'text': '🚜 Ta\'minotchilar'}],
            [{'text': '🏪 Shift Holati'}, {'text': '💎 Bonus Liderlar'}],
            [{'text': '👥 Mijozlar Ro\'yxati'}, {'text': '📑 Oxirgi Savdolar'}],
            [{'text': '❓ Yordam'}]
        ],
        'resize_keyboard': True,
        'one_time_keyboard': False
    }




def send_customer_excel_backup(chat_id=None):
    """Barcha mijozlarni Excelga yozadi va Telegramga jo'natadi."""
    if not chat_id:
        chat_id = CHAT_ID
        
    from pos.models import Customer
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    try:
        # Create Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mijozlar"
        ws.views.sheetView[0].showGridLines = True
        
        # Style variables (Green & Gold theme)
        header_fill = PatternFill(start_color='1B6B4A', end_color='1B6B4A', fill_type='solid')
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        row_alt_fill = PatternFill(start_color='F3F8F5', end_color='F3F8F5', fill_type='solid')
        
        headers = ["ID", "F.I.Sh", "Telefon", "Qarz (so'm)", "Kredit Limiti", "Bonus ballari", "Ro'yxatdan o'tgan sana", "Qora ro'yxat", "Izoh"]
        ws.row_dimensions[1].height = 28
        
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border
            
        customers = Customer.objects.all().order_by('id')
        for row_idx, c in enumerate(customers, 2):
            ws.row_dimensions[row_idx].height = 20
            fullname = f"{c.first_name} {c.last_name or ''}".strip()
            created_str = c.created_at.strftime('%d.%m.%Y %H:%M') if c.created_at else ''
            blacklist_str = "Ha" if c.is_blacklisted else "Yo'q"
            
            row_data = [
                c.custom_id or f"CUST-{c.id}",
                fullname,
                c.phone or '',
                float(c.debt_amount),
                float(c.debt_limit),
                int(c.bonus_points),
                created_str,
                blacklist_str,
                c.note or ''
            ]
            
            is_alt = (row_idx % 2 == 1)
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                if is_alt:
                    cell.fill = row_alt_fill
                    
                # Format numbers
                if col_idx in [4, 5]:
                    cell.number_format = '#,##0'
                elif col_idx == 6:
                    cell.number_format = '#,##0'
                    
        # Adjust column widths
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        # Create media directory if not exists
        os.makedirs('media', exist_ok=True)
        file_path = os.path.join('media', 'mijozlar_royxati.xlsx')
        wb.save(file_path)
        
        # Send via telegram
        caption = f"👥 *MeatFlow Mijozlar Zaxira Nusxasi*\n\nJami mijozlar soni: {customers.count()} ta\nSana: {timezone.localtime(timezone.now()).strftime('%d.%m.%Y %H:%M')}"
        send_document(chat_id, file_path, caption=caption)
        
    except Exception as e:
        print(f"[TG Bot] Customer Excel generation error: {e}")


def handle_mijozlar(chat_id):
    """Mijozlar ro'yxatini Excel shaklida jo'natadi."""
    send_message(chat_id, "⏳ Mijozlar ro'yxati shakllantirilmoqda, iltimos kuting...")
    send_customer_excel_backup(chat_id=chat_id)


def handle_oxirgi_savdolar(chat_id):
    """Oxirgi 5 ta savdo haqida ma'lumot beradi."""
    from pos.models import Sale
    sales = Sale.objects.all().order_by('-id')[:5]
    if not sales:
        send_message(chat_id, "ℹ️ Tizimda hali savdolar mavjud emas.")
        return
        
    text = "📑 **Oxirgi 5 ta savdo:**\n\n"
    for s in sales:
        time_str = timezone.localtime(s.created_at).strftime('%d.%m.%Y %H:%M')
        cust_name = f"{s.customer.first_name} {s.customer.last_name or ''}".strip() if s.customer else "Anonim Mijoz"
        text += (
            f"🔹 **Chek #{s.id}** ({time_str})\n"
            f"👤 Mijoz: {cust_name}\n"
            f"💰 Jami: {s.total_amount:,.0f} so'm\n"
            f"💵 To'lov: {s.final_paid:,.0f} so'm ({s.get_payment_method_display()})\n"
            f"📉 Chegirma: {s.discount_amount:,.0f} so'm | 💎 Bonus: {s.bonus_used:,.0f} ball\n"
            f"➕ Qarz: {s.debt_added:,.0f} so'm\n\n"
        )
    send_message(chat_id, text)


def handle_oxirgi_soyimlar(chat_id):
    """Oxirgi 5 ta so'yim (chorva qabuli) haqida ma'lumot beradi."""
    from pos.models import Slaughter
    slaughters = Slaughter.objects.all().order_by('-id')[:5]
    if not slaughters:
        send_message(chat_id, "ℹ️ Tizimda hali so'yimlar mavjud emas.")
        return
        
    text = "🐂 **Oxirgi 5 ta so'yim qabuli:**\n\n"
    for s in slaughters:
        time_str = timezone.localtime(s.created_at).strftime('%d.%m.%Y %H:%M')
        partner = ""
        if s.supplier:
            partner = f"Ta'minotchi: {s.supplier.first_name}"
        elif s.customer:
            partner = f"Mijoz: {s.customer.first_name}"
        else:
            partner = "Nomalum hamkor"
            
        animal_type_str = "Mol" if s.animal_type == 'mol' else "Qo'y"
        text += (
            f"🔹 **So'yim #{s.id}** ({time_str})\n"
            f"🐄 Turi: {animal_type_str}\n"
            f"⚖️ Og'irlik: {s.total_weight:.3f} kg\n"
            f"💰 Narxi: {s.purchase_price_per_kg:,.0f} so'm/kg\n"
            f"🤝 Jami qiymat: {s.total_cost:,.0f} so'm\n"
            f"👤 Hamkor: {partner}\n\n"
        )
    send_message(chat_id, text)


def back_button_keyboard():
    """Orqaga qaytish tugmasi."""
    return {
        'inline_keyboard': [
            [{'text': '🔙 Asosiy Menyu', 'callback_data': 'cmd_menu'}],
        ]
    }


# ══════════════════════════════════════════════════════════
# COMMAND HANDLERS
# ══════════════════════════════════════════════════════════

def handle_start(chat_id):
    """Botni ishga tushirish — salomlashuv va asosiy menyu."""
    text = (
        "🥩 *Baxmal Meat Kassa Bot*\n\n"
        "Assalomu alaykum! Men sizning kassa yordamchingizman.\n"
        "Pastdagi tugmalar orqali kerakli ma'lumotlarni tezkor olishingiz mumkin:"
    )
    send_message(chat_id, text, reply_markup=main_menu_reply_keyboard())


def handle_hisobot(chat_id, target_date=None, message_id=None):
    """Savdolar xulosasi — bugungi yoki berilgan sanadagi."""
    from pos.models import Sale, SaleItem, CashierShift

    if target_date is None:
        target_date = timezone.localdate()

    day_label = "📊 *Bugungi" if target_date == timezone.localdate() else f"📅 *{target_date.strftime('%d.%m.%Y')}"
    
    sales = Sale.objects.filter(created_at__date=target_date)
    total_count = sales.count()

    if total_count == 0:
        text = f"{day_label} Savdo Hisoboti*\n\n😔 Bu kunda hech qanday savdo qilinmagan."
        kb = back_button_keyboard()
        if message_id:
            edit_message(chat_id, message_id, text, reply_markup=kb)
        else:
            send_message(chat_id, text, reply_markup=kb)
        return

    total_sum = sales.aggregate(s=Sum('total_amount'))['s'] or Decimal('0')
    total_paid = sales.aggregate(s=Sum('final_paid'))['s'] or Decimal('0')
    total_debt = sales.aggregate(s=Sum('debt_added'))['s'] or Decimal('0')
    total_bonus = sales.aggregate(s=Sum('bonus_used'))['s'] or Decimal('0')
    total_discount = sales.aggregate(s=Sum('discount_amount'))['s'] or Decimal('0')

    naqd_count = sales.filter(payment_method='naqd').count()
    karta_count = sales.filter(payment_method='karta').count()
    nasiya_count = sales.filter(payment_method='nasiya').count()

    # Mahsulot taqsimoti
    items = SaleItem.objects.filter(sale__created_at__date=target_date)
    product_stats = items.values('product__name').annotate(
        total_weight=Sum('weight'),
        total_sum=Sum('item_total')
    ).order_by('-total_sum')

    products_text = ""
    for p in product_stats[:8]:
        name = p['product__name']
        kg = p['total_weight'] or 0
        sm = p['total_sum'] or 0
        products_text += f"  • {name}: *{kg:.1f} kg* — {int(sm):,} so'm\n"

    text = (
        f"{day_label} Savdo Hisoboti*\n"
        f"{'━' * 28}\n\n"
        f"🛒 *Jami savdolar:* {total_count} ta\n"
        f"💵 *Jami summa:* {int(total_sum):,} so'm\n"
        f"✅ *To'langan:* {int(total_paid):,} so'm\n"
        f"📋 *Nasiyaga:* {int(total_debt):,} so'm\n"
        f"💎 *Bonus sarflangan:* {int(total_bonus):,} so'm\n"
        f"🏷️ *Chegirma:* {int(total_discount):,} so'm\n\n"
        f"📌 *To'lov usullari:*\n"
        f"  💵 Naqd: {naqd_count} ta\n"
        f"  💳 Karta: {karta_count} ta\n"
        f"  📝 Nasiya: {nasiya_count} ta\n\n"
        f"📦 *Sotilgan mahsulotlar:*\n"
        f"{products_text}"
    )

    kb = back_button_keyboard()
    if message_id:
        edit_message(chat_id, message_id, text, reply_markup=kb)
    else:
        send_message(chat_id, text, reply_markup=kb)


def handle_qarz(chat_id, message_id=None):
    """Top 10 eng ko'p qarzdor mijozlar."""
    from pos.models import Customer

    debtors = Customer.objects.filter(debt_amount__gt=0).order_by('-debt_amount')[:10]

    if not debtors:
        text = "💰 *Qarzdorlar Ro'yxati*\n\n✅ Hech kim qarzdor emas! Barcha hisob-kitoblar to'liq."
    else:
        total_debt = Customer.objects.filter(debt_amount__gt=0).aggregate(s=Sum('debt_amount'))['s'] or 0
        debtor_count = Customer.objects.filter(debt_amount__gt=0).count()

        text = (
            f"💰 *Top Qarzdor Mijozlar*\n"
            f"{'━' * 28}\n\n"
            f"📊 Jami qarzdorlar: *{debtor_count}* ta\n"
            f"💵 Jami nasiya: *{int(total_debt):,}* so'm\n\n"
        )
        for i, c in enumerate(debtors, 1):
            emoji = "🔴" if c.debt_amount > 500000 else "🟡" if c.debt_amount > 200000 else "🟢"
            name = f"{c.first_name} {c.last_name or ''}"
            text += f"{emoji} *{i}.* {name} (`{c.custom_id}`)\n    💳 {int(c.debt_amount):,} so'm / Limit: {int(c.debt_limit):,} so'm\n"

    kb = back_button_keyboard()
    if message_id:
        edit_message(chat_id, message_id, text, reply_markup=kb)
    else:
        send_message(chat_id, text, reply_markup=kb)


def handle_zaxira(chat_id, message_id=None):
    """Ombordagi barcha mahsulotlar zaxirasi."""
    from pos.models import Stock

    stocks = Stock.objects.select_related('product').filter(product__is_active=True).order_by('product__name')

    if not stocks:
        text = "📦 *Ombor Zaxiralari*\n\n😔 Hech qanday mahsulot topilmadi."
    else:
        text = f"📦 *Ombor Zaxiralari*\n{'━' * 28}\n\n"
        for s in stocks:
            qty = s.quantity
            if qty <= 0:
                emoji = "🔴"
                status = "TUGAGAN"
            elif qty < 5:
                emoji = "🟡"
                status = "KAM"
            elif qty < 20:
                emoji = "🟢"
                status = ""
            else:
                emoji = "✅"
                status = ""
            
            status_text = f" ⚠️ *{status}*" if status else ""
            text += f"{emoji} *{s.product.name}:* {qty:.1f} kg{status_text}\n"

    kb = back_button_keyboard()
    if message_id:
        edit_message(chat_id, message_id, text, reply_markup=kb)
    else:
        send_message(chat_id, text, reply_markup=kb)


def handle_taminotchi(chat_id, message_id=None):
    """Ta'minotchilar balansi."""
    from pos.models import Supplier

    suppliers = Supplier.objects.all().order_by('-our_debt')

    if not suppliers:
        text = "🚜 *Ta'minotchilar*\n\n😔 Hech qanday ta'minotchi topilmadi."
    else:
        total_debt = suppliers.aggregate(s=Sum('our_debt'))['s'] or 0
        text = (
            f"🚜 *Ta'minotchilar Balansi*\n"
            f"{'━' * 28}\n\n"
            f"💵 Bizning jami qarzimiz: *{int(total_debt):,}* so'm\n\n"
        )
        for s in suppliers:
            emoji = "🔴" if s.our_debt > 1000000 else "🟡" if s.our_debt > 0 else "✅"
            name = f"{s.first_name} {s.last_name or ''}"
            if s.our_debt > 0:
                text += f"{emoji} *{name}* (`{s.custom_id}`)\n    Qarzimiz: {int(s.our_debt):,} so'm\n"
            else:
                text += f"{emoji} *{name}* — Qarz yo'q ✅\n"

    kb = back_button_keyboard()
    if message_id:
        edit_message(chat_id, message_id, text, reply_markup=kb)
    else:
        send_message(chat_id, text, reply_markup=kb)


def handle_shift(chat_id, message_id=None):
    """Hozirgi kassa shifti holati."""
    from pos.models import CashierShift

    active = CashierShift.objects.filter(is_open=True).first()

    if active:
        opened = timezone.localtime(active.opened_at).strftime('%H:%M — %d.%m.%Y')
        sales_count = active.sales.count() if hasattr(active, 'sales') else 0
        sales_sum = active.sales.aggregate(s=Sum('final_paid'))['s'] or 0 if hasattr(active, 'sales') else 0
        text = (
            f"🏪 *Kassa Shifti Holati*\n"
            f"{'━' * 28}\n\n"
            f"🟢 *SHIFT OCHIQ*\n\n"
            f"👤 Kassir: *{active.cashier.username}*\n"
            f"🕐 Ochilgan: {opened}\n"
            f"🛒 Savdolar: *{sales_count}* ta\n"
            f"💵 Jami tushumlar: *{int(sales_sum):,}* so'm"
        )
    else:
        last = CashierShift.objects.filter(is_open=False).order_by('-closed_at').first()
        if last:
            closed = timezone.localtime(last.closed_at).strftime('%H:%M — %d.%m.%Y') if last.closed_at else "—"
            text = (
                f"🏪 *Kassa Shifti Holati*\n"
                f"{'━' * 28}\n\n"
                f"🔴 *SHIFT YOPILGAN*\n\n"
                f"Oxirgi shift: {closed}\n"
                f"Kassir: *{last.cashier.username}*"
            )
        else:
            text = "🏪 *Kassa Shifti Holati*\n\n🔴 Hali hech qanday shift ochilmagan."

    kb = back_button_keyboard()
    if message_id:
        edit_message(chat_id, message_id, text, reply_markup=kb)
    else:
        send_message(chat_id, text, reply_markup=kb)


def handle_bonus(chat_id, message_id=None):
    """Top 10 eng ko'p bonus to'plagan mijozlar."""
    from pos.models import Customer

    top = Customer.objects.filter(bonus_points__gt=0).order_by('-bonus_points')[:10]

    if not top:
        text = "💎 *Bonus Liderlar*\n\n😔 Hech kim bonus to'plamagan."
    else:
        text = f"💎 *Top Bonus Liderlar*\n{'━' * 28}\n\n"
        for i, c in enumerate(top, 1):
            medal = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else f"  {i}."
            name = f"{c.first_name} {c.last_name or ''}"
            text += f"{medal} *{name}* — {c.bonus_points:,} ball\n"

    kb = back_button_keyboard()
    if message_id:
        edit_message(chat_id, message_id, text, reply_markup=kb)
    else:
        send_message(chat_id, text, reply_markup=kb)


def handle_yordam(chat_id, message_id=None):
    """Yordam — barcha buyruqlar ro'yxati."""
    text = (
        "❓ *Yordam — Buyruqlar Ro'yxati*\n"
        f"{'━' * 28}\n\n"
        "📊 `/hisobot` — Bugungi savdo xulosasi\n"
        "💰 `/qarz` — Top qarzdor mijozlar\n"
        "📦 `/zaxira` — Ombor zaxiralari\n"
        "🚜 `/taminotchi` — Ta'minotchilar balansi\n"
        "🏪 `/shift` — Kassa shifti holati\n"
        "💎 `/bonus` — Bonus liderlar\n"
        "👥 `/mijozlar` — Mijozlar ro'yxati (Excel fayl)\n"
        "📑 `/savdolar` — Oxirgi 5 ta savdo cheklari\n"
        "🐮 `/soyimlar` — Oxirgi 5 ta so'yim qabullari\n"
        "📅 `/savdo 2026-07-20` — Sanadagi savdolar\n"
        "❓ `/yordam` — Shu ro'yxat\n\n"
        "💡 *Shuningdek pastdagi doimiy tugmalardan ham foydalanishingiz mumkin!*"
    )
    kb = back_button_keyboard()
    if message_id:
        edit_message(chat_id, message_id, text, reply_markup=kb)
    else:
        send_message(chat_id, text, reply_markup=kb)


# ══════════════════════════════════════════════════════════
# DAILY REPORT GENERATOR
# ══════════════════════════════════════════════════════════

def generate_daily_report(target_date=None):
    """Kunlik to'liq hisobotni generatsiya qilish va guruhga yuborish."""
    from pos.models import Sale, SaleItem, Customer, Stock, Supplier, CashTransaction

    if target_date is None:
        target_date = timezone.localdate()

    sales = Sale.objects.filter(created_at__date=target_date)
    total_count = sales.count()
    total_sum = sales.aggregate(s=Sum('total_amount'))['s'] or Decimal('0')
    total_paid = sales.aggregate(s=Sum('final_paid'))['s'] or Decimal('0')
    total_debt = sales.aggregate(s=Sum('debt_added'))['s'] or Decimal('0')

    # To'lov usullari
    naqd_sum = sales.filter(payment_method='naqd').aggregate(s=Sum('final_paid'))['s'] or 0
    karta_sum = sales.filter(payment_method='karta').aggregate(s=Sum('final_paid'))['s'] or 0
    nasiya_sum = sales.filter(payment_method='nasiya').aggregate(s=Sum('debt_added'))['s'] or 0

    # Mahsulotlar
    items = SaleItem.objects.filter(sale__created_at__date=target_date)
    product_stats = items.values('product__name').annotate(
        total_weight=Sum('weight'),
        total_sum=Sum('item_total')
    ).order_by('-total_sum')

    products_text = ""
    for p in product_stats:
        name = p['product__name']
        kg = p['total_weight'] or 0
        sm = p['total_sum'] or 0
        products_text += f"  • {name}: {kg:.1f} kg — {int(sm):,} so'm\n"

    # Ombor holati
    stocks = Stock.objects.select_related('product').filter(product__is_active=True)
    low_stocks = []
    for s in stocks:
        if s.quantity < 5:
            low_stocks.append(f"  ⚠️ {s.product.name}: {s.quantity:.1f} kg")

    stock_warning = ""
    if low_stocks:
        stock_warning = "\n\n🚨 *KAM ZAXIRA OGOHLANTIRISHLARI:*\n" + "\n".join(low_stocks)

    # Jami qarz holati
    total_all_debt = Customer.objects.filter(debt_amount__gt=0).aggregate(s=Sum('debt_amount'))['s'] or 0
    debtor_count = Customer.objects.filter(debt_amount__gt=0).count()

    # Kassaga tushum
    cash_in = CashTransaction.objects.filter(
        created_at__date=target_date,
        transaction_type='in'
    ).aggregate(s=Sum('amount'))['s'] or 0
    cash_out = CashTransaction.objects.filter(
        created_at__date=target_date,
        transaction_type='out'
    ).aggregate(s=Sum('amount'))['s'] or 0

    date_str = target_date.strftime('%d.%m.%Y')
    text = (
        f"📋 *BAXMAL MEAT — KUNLIK HISOBOT*\n"
        f"📅 *{date_str}*\n"
        f"{'━' * 30}\n\n"
        f"🛒 *SAVDOLAR:*\n"
        f"  Jami: *{total_count}* ta savdo\n"
        f"  Umumiy summa: *{int(total_sum):,}* so'm\n"
        f"  To'langan: *{int(total_paid):,}* so'm\n"
        f"  Nasiyaga: *{int(total_debt):,}* so'm\n\n"
        f"💳 *TO'LOV USULLARI:*\n"
        f"  💵 Naqd: {int(naqd_sum):,} so'm\n"
        f"  💳 Karta: {int(karta_sum):,} so'm\n"
        f"  📝 Nasiya: {int(nasiya_sum):,} so'm\n\n"
        f"📦 *SOTILGAN MAHSULOTLAR:*\n"
        f"{products_text}\n"
        f"🏦 *KASSA OQIMI:*\n"
        f"  ⬆️ Kirim: {int(cash_in):,} so'm\n"
        f"  ⬇️ Chiqim: {int(cash_out):,} so'm\n\n"
        f"💰 *UMUMIY NASIYA HOLATI:*\n"
        f"  Qarzdorlar: *{debtor_count}* ta\n"
        f"  Jami nasiya: *{int(total_all_debt):,}* so'm"
        f"{stock_warning}"
    )

    send_message(CHAT_ID, text)
    return text


# ══════════════════════════════════════════════════════════
# LOW STOCK ALERT (savdodan keyin chaqiriladi)
# ══════════════════════════════════════════════════════════

def check_low_stock_alert(product_name, remaining_qty):
    """Agar mahsulot zaxirasi 5 kg dan past bo'lsa, Telegramga ogohlantirish yuborish."""
    if remaining_qty < 5 and remaining_qty >= 0:
        text = (
            f"🚨 *KAM ZAXIRA OGOHLANTIRISHLARI!*\n\n"
            f"📦 *{product_name}* zaxirasi juda kam!\n"
            f"📉 Qolgan: *{remaining_qty:.1f} kg*\n\n"
            f"Iltimos, yangi partiya qo'shing yoki ta'minotchiga buyurtma bering."
        )
        send_message(CHAT_ID, text)


# ══════════════════════════════════════════════════════════
# DISPATCHER — Buyruq va callback'larni yo'naltirish
# ══════════════════════════════════════════════════════════

def handle_reklama(chat_id, text_raw):
    """Barcha mijozlarga ommaviy reklama/bildirishnoma yuborish."""
    parts = text_raw.split(maxsplit=1)
    if len(parts) < 2:
        msg = (
            "📢 *OMMAVIY REKLAMA YUBORISH*\n"
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            "Foydalanish: `/reklama [Sizning reklama yoki e'lon matningiz]`\n\n"
            "Misol:\n"
            "`/reklama 🥩 Bugun do'konimizga yangi so'yilgan yosh mol go'shti keldi! Zaxira chegaralangan, bot orqali buyurtma bering.`"
        )
        send_message(chat_id, msg)
        return

    broadcast_text = parts[1].strip()
    from pos.models import Customer
    from pos.customer_bot import send_message as send_cust_msg

    customers = Customer.objects.filter(telegram_chat_id__isnull=False).exclude(telegram_chat_id='')
    total_count = customers.count()
    if total_count == 0:
        send_message(chat_id, "ℹ️ Telegram botga ulangan mijozlar topilmadi.")
        return

    sent_count = 0
    fail_count = 0

    kb = {
        'inline_keyboard': [
            [{'text': '🛒 Hozir Buyurtma Berish', 'callback_data': 'cmd_order'}]
        ]
    }

    send_message(chat_id, f"⏳ {total_count} ta mijozga reklama yuborilmoqda...")

    for cust in customers:
        try:
            res = send_cust_msg(cust.telegram_chat_id, f"📢 *BAXMAL MEAT — E'LON*\n\n{broadcast_text}", reply_markup=kb)
            if res and res.get('ok'):
                sent_count += 1
            else:
                fail_count += 1
        except Exception:
            fail_count += 1

    report = (
        f"✅ *OMMAVIY REKLAMA YUBORILDI*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"👥 *Jami mijozlar:* {total_count} ta\n"
        f"🟢 *Muvaffaqiyatli yetkazildi:* {sent_count} ta\n"
        f"🔴 *Yetkazilmadi (bloklangan):* {fail_count} ta"
    )
    send_message(chat_id, report)


def handle_decay(chat_id, message_id=None):
    """Vitrinada 2 kundan ortiq turgan go'sht partiyalari bo'yicha AI tavsiya."""
    from pos.models import StockBatch
    batches = StockBatch.objects.filter(current_quantity__gt=Decimal('0.05')).select_related('product').order_by('created_at')
    
    ai_recs = []
    for b in batches:
        rec = b.get_ai_recommendation()
        if rec:
            ai_recs.append(rec)
            
    if not ai_recs:
        text = (
            "🥩 *VITRINA & ZAXIRA HOLATI*\n\n"
            "✅ *Barcha partiyalar yangi!* Vitrinada 2 kundan ortiq turib qolgan go'sht partiyalari mavjud emas."
        )
    else:
        text = (
            "🥩 *VITRINA & QURISH ZARARI — AI TAVSIYALARI*\n"
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            "Do'kon mudiriga diqqat! Vitrinada 2 kundan ortiq turgan partiyalar aniqlandi:\n\n"
        )
        for idx, w in enumerate(ai_recs, 1):
            text += (
                f"{idx}. ⚠️ *{w['product_name']}* (Partiya #{w['batch_id']})\n"
                f"   ⏳ *Turgan vaqti:* `{w['days']} kun` | 📦 *Qoldiq:* `{w['quantity']:.2f} kg`\n"
                f"   📉 *Taxminiy yo'qotish:* `{w['decay_loss_kg']:.3f} kg` (~`{w['decay_loss_sum']:,.0f}` so'm)\n"
                f"   💡 *AI Maslahat:* _{w['message']}_\n\n"
            )
        text += (
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            "⚡ *Tavsiya:* Zararni kamaytirish uchun ushbu partiyalarni qiymaga aylantiring yoki maxsus chegirma e'lon qiling."
        )
        
    kb = back_button_keyboard()
    if message_id:
        edit_message(chat_id, message_id, text, reply_markup=kb)
    else:
        send_message(chat_id, text, reply_markup=kb)


def dispatch_command(chat_id, text_raw):
    """Matnli buyruqlarni tegishli handlerga yo'naltirish."""
    text = text_raw.strip().lower()

    # Remove @bot_username if present
    if '@' in text:
        text = text.split('@')[0]

    if text == '/start':
        handle_start(chat_id)
    elif text in ['/z_report', '🌙 kechki z-hisobot', '/digest']:
        send_daily_executive_digest(chat_id=chat_id)
    elif text in ['/decay', '/qurish', '🥩 vitrina nazorati', '📉 vitrina & qurish']:
        handle_decay(chat_id)
    elif text in ['/hisobot', '📊 bugungi hisobot']:
        handle_hisobot(chat_id)


    elif text in ['/qarz', '💰 qarzdorlar']:
        handle_qarz(chat_id)
    elif text in ['/zaxira', '📦 zaxira (ombor)', '📦 zaxira']:
        handle_zaxira(chat_id)
    elif text in ['/taminotchi', '/taminotchilar', '🚜 ta\'minotchilar']:
        handle_taminotchi(chat_id)
    elif text in ['/shift', '🏪 shift holati']:
        handle_shift(chat_id)
    elif text in ['/bonus', '💎 bonus liderlar']:
        handle_bonus(chat_id)
    elif text in ['/mijozlar', '👥 mijozlar ro\'yxati']:
        handle_mijozlar(chat_id)
    elif text in ['/savdolar', '📑 oxirgi savdolar']:
        handle_oxirgi_savdolar(chat_id)
    elif text in ['/soyimlar', '🐮 oxirgi so\'yimlar']:
        handle_oxirgi_soyimlar(chat_id)
    elif text in ['/yordam', '/help', '❓ yordam']:
        handle_yordam(chat_id)
    elif text.startswith('/reklama') or text.startswith('/broadcast'):
        handle_reklama(chat_id, text_raw)
    elif text == '📅 kechagi hisobot':
        yesterday = timezone.localdate() - timedelta(days=1)
        handle_hisobot(chat_id, target_date=yesterday)
    elif text.startswith('/savdo'):
        parts = text_raw.strip().split()
        if len(parts) >= 2:
            try:
                target = datetime.strptime(parts[1], '%Y-%m-%d').date()
                handle_hisobot(chat_id, target_date=target)
            except ValueError:
                send_message(chat_id, "❌ Noto'g'ri sana formati!\n\nTo'g'ri format: `/savdo 2026-07-20`")
        else:
            handle_hisobot(chat_id)
    else:
        # 🎙️ GEMINI OVOZLI VA MATNLI AI QASSOB / BIZNES YORDAMCHISI
        send_message(chat_id, "🧠 *AI Qassob tahlil qilmoqda...*")
        from .voice_ai_service import query_gemini_ai_qassob
        reply = query_gemini_ai_qassob(user_prompt=text_raw)
        send_message(chat_id, reply)


def dispatch_voice_message(chat_id, voice_file_id):
    """Do'kon egasining ovozli xabarini qabul qilib, Gemini AI orqali tahlil qilish."""
    send_message(chat_id, "🎙️ *Ovozingiz tinglanmoqda va do'kon ma'lumotlari tahlil qilinmoqda...*")
    from .voice_ai_service import handle_telegram_voice_message
    ai_reply = handle_telegram_voice_message(chat_id, voice_file_id)
    send_message(chat_id, ai_reply)


def edit_photo_caption(chat_id, message_id, caption, reply_markup=None, parse_mode='HTML'):
    """Foto xabar tagidagi matnni tahrirlash va tugmalarni yangilash/o'chirish."""
    import json
    payload = {
        'chat_id': chat_id,
        'message_id': message_id,
        'caption': caption,
        'parse_mode': parse_mode,
        'reply_markup': json.dumps(reply_markup if reply_markup is not None else {'inline_keyboard': []})
    }
    try:
        r = requests.post(f"{API_URL}/editMessageCaption", json=payload, timeout=10)
        return r.json()
    except Exception as e:
        print(f"[TG Bot editMessageCaption Error]: {e}")
        return None


def handle_proof_approval(chat_id, message_id, proof_id, callback_query_id):
    """To'lov chekini tasdiqlab qarzni yopish."""
    try:
        from pos.models import PaymentProof, CashTransaction, CustomerLog
        from django.contrib.auth import get_user_model
        User = get_user_model()
        admin_user = User.objects.filter(is_superuser=True).first()

        proof = PaymentProof.objects.get(id=proof_id)
        if proof.is_verified:
            answer_callback(callback_query_id, "Ushbu to'lov allaqachon tasdiqlangan!")
            return

        customer = proof.customer
        proof.is_verified = True
        proof.save()

        # Deduct debt
        customer.debt_amount -= proof.amount
        customer.save()

        # Record cash transaction
        CashTransaction.objects.create(
            transaction_type='in',
            amount=proof.amount,
            category='debt_pay',
            payment_method='karta' if str(proof.provider).lower() in ['click', 'payme', 'karta', 'card'] else 'naqd',
            description=f"Onlayn to'lov ({proof.provider.upper()}) tasdiqlandi. Mijoz: {customer.first_name} ({customer.custom_id})",
            created_by=admin_user,
            customer=customer
        )

        # Record customer log
        CustomerLog.objects.create(
            customer=customer,
            log_type='debt_pay',
            title=f"💳 To'lov Tasdiqlandi ({proof.provider.upper()})",
            message=f"To'lovingiz ({proof.amount:,.0f} so'm) admin tomonidan tasdiqlandi va qarzingizdan yopildi. Qolgan qarz: {customer.debt_amount:,.0f} so'm",
            amount=proof.amount
        )

        # Send customer Telegram push
        if customer.telegram_chat_id:
            try:
                from pos.customer_bot import send_message as send_cust_msg
                cust_text = (
                    f"🎉 *TO'LOVINGIZ TASDIQLANDI!*\n"
                    f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                    f"💰 *Qabul qilingan summa:* `{proof.amount:,.0f} so'm`\n"
                    f"💳 *Tizim:* {proof.provider.upper()}\n"
                    f"📉 *Qolgan qarzingiz:* `{customer.debt_amount:,.0f} so'm`\n\n"
                    f"Rahmat! Xaridingiz barakali bo'lsin! 🥩✨"
                )
                send_cust_msg(customer.telegram_chat_id, cust_text)
            except Exception as tg_cust_err:
                print(f"[Cust TG Error]: {tg_cust_err}")

        success_caption = (
            f"✅ <b>TO'LOV TASDIQLANDI VA QARZDAN YOPILDI!</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"👤 <b>Mijoz:</b> {customer.first_name} ({customer.phone})\n"
            f"💰 <b>To'langan summa:</b> {proof.amount:,.0f} so'm\n"
            f"📉 <b>Qolgan qarz:</b> {customer.debt_amount:,.0f} so'm\n"
            f"⚡ <b>Kassa tushumi qayd etildi.</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"
        )
        edit_photo_caption(chat_id, message_id, success_caption, reply_markup={'inline_keyboard': []})
        answer_callback(callback_query_id, "✅ To'lov tasdiqlandi va qarz yopildi!")
    except Exception as e:
        print(f"[Proof Approve Error]: {e}")
        answer_callback(callback_query_id, f"Xato: {str(e)[:50]}")


def handle_proof_rejection(chat_id, message_id, proof_id, callback_query_id):
    """To'lov chekini rad etish."""
    try:
        from pos.models import PaymentProof, CustomerLog
        proof = PaymentProof.objects.get(id=proof_id)
        customer = proof.customer

        CustomerLog.objects.create(
            customer=customer,
            log_type='debt_pay',
            title="❌ To'lov Cheki Rad Etildi",
            message=f"{proof.amount:,.0f} so'mlik to'lov cheki admin tomonidan qabul qilinmadi. Iltimos, qayta tekshiring.",
            amount=Decimal('0.00')
        )

        if customer.telegram_chat_id:
            try:
                from pos.customer_bot import send_message as send_cust_msg
                cust_text = (
                    f"❌ *TO'LOV CHEKI RAD ETILDI*\n"
                    f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                    f"💰 *Summa:* `{proof.amount:,.0f} so'm`\n"
                    f"Iltimos, to'lov chekini qayta yuklang yoki do'kon bilan bog'laning."
                )
                send_cust_msg(customer.telegram_chat_id, cust_text)
            except Exception as tg_cust_err:
                pass

        reject_caption = (
            f"❌ <b>TO'LOV CHEKI RAD ETILDI</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"👤 <b>Mijoz:</b> {customer.first_name} ({customer.phone})\n"
            f"💰 <b>Summa:</b> {proof.amount:,.0f} so'm\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"
        )
        edit_photo_caption(chat_id, message_id, reject_caption, reply_markup={'inline_keyboard': []})
        answer_callback(callback_query_id, "❌ To'lov cheki rad etildi.")
    except Exception as e:
        print(f"[Proof Reject Error]: {e}")
        answer_callback(callback_query_id, f"Xato: {str(e)[:50]}")


def safe_edit_message_or_caption(chat_id, message_id, text, reply_markup=None):
    """Matn yoki rasm xabarini xavfsiz tahrirlash."""
    res = edit_message(chat_id, message_id, text, reply_markup=reply_markup)
    if not res or not res.get('ok'):
        return edit_photo_caption(chat_id, message_id, text, reply_markup=reply_markup, parse_mode='HTML')
    return res


def dispatch_callback(chat_id, message_id, callback_data, callback_query_id):
    """Inline tugma callback'larini tegishli handlerga yo'naltirish."""
    answer_callback(callback_query_id, "⏳ Bajarilmoqda...")
    from pos.models import B2BOrder, CustomerLog
    from decimal import Decimal

    if callback_data.startswith('adm_appr_'):
        order_id = int(callback_data.replace('adm_appr_', ''))
        try:
            order = B2BOrder.objects.get(id=order_id)
            order.status = 'approved'
            order.save()
            
            CustomerLog.objects.create(
                customer=order.customer,
                log_type='bonus',
                title="Do'kon xabari",
                message=f"✅ Buyurtmangiz (#{order.id}) do'kon tomonidan tasdiqlandi va tayyorlanmoqda!",
                details={'source': 'telegram_admin', 'order_id': order.id, 'new_status': 'approved'},
                amount=Decimal('0.00')
            )
            notify_customer_order_status(order, "✅ Admin tomonidan tasdiqlandi")
            
            next_kb = {
                'inline_keyboard': [
                    [
                        {'text': '🥩 Qadoqlashga o\'tkazish', 'callback_data': f'adm_prep_{order_id}'},
                        {'text': '🚚 Kuryerga berish', 'callback_data': f'adm_ship_{order_id}'}
                    ],
                    [
                        {'text': '❌ Rad etish', 'callback_data': f'adm_rej_{order_id}'}
                    ]
                ]
            }
            caption = (
                f"✅ <b>BUYURTMA #{order.id} TASDIQLANDI!</b>\n"
                f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                f"👤 <b>Mijoz:</b> {order.customer.first_name} ({order.customer.phone})\n"
                f"📦 <b>Mahsulot:</b> {order.product.name} ({order.requested_weight} kg)\n"
                f"💰 <b>Summa:</b> {int(order.requested_weight * order.product.price_per_kg):,} so'm\n\n"
                f"<i>Mijoz Telegramiga tasdiqlash xabari yuborildi.</i>"
            )
            safe_edit_message_or_caption(chat_id, message_id, caption, reply_markup=next_kb)
            answer_callback(callback_query_id, "✅ Buyurtma tasdiqlandi!")
        except Exception as e:
            send_message(chat_id, f"⚠️ Xato: {e}")

    elif callback_data.startswith('adm_prep_'):
        order_id = int(callback_data.replace('adm_prep_', ''))
        try:
            order = B2BOrder.objects.get(id=order_id)
            order.status = 'preparing'
            order.save()
            
            CustomerLog.objects.create(
                customer=order.customer,
                log_type='bonus',
                title="Do'kon xabari",
                message=f"🥩 Buyurtmangiz (#{order.id}) go'sht tortilib qadoqlanmoqda!",
                details={'source': 'telegram_admin', 'order_id': order.id, 'new_status': 'preparing'},
                amount=Decimal('0.00')
            )
            notify_customer_order_status(order, "🥩 Go'sht tortilmoqda / Qadoqlanmoqda")
            
            next_kb = {
                'inline_keyboard': [
                    [{'text': '🚚 Kuryerga berish', 'callback_data': f'adm_ship_{order_id}'}],
                    [{'text': '❌ Rad etish', 'callback_data': f'adm_rej_{order_id}'}]
                ]
            }
            caption = (
                f"🥩 <b>BUYURTMA #{order.id} QADOQLANMOQDA!</b>\n"
                f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                f"👤 <b>Mijoz:</b> {order.customer.first_name} ({order.customer.phone})\n"
                f"📦 <b>Mahsulot:</b> {order.product.name} ({order.requested_weight} kg)"
            )
            safe_edit_message_or_caption(chat_id, message_id, caption, reply_markup=next_kb)
            answer_callback(callback_query_id, "🥩 Qadoqlashga o'tkazildi!")
        except Exception as e:
            send_message(chat_id, f"⚠️ Xato: {e}")

    elif callback_data.startswith('adm_ship_'):
        order_id = int(callback_data.replace('adm_ship_', ''))
        try:
            order = B2BOrder.objects.get(id=order_id)
            order.status = 'shipping'
            order.save()
            
            CustomerLog.objects.create(
                customer=order.customer,
                log_type='bonus',
                title="Do'kon xabari",
                message=f"🚚 Buyurtmangiz (#{order.id}) kuryerga topshirildi va manzil tomon yo'lga chiqdi!",
                details={'source': 'telegram_admin', 'order_id': order.id, 'new_status': 'shipping'},
                amount=Decimal('0.00')
            )
            notify_customer_order_status(order, "🚚 Kuryer yo'lda (Yetkazib berilmoqda)")
            
            nav_buttons = []
            if order.latitude and order.longitude:
                nav_buttons.append({'text': '🚗 Yandex Navigator', 'url': f"https://yandex.uz/maps/?rtext=~{order.latitude},{order.longitude}"})
                nav_buttons.append({'text': '🗺️ Google Maps', 'url': f"https://maps.google.com/?q={order.latitude},{order.longitude}"})
            elif order.delivery_address:
                import urllib.parse
                encoded_addr = urllib.parse.quote(order.delivery_address)
                nav_buttons.append({'text': '🚗 Yandex Xarita', 'url': f"https://yandex.uz/maps/?text={encoded_addr}"})

            call_btn = []
            if order.customer.phone:
                clean_phone = order.customer.phone.replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
                call_btn.append({'text': f"📞 Tel: {order.customer.phone}", 'url': f"tel:{clean_phone}"})

            inline_rows = [
                [{'text': '🎉 Topshirildi (Yakunlash)', 'callback_data': f'adm_comp_{order_id}'}]
            ]
            if nav_buttons:
                inline_rows.append(nav_buttons)
            if call_btn:
                inline_rows.append(call_btn)

            next_kb = {'inline_keyboard': inline_rows}
            caption = (
                f"🚚 <b>BUYURTMA #{order.id} KURYERGA TOPSHIRILDI!</b>\n"
                f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                f"👤 <b>Mijoz:</b> {order.customer.first_name} ({order.customer.phone})\n"
                f"📦 <b>Mahsulot:</b> {order.product.name} ({order.requested_weight} kg)\n"
                f"📍 <b>Manzil:</b> {order.delivery_address or 'Belgilanmagan'}\n\n"
                f"<i>Quyidagi 1-bosishda Navigator tugmalari orqali marshrutni ochishingiz mumkin.</i>"
            )
            safe_edit_message_or_caption(chat_id, message_id, caption, reply_markup=next_kb)
            answer_callback(callback_query_id, "🚚 Kuryerga berildi! Navigator tayyor.")
        except Exception as e:
            send_message(chat_id, f"⚠️ Xato: {e}")

    elif callback_data.startswith('adm_comp_'):
        order_id = int(callback_data.replace('adm_comp_', ''))
        try:
            order = B2BOrder.objects.get(id=order_id)
            order.status = 'completed'
            order.save()
            
            CustomerLog.objects.create(
                customer=order.customer,
                log_type='bonus',
                title="Do'kon xabari",
                message=f"🎉 Buyurtmangiz (#{order.id}) muvaffaqiyatli yetkazildi! Xaridingiz barakali bo'lsin!",
                details={'source': 'telegram_admin', 'order_id': order.id, 'new_status': 'completed'},
                amount=Decimal('0.00')
            )
            notify_customer_order_status(order, "🎉 Yetkazildi (Yakunlandi)")
            caption = (
                f"🎉 <b>BUYURTMA #{order.id} YAKUNLANDI (TOPSHIRILDI)!</b>\n"
                f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                f"👤 <b>Mijoz:</b> {order.customer.first_name} ({order.customer.phone})\n"
                f"<i>Buyurtma muvaffaqiyatli topshirildi va yopildi.</i>"
            )
            safe_edit_message_or_caption(chat_id, message_id, caption, reply_markup={'inline_keyboard': []})
            answer_callback(callback_query_id, "🎉 Buyurtma yakunlandi!")
        except Exception as e:
            send_message(chat_id, f"⚠️ Xato: {e}")

    elif callback_data.startswith('adm_rej_'):
        order_id = int(callback_data.replace('adm_rej_', ''))
        try:
            order = B2BOrder.objects.get(id=order_id)
            order.status = 'rejected'
            order.save()
            
            CustomerLog.objects.create(
                customer=order.customer,
                log_type='bonus',
                title="Do'kon xabari",
                message=f"❌ Buyurtmangiz (#{order.id}) do'kon tomonidan rad etildi.",
                details={'source': 'telegram_admin', 'order_id': order.id, 'new_status': 'rejected'},
                amount=Decimal('0.00')
            )
            notify_customer_order_status(order, "❌ Rad etildi (To'lov cheki mos kelmadi)")
            caption = (
                f"❌ <b>BUYURTMA #{order.id} RAD ETILDI.</b>\n"
                f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                f"👤 <b>Mijoz:</b> {order.customer.first_name} ({order.customer.phone})"
            )
            safe_edit_message_or_caption(chat_id, message_id, caption, reply_markup={'inline_keyboard': []})
            answer_callback(callback_query_id, "❌ Buyurtma rad etildi.")
        except Exception as e:
            send_message(chat_id, f"⚠️ Xato: {e}")

    elif callback_data.startswith('adm_proof_appr_'):
        proof_id = int(callback_data.replace('adm_proof_appr_', ''))
        handle_proof_approval(chat_id, message_id, proof_id, callback_query_id)

    elif callback_data.startswith('adm_proof_rej_'):
        proof_id = int(callback_data.replace('adm_proof_rej_', ''))
        handle_proof_rejection(chat_id, message_id, proof_id, callback_query_id)

    elif callback_data == 'cmd_menu':
        handle_start_edit(chat_id, message_id)
    elif callback_data == 'cmd_hisobot':
        handle_hisobot(chat_id, message_id=message_id)
    elif callback_data == 'cmd_z_report':
        send_daily_executive_digest(chat_id=chat_id)
    elif callback_data == 'cmd_decay':
        handle_decay(chat_id, message_id=message_id)
    elif callback_data == 'cmd_qarz':

        handle_qarz(chat_id, message_id=message_id)
    elif callback_data == 'cmd_zaxira':
        handle_zaxira(chat_id, message_id=message_id)
    elif callback_data == 'cmd_taminotchi':
        handle_taminotchi(chat_id, message_id=message_id)
    elif callback_data == 'cmd_shift':
        handle_shift(chat_id, message_id=message_id)
    elif callback_data == 'cmd_bonus':
        handle_bonus(chat_id, message_id=message_id)
    elif callback_data == 'cmd_kecha':
        yesterday = timezone.localdate() - timedelta(days=1)
        handle_hisobot(chat_id, target_date=yesterday, message_id=message_id)


def notify_customer_order_status(order, new_status_text):
    """Mijozga buyurtma statusi haqida bildirishnoma yuborish."""
    if not order.customer or not order.customer.telegram_chat_id:
        return

    chat_id = order.customer.telegram_chat_id
    try:
        from .customer_bot import send_message as send_customer_msg, main_menu_reply_keyboard, render_telegram_stepper_bar
        stepper_bar = render_telegram_stepper_bar(order.status)
        msg = (
            f"🔔 *BUYURTMANINGIZ STATUSI O'ZGARDI!* (Buyurtma #{order.id})\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"📦 *Mahsulot:* {order.product.name} ({order.requested_weight} kg)\n\n"
            f"{stepper_bar}\n\n"
            f"📊 *Yangi status:* *{new_status_text}*\n\n"
            f"📌 Real vaqt rejimida kuzatish uchun *'📌 Buyurtma Statusi (Live)'* tugmasini bosing."
        )
        send_customer_msg(chat_id, msg, reply_markup=main_menu_reply_keyboard())
    except Exception as e:
        print(f"[Notify Customer Error]: {e}")


def handle_start_edit(chat_id, message_id):
    """Start menyusini tahrirlash (tugma orqali qaytganda)."""
    text = (
        "🥩 *Baxmal Meat Kassa Bot*\n\n"
        "Quyidagi tugmalardan birini tanlang:"
    )
    edit_message(chat_id, message_id, text, reply_markup=main_menu_keyboard())


def send_daily_executive_digest(chat_id=None, target_date=None):
    """Do'kon egasi va rahbariyat uchun kechki avtomatik Z-Hisobot (Executive Digest)."""
    from pos.models import Sale, SaleItem, CashierShift, CashTransaction, Stock, Slaughter, StoreSetting
    
    if target_date is None:
        target_date = timezone.localdate()
    if not chat_id:
        chat_id = CHAT_ID
    if not chat_id:
        print("[Daily Digest Error]: TELEGRAM_CHAT_ID is not configured.")
        return False

    date_str = target_date.strftime('%d.%m.%Y')
    sales = Sale.objects.filter(created_at__date=target_date)
    total_sales_count = sales.count()
    
    total_revenue = sales.aggregate(s=Sum('total_amount'))['s'] or Decimal('0.00')
    naqd_total = sales.filter(payment_method='naqd').aggregate(s=Sum('final_paid'))['s'] or Decimal('0.00')
    karta_total = sales.filter(payment_method='karta').aggregate(s=Sum('final_paid'))['s'] or Decimal('0.00')
    qr_total = sales.filter(payment_method='qr').aggregate(s=Sum('final_paid'))['s'] or Decimal('0.00')
    nasiya_total = sales.filter(payment_method='nasiya').aggregate(s=Sum('debt_added'))['s'] or Decimal('0.00')
    bonus_used_total = sales.aggregate(s=Sum('bonus_used'))['s'] or Decimal('0.00')
    discount_total = sales.aggregate(s=Sum('discount_amount'))['s'] or Decimal('0.00')

    # Sold meat breakdown
    sale_items = SaleItem.objects.filter(sale__created_at__date=target_date)
    total_weight_sold = sale_items.aggregate(w=Sum('weight'))['w'] or Decimal('0.000')
    
    top_products = sale_items.values('product__name').annotate(
        weight=Sum('weight'),
        total=Sum('item_total')
    ).order_by('-total')[:5]

    product_lines = ""
    for idx, p in enumerate(top_products, 1):
        product_lines += f"  {idx}. *{p['product__name']}*: `{p['weight']:.1f} kg` — `{int(p['total']):,}` so'm\n"

    # Cash Transactions (Kirim / Chiqim)
    cash_ins = CashTransaction.objects.filter(created_at__date=target_date, transaction_type='in').aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
    cash_outs = CashTransaction.objects.filter(created_at__date=target_date, transaction_type='out').aggregate(s=Sum('amount'))['s'] or Decimal('0.00')

    # Shifts today
    shifts = CashierShift.objects.filter(opened_at__date=target_date)
    shift_status_lines = ""
    for sh in shifts:
        status_icon = "🟢 Ochiq" if sh.is_open else "🔴 Yopilgan"
        diff_str = f"(Farq: `{sh.cash_difference:,.0f}` so'm)" if not sh.is_open else ""
        shift_status_lines += f"  • Shift #{sh.id} ({sh.cashier.username}): {status_icon} {diff_str}\n"

    # Low stock alerts (< 5 kg)
    low_stocks = Stock.objects.filter(product__is_active=True, quantity__lte=Decimal('5.000')).select_related('product')
    low_stock_lines = ""
    for ls in low_stocks[:4]:
        low_stock_lines += f"  ⚠️ *{ls.product.name}*: faqat `{ls.quantity:.3f} kg` qoldi!\n"

    # Slaughters today
    slaughters_today = Slaughter.objects.filter(created_at__date=target_date)
    slaughter_count = slaughters_today.count()
    slaughter_weight = slaughters_today.aggregate(w=Sum('total_weight'))['w'] or Decimal('0.000')

    # Total stock weight
    total_stock_weight = Stock.objects.filter(product__is_active=True).aggregate(s=Sum('quantity'))['s'] or Decimal('0.000')

    store = StoreSetting.objects.filter(is_active=True).first()
    store_title = store.name.upper() if store else "BAXMAL MEAT"

    msg = (
        f"🌙 *{store_title} — KUNLIK Z-HISOBOT (21:00)*\n"
        f"📅 *Sana:* `{date_str}`\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"💰 *Jami tushum:* `{int(total_revenue):,}` so'm (Naqd: `{int(naqd_total):,}`, Karta: `{int(karta_total):,}`, Nasiya: `{int(nasiya_total):,}`)\n"
        f"🥩 *Sotilgan go'sht:* `{total_weight_sold:.1f} kg`\n"
        f"👥 *Xaridorlar soni:* `{total_sales_count} ta`\n"
        f"📉 *Ombordagi qoldiq:* `{total_stock_weight:.1f} kg`\n\n"
        f"📊 *To'lovlar Taqsimoti:*\n"
        f"  💵 *Naqd:* `{int(naqd_total):,}` so'm\n"
        f"  💳 *Plastik Karta:* `{int(karta_total):,}` so'm\n"
        f"  📱 *QR To'lov:* `{int(qr_total):,}` so'm\n"
        f"  📋 *Nasiya (Qarz):* `{int(nasiya_total):,}` so'm\n"
        f"  💎 *Bonus sarflandi:* `{int(bonus_used_total):,}` so'm\n"
        f"  🏷️ *Chegirmalar:* `{int(discount_total):,}` so'm\n\n"
        f"💸 *Kassa Kirim / Chiqim:*\n"
        f"  📥 Qo'shimcha Kirim: `+{int(cash_ins):,}` so'm\n"
        f"  📤 Chiqim / Xarajat: `-{int(cash_outs):,}` so'm\n\n"
        f"🏆 *Eng ko'p sotilgan go'shtlar:*\n"
        f"{product_lines or '  Savdo bo`lmadi'}\n"
    )


    if slaughter_count > 0:
        msg += f"🐂 *Bugungi so'yim:* `{slaughter_count} bosh` ({slaughter_weight:.1f} kg toza go'sht)\n\n"

    if shift_status_lines:
        msg += f"🏪 *Kassa Smenalari:*\n{shift_status_lines}\n"

    if low_stock_lines:
        msg += f"📦 *Kam qolgan zaxiralar:*\n{low_stock_lines}\n"

    # Vitrinada 2+ kun turgan partiyalar bo'yicha AI tavsiya
    try:
        from .models import StockBatch
        aging_batches = StockBatch.objects.filter(current_quantity__gt=Decimal('0.05')).select_related('product')
        ai_decay_lines = ""
        for ab in aging_batches:
            ai_rec = ab.get_ai_recommendation()
            if ai_rec:
                ai_decay_lines += f"  ⚠️ *{ai_rec['product_name']}* (#{ai_rec['batch_id']}): vitrinada `{ai_rec['days']} kun` turibdi (`{ai_rec['quantity']:.1f} kg`). _Qiymaga aylantirish yoki tezroq sotish tavsiya etiladi!_\n"
        if ai_decay_lines:
            msg += f"🧠 *Vitrina & Qurish Zarari — AI Tavsiya:*\n{ai_decay_lines}\n"
    except Exception as e:
        print(f"[AI Decay Digest Error]: {e}")

    msg += (
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"✨ *MeatFlow Pro Executive Intelligence*"
    )


    site_url = os.environ.get('SITE_URL', 'https://baxmalmeat.uz')
    reply_markup = {
        'inline_keyboard': [
            [
                {'text': '📊 Global Analitika (Veb)', 'url': f"{site_url}/pos/global-report/"},
                {'text': '👥 Qarzdorlar', 'callback_data': 'cmd_qarz'}
            ],
            [
                {'text': '🔄 Qayta Yangilash', 'callback_data': 'cmd_z_report'}
            ]
        ]
    }
    return send_message(chat_id, msg, reply_markup=reply_markup)

